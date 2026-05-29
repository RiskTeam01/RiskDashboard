import hashlib
import html
import json
import os
import platform
import re
import secrets
import shutil
import threading
import time
import traceback
import uuid
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import (
    FileResponse,
    HTMLResponse,
    JSONResponse,
    RedirectResponse,
)
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook


# ============================================================
# PHILLIPCAPITAL RISK MANAGEMENT
# CREDIT WORKSHEET PROCESSOR
# ------------------------------------------------------------
# FastAPI browser app:
#   - Login gate (all pages require sign-in)
#   - Settings page: manage logins, cleanup, upload limits
#   - Home: upload one PDF or batch upload multiple PDFs
#   - Live batch progress screen (per-PDF status)
#   - Generates one Excel workbook per PDF + an audit report
#   - If one PDF: downloads one .xlsx
#   - If multiple PDFs: downloads one .zip containing all .xlsx outputs
#   - Console: combined field results + raw helper codes + logs
#   - Outputs: clean workbook/audit download cards
#   - Auto-detects Excel template from /templates
#   - Uses /assets/phillipcapital_logo.png for logo and favicon
#
# RUN:
#   python -m uvicorn server:app --reload --host 0.0.0.0 --port 8000
# ============================================================


APP_ROOT = Path(__file__).resolve().parent

TEMPLATE_DIR = APP_ROOT / "templates"
UPLOAD_DIR = APP_ROOT / "uploads"
OUTPUT_DIR = APP_ROOT / "outputs"
LOG_DIR = APP_ROOT / "logs"
ASSETS_DIR = APP_ROOT / "assets"
AUDIT_DIR = APP_ROOT / "audits"

USERS_FILE = APP_ROOT / "users.json"
CONFIG_FILE = APP_ROOT / "app_config.json"

SHEET_NAME = "Sheet1"
PREFERRED_TEMPLATE_FILENAME = "Buckler Excel Credit WS Template.xlsx"

SESSION_COOKIE = "pc_session"

for folder in [TEMPLATE_DIR, UPLOAD_DIR, OUTPUT_DIR, LOG_DIR, ASSETS_DIR, AUDIT_DIR]:
    folder.mkdir(parents=True, exist_ok=True)


# ============================================================
# APP CONFIG (editable in Settings)
# ============================================================

DEFAULT_CONFIG = {
    "max_pdf_size_mb": 50,
    "max_batch_size": 20,
    "session_timeout_minutes": 60,
    "cleanup_uploads_days": 7,
    "cleanup_outputs_days": 30,
    "cleanup_logs_days": 30,
}


def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            merged = dict(DEFAULT_CONFIG)
            merged.update({k: data[k] for k in data if k in DEFAULT_CONFIG})
            return merged
        except Exception:
            pass
    return dict(DEFAULT_CONFIG)


def save_config(config: dict):
    clean = dict(DEFAULT_CONFIG)
    for key in DEFAULT_CONFIG:
        if key in config:
            clean[key] = config[key]
    CONFIG_FILE.write_text(json.dumps(clean, indent=2), encoding="utf-8")


# ============================================================
# AUTH: USERS + PASSWORD HASHING
# ============================================================

# Preset logins (created on first run). Passwords are stored hashed.
PRESET_USERS = [
    {"username": "kpage@phillipcapital.com", "password": "Welcome2"},
    {"username": "haotian@phillipcapital.com", "password": "Welcome3"},
    {"username": "hcurtis@phillipcapital.com", "password": "Welcome 1"},
]


def hash_password(password: str, salt: Optional[str] = None) -> tuple[str, str]:
    if salt is None:
        salt = secrets.token_hex(16)
    derived = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), bytes.fromhex(salt), 100_000
    )
    return salt, derived.hex()


def verify_password(password: str, salt: str, hashed: str) -> bool:
    try:
        _, candidate = hash_password(password, salt)
        return secrets.compare_digest(candidate, hashed)
    except Exception:
        return False


def load_users() -> dict:
    if USERS_FILE.exists():
        try:
            return json.loads(USERS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_users(users: dict):
    USERS_FILE.write_text(json.dumps(users, indent=2), encoding="utf-8")


def seed_preset_users():
    users = load_users()

    if users:
        return

    for preset in PRESET_USERS:
        salt, hashed = hash_password(preset["password"])
        users[preset["username"].lower()] = {
            "username": preset["username"],
            "salt": salt,
            "hash": hashed,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "created_by": "system",
        }

    save_users(users)


def add_user(username: str, password: str, created_by: str) -> tuple[bool, str]:
    username = username.strip()
    if not username:
        return False, "Username cannot be empty."
    if not password:
        return False, "Password cannot be empty."

    users = load_users()
    key = username.lower()

    if key in users:
        return False, f"User '{username}' already exists."

    salt, hashed = hash_password(password)
    users[key] = {
        "username": username,
        "salt": salt,
        "hash": hashed,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "created_by": created_by,
    }
    save_users(users)
    return True, f"User '{username}' added."


def delete_user(username: str, current_user: str) -> tuple[bool, str]:
    users = load_users()
    key = username.strip().lower()

    if key not in users:
        return False, "User not found."

    if key == current_user.strip().lower():
        return False, "You cannot delete the account you are currently signed in with."

    if len(users) <= 1:
        return False, "Cannot delete the last remaining user."

    removed = users.pop(key)
    save_users(users)
    return True, f"User '{removed.get('username', username)}' removed."


def update_password(username: str, new_password: str, actor: str) -> tuple[bool, str]:
    if not new_password:
        return False, "Password cannot be empty."

    users = load_users()
    key = username.strip().lower()

    if key not in users:
        return False, "User not found."

    record = users[key]
    salt, hashed = hash_password(new_password)
    record["salt"] = salt
    record["hash"] = hashed
    record["password_updated_at"] = datetime.now().isoformat(timespec="seconds")
    record["password_updated_by"] = actor
    users[key] = record
    save_users(users)

    # If an admin updates someone else's password, drop that user's other sessions
    # so the old password cannot continue being used via a stale cookie.
    actor_key = actor.strip().lower()
    if key != actor_key:
        stale_tokens = [
            tok for tok, sess in SESSIONS.items()
            if sess.get("username", "").strip().lower() == key
        ]
        for tok in stale_tokens:
            SESSIONS.pop(tok, None)

    return True, f"Password updated for '{record.get('username', username)}'."


def authenticate(username: str, password: str) -> Optional[str]:
    users = load_users()
    key = username.strip().lower()
    record = users.get(key)

    if not record:
        return None

    if verify_password(password, record.get("salt", ""), record.get("hash", "")):
        return record.get("username", username)

    return None


# ============================================================
# SESSIONS
# ============================================================

SESSIONS: dict[str, dict] = {}


def create_session(username: str) -> str:
    token = secrets.token_urlsafe(32)
    now = time.time()
    SESSIONS[token] = {
        "username": username,
        "created": now,
        "last_active": now,
    }
    return token


def validate_session(token: Optional[str]) -> Optional[str]:
    if not token:
        return None

    session = SESSIONS.get(token)
    if not session:
        return None

    timeout_seconds = load_config()["session_timeout_minutes"] * 60
    if time.time() - session["last_active"] > timeout_seconds:
        SESSIONS.pop(token, None)
        return None

    session["last_active"] = time.time()
    return session["username"]


def destroy_session(token: Optional[str]):
    if token:
        SESSIONS.pop(token, None)


def current_user_or_none(request: Request) -> Optional[str]:
    token = request.cookies.get(SESSION_COOKIE)
    return validate_session(token)


def require_api_user(request: Request) -> str:
    user = current_user_or_none(request)
    if not user:
        raise HTTPException(status_code=401, detail="Login required.")
    return user


seed_preset_users()


# ============================================================
# FIELD MAP
# ============================================================

FIELD_DEFINITIONS = [
    {"expression": "940", "label": "Total Assets", "excel_cell": "B4"},
    {"expression": "1800", "label": "Total Equity", "excel_cell": "B6"},
    {"expression": "750", "label": "Cash and cash equivalents", "excel_cell": "B8"},
    {"expression": "760", "label": "Cash segregated under federal", "excel_cell": "B10"},
    {"expression": "770", "label": "Fails to Deliver", "excel_cell": "B12"},
    {"expression": "780", "label": "Stocks Borrowed", "excel_cell": "B14"},
    {"expression": "800", "label": "Clearing Org receivables", "excel_cell": "B16"},
    {"expression": "810", "label": "Others", "excel_cell": "B18"},
    {"expression": "820", "label": "Customer Receivables", "excel_cell": "B20"},
    {"expression": "840", "label": "Reverse Repos", "excel_cell": "B22"},
    {"expression": "292", "label": "Trade Date Receivable", "excel_cell": "B24"},
    {"expression": "12019", "label": "Marketable securities", "excel_cell": "B26"},
    {"expression": "740", "label": "Non-allowable assets", "excel_cell": "B28"},
    {"expression": "890", "label": "Secured Demand Notes", "excel_cell": "B32"},
    {"expression": "1760", "label": "Total Liabilities", "excel_cell": "B34"},

    {"expression": "1490+1500", "label": "Fails to Receive", "excel_cell": "D12"},
    {"expression": "1510+1520", "label": "Stocks Loaned", "excel_cell": "D14"},
    {"expression": "1550+1560", "label": "Clearing Org payables", "excel_cell": "D16"},
    {"expression": "1570", "label": "Other", "excel_cell": "D18"},
    {"expression": "1580+1590", "label": "Customer payables", "excel_cell": "D20"},
    {"expression": "1480", "label": "Repos", "excel_cell": "D22"},
    {"expression": "1686", "label": "Obligation to rtn Securities Collateral", "excel_cell": "D26"},
    {"expression": "1480", "label": "All other liabilities", "excel_cell": "D30"},
    {"expression": "1730", "label": "Securities borrowings", "excel_cell": "D32"},
]


AMOUNT_PATTERN = re.compile(
    r"""
    ^\$?
    \(?
    -?
    (?:
        \d{1,3}(?:,\d{3})+ |
        \d+
    )
    (?:\.\d+)?
    \)?
    $
    """,
    re.VERBOSE,
)


JOB_RESULTS: dict[str, dict] = {}
LATEST_JOB_ID: Optional[str] = None

# Live batch progress jobs (in-memory)
BATCH_JOBS: dict[str, dict] = {}
BATCH_JOBS_LOCK = threading.Lock()


# ============================================================
# DATA MODELS
# ============================================================

@dataclass
class WordItem:
    page_index: int
    page_number: int
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    block_no: int
    line_no: int
    word_no: int

    @property
    def y_center(self) -> float:
        return (self.y0 + self.y1) / 2


@dataclass
class CodeOccurrence:
    code: str
    page_number: int
    x0: float
    y0: float
    x1: float
    y1: float
    nearby_amount_text: str = ""
    nearby_context: str = ""
    note: str = ""
    confidence_score: int = 0
    selected: bool = False

    @property
    def location_text(self) -> str:
        return f"Page {self.page_number}, x={round(self.x0, 2)}, y={round(self.y0, 2)}"


@dataclass
class FieldSpec:
    expression: str
    label: str
    excel_cell: str
    codes: list[str] = field(default_factory=list)


@dataclass
class FieldResult:
    expression: str
    label: str
    excel_cell: str
    display_value: str
    numeric_value: Optional[float]
    should_write_blank: bool
    status: str
    difficulty: str
    notes: str
    component_details: str


# ============================================================
# TEMPLATE / ASSET HELPERS
# ============================================================

def detect_backend_template() -> Path:
    preferred = TEMPLATE_DIR / PREFERRED_TEMPLATE_FILENAME

    if preferred.exists() and not preferred.name.startswith("~$"):
        return preferred

    candidates = [
        path for path in TEMPLATE_DIR.glob("*.xlsx")
        if not path.name.startswith("~$")
    ]

    if not candidates:
        raise FileNotFoundError(
            f"No Excel template found in: {TEMPLATE_DIR}. "
            "Place your blank .xlsx template inside the templates folder."
        )

    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def get_template_status() -> dict:
    try:
        template_path = detect_backend_template()
        return {
            "exists": True,
            "filename": template_path.name,
            "folder": str(TEMPLATE_DIR),
            "path": str(template_path),
            "error": "",
        }
    except Exception as e:
        return {
            "exists": False,
            "filename": "",
            "folder": str(TEMPLATE_DIR),
            "path": "",
            "error": str(e),
        }


def get_logo_path() -> Optional[Path]:
    preferred = ASSETS_DIR / "phillipcapital_logo.png"

    if preferred.exists():
        return preferred

    candidates = [
        path for path in ASSETS_DIR.glob("*.png")
        if not path.name.startswith("~$")
    ]

    if candidates:
        candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return candidates[0]

    return None


def get_logo_url() -> str:
    logo_path = get_logo_path()
    if logo_path:
        return f"/assets/{logo_path.name}"
    return ""


# ============================================================
# ENGINE
# ============================================================

class CreditWorksheetEngine:
    def __init__(self):
        self.debug_lines: list[str] = []

    def log(self, message: str):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.debug_lines.append(f"[{timestamp}] {message}")

    def save_debug_log(self, job_id: str, suffix: str = "") -> Path:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        safe_suffix = f"_{suffix}" if suffix else ""
        log_path = LOG_DIR / f"credit_ws_debug_{job_id}{safe_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        log_path.write_text("\n".join(self.debug_lines), encoding="utf-8")
        return log_path

    def get_field_specs(self) -> list[FieldSpec]:
        specs = []

        for item in FIELD_DEFINITIONS:
            expression = item["expression"].replace(" ", "")
            codes = expression.split("+")
            specs.append(
                FieldSpec(
                    expression=expression,
                    label=item["label"],
                    excel_cell=item["excel_cell"],
                    codes=codes,
                )
            )

        return specs

    def get_requested_codes(self, field_specs: list[FieldSpec]) -> list[str]:
        return sorted(
            set(code for spec in field_specs for code in spec.codes),
            key=lambda value: int(value),
        )

    def extract_pdf_field_results(self, pdf_path: Path):
        try:
            import fitz
        except ImportError:
            raise RuntimeError("PyMuPDF is not installed. Run: python -m pip install pymupdf")

        field_specs = self.get_field_specs()
        requested_codes = self.get_requested_codes(field_specs)

        self.log("=" * 100)
        self.log("Starting PDF extraction")
        self.log(f"PDF: {pdf_path}")
        self.log(f"Field definitions: {len(field_specs)}")
        self.log(f"Requested helper codes: {', '.join(requested_codes)}")
        self.log("=" * 100)

        doc = fitz.open(pdf_path)

        all_words: list[WordItem] = []
        page_count = len(doc)
        total_text_chars = 0
        pages_with_text = 0
        total_images = 0

        self.log(f"[PDF] Page count: {page_count}")

        for page_index in range(page_count):
            page = doc[page_index]
            text = page.get_text("text") or ""
            words_raw = page.get_text("words") or []
            images = page.get_images(full=True) or []

            text_chars = len(text.strip())
            total_text_chars += text_chars
            total_images += len(images)

            if text.strip():
                pages_with_text += 1

            self.log(
                f"[PAGE {page_index + 1}] "
                f"text chars={text_chars}, "
                f"words={len(words_raw)}, "
                f"images={len(images)}, "
                f"size={round(page.rect.width, 2)}x{round(page.rect.height, 2)}"
            )

            if page_index == 0:
                preview = text.strip().replace("\n", " ")
                preview = preview[:900] + ("..." if len(preview) > 900 else "")
                self.log(f"[PAGE 1 TEXT PREVIEW] {preview}")

            for word in words_raw:
                try:
                    x0, y0, x1, y1, word_text, block_no, line_no, word_no = word
                    clean_text = str(word_text).strip()

                    if clean_text:
                        all_words.append(
                            WordItem(
                                page_index=page_index,
                                page_number=page_index + 1,
                                x0=x0,
                                y0=y0,
                                x1=x1,
                                y1=y1,
                                text=clean_text,
                                block_no=block_no,
                                line_no=line_no,
                                word_no=word_no,
                            )
                        )
                except Exception as e:
                    self.log(f"[WARN] Failed parsing word on page {page_index + 1}: {e}")

        doc.close()

        readability = self.calculate_readability(
            page_count=page_count,
            pages_with_text=pages_with_text,
            total_text_chars=total_text_chars,
            total_words=len(all_words),
            total_images=total_images,
        )

        readability_info = {
            "page_count": page_count,
            "pages_with_text": pages_with_text,
            "total_text_chars": total_text_chars,
            "total_words": len(all_words),
            "total_images": total_images,
            "readability": readability,
        }

        self.log("-" * 100)
        self.log("[READABILITY SUMMARY]")
        self.log(f"Pages with selectable text: {pages_with_text}/{page_count}")
        self.log(f"Total extracted text characters: {total_text_chars}")
        self.log(f"Total extracted word items: {len(all_words)}")
        self.log(f"Total embedded images detected: {total_images}")
        self.log(f"Readability result: {readability}")
        self.log("-" * 100)

        occurrences_by_code = self.find_requested_code_occurrences(
            all_words=all_words,
            requested_codes=requested_codes,
        )

        self.select_best_occurrences(occurrences_by_code)

        for code in requested_codes:
            occurrences = occurrences_by_code.get(code, [])
            self.log("")
            self.log(f"[CODE {code}] occurrences found: {len(occurrences)}")

            if not occurrences:
                self.log("  - Missing from selectable text.")
                continue

            for occ in occurrences:
                selected_text = "YES" if occ.selected else "NO"
                self.log(
                    f"  - selected={selected_text}, "
                    f"page={occ.page_number}, "
                    f"x={round(occ.x0, 2)}-{round(occ.x1, 2)}, "
                    f"y={round(occ.y0, 2)}-{round(occ.y1, 2)}, "
                    f"amount='{occ.nearby_amount_text}', "
                    f"score={occ.confidence_score}, "
                    f"note='{occ.note}', "
                    f"context='{occ.nearby_context}'"
                )

        field_results: list[FieldResult] = []

        for spec in field_specs:
            result = self.evaluate_field(spec, occurrences_by_code)
            field_results.append(result)

            self.log("")
            self.log(f"[FIELD] {spec.expression} = {spec.label} -> {spec.excel_cell}")
            self.log(f"  Preview value: {result.display_value}")
            self.log(f"  Status: {result.status}")
            self.log(f"  Difficulty: {result.difficulty}")
            self.log(f"  Notes: {result.notes}")
            self.log(f"  Details: {result.component_details}")

        summary = self.build_overall_summary(
            field_results=field_results,
            occurrences_by_code=occurrences_by_code,
            requested_codes=requested_codes,
            readability_info=readability_info,
        )

        self.log("")
        self.log("=" * 100)
        self.log("[OVERALL PDF AUTOMATION READINESS]")
        for line in summary:
            self.log(line)
        self.log("=" * 100)

        return field_results, occurrences_by_code, readability_info, summary

    def calculate_readability(
        self,
        page_count: int,
        pages_with_text: int,
        total_text_chars: int,
        total_words: int,
        total_images: int,
    ) -> str:
        if total_words == 0 and total_text_chars == 0:
            if total_images > 0:
                return "HARD - likely scanned/image PDF; OCR would be needed"
            return "HARD - no selectable text detected"

        avg_words_per_page = total_words / max(page_count, 1)

        if pages_with_text == page_count and avg_words_per_page >= 30:
            return "EASY - selectable text detected on all pages"

        if pages_with_text > 0 and total_words >= 50:
            return "MEDIUM - some selectable text detected; layout needs testing"

        return "HARD - very limited selectable text detected"

    def normalize_code_token(self, text: str) -> str:
        cleaned = str(text).strip()
        cleaned = cleaned.replace("[", "").replace("]", "")
        cleaned = cleaned.replace("(", "").replace(")", "")
        cleaned = cleaned.replace("{", "").replace("}", "")
        cleaned = cleaned.replace(".", "")
        cleaned = cleaned.strip()
        return cleaned

    def is_amount_like(self, text: str) -> bool:
        cleaned = str(text).strip()

        if not cleaned:
            return False

        possible_code = (
            cleaned.replace("$", "")
            .replace("-", "")
            .replace("(", "")
            .replace(")", "")
            .strip()
        )

        if "," not in cleaned and possible_code.isdigit() and len(possible_code) <= 5:
            return False

        return bool(AMOUNT_PATTERN.match(cleaned))

    def amount_to_number(self, text: str):
        if not text:
            return None

        cleaned = str(text).strip()
        negative = False

        if cleaned.startswith("(") and cleaned.endswith(")"):
            negative = True

        cleaned = cleaned.replace("$", "")
        cleaned = cleaned.replace(",", "")
        cleaned = cleaned.replace("(", "")
        cleaned = cleaned.replace(")", "")
        cleaned = cleaned.strip()

        try:
            if "." in cleaned:
                value = float(cleaned)
            else:
                value = int(cleaned)

            if negative:
                value *= -1

            return value

        except Exception:
            return None

    def format_number_for_display(self, value: Optional[float]) -> str:
        if value is None:
            return ""

        if abs(value - int(value)) < 0.000001:
            return f"{int(value):,}"

        return f"{value:,.2f}"

    def find_requested_code_occurrences(
        self,
        all_words: list[WordItem],
        requested_codes: list[str],
    ) -> dict[str, list[CodeOccurrence]]:
        requested_set = set(requested_codes)
        occurrences_by_code = {code: [] for code in requested_codes}

        for word in all_words:
            normalized = self.normalize_code_token(word.text)

            if normalized not in requested_set:
                continue

            nearby_amount, nearby_context, note, score = self.find_left_side_amount_preview(
                target_word=word,
                all_words=all_words,
            )

            occurrence = CodeOccurrence(
                code=normalized,
                page_number=word.page_number,
                x0=word.x0,
                y0=word.y0,
                x1=word.x1,
                y1=word.y1,
                nearby_amount_text=nearby_amount,
                nearby_context=nearby_context,
                note=note,
                confidence_score=score,
            )

            occurrences_by_code[normalized].append(occurrence)

        return occurrences_by_code

    def find_left_side_amount_preview(
        self,
        target_word: WordItem,
        all_words: list[WordItem],
    ):
        y_tolerance = 5.75
        max_left_distance = 260

        same_page_words = [w for w in all_words if w.page_index == target_word.page_index]

        same_row_left_words = [
            w for w in same_page_words
            if w.x1 < target_word.x0
            and abs(w.y_center - target_word.y_center) <= y_tolerance
            and (target_word.x0 - w.x1) <= max_left_distance
        ]

        same_row_left_words.sort(key=lambda w: w.x1, reverse=True)

        nearby_context = " ".join(w.text for w in same_row_left_words[:12])

        for candidate in same_row_left_words:
            candidate_text = candidate.text.strip()

            if self.is_amount_like(candidate_text):
                score = 80

                distance = target_word.x0 - candidate.x1
                if distance <= 80:
                    score += 10
                elif distance <= 150:
                    score += 5

                if "," in candidate_text:
                    score += 10

                return (
                    candidate_text,
                    nearby_context,
                    "Nearby amount-like value found to the left",
                    score,
                )

        if nearby_context:
            return (
                "",
                nearby_context,
                "Code found, but no amount-like value was found on the same row. Treat as valid blank unless later rules say otherwise.",
                55,
            )

        return (
            "",
            "",
            "Code found, but left-side row area appears blank or not selectable. Treat as valid blank unless later rules say otherwise.",
            50,
        )

    def select_best_occurrences(self, occurrences_by_code: dict[str, list[CodeOccurrence]]):
        for code, occurrences in occurrences_by_code.items():
            if not occurrences:
                continue

            with_amount = [occ for occ in occurrences if occ.nearby_amount_text]

            if len(occurrences) == 1:
                occurrences[0].selected = True
                continue

            if len(with_amount) == 1:
                with_amount[0].selected = True
                with_amount[0].note += " | Auto-selected because it is the only occurrence with a nearby amount."
                continue

            if len(with_amount) > 1:
                with_amount.sort(key=lambda occ: occ.confidence_score, reverse=True)
                with_amount[0].selected = True
                with_amount[0].note += " | Auto-selected by highest score, but multiple amount matches exist. Needs review."
                continue

            occurrences.sort(key=lambda occ: occ.confidence_score, reverse=True)
            occurrences[0].selected = True
            occurrences[0].note += " | Auto-selected blank candidate, but multiple blank/code-reference matches exist. Needs review."

    def evaluate_field(
        self,
        spec: FieldSpec,
        occurrences_by_code: dict[str, list[CodeOccurrence]],
    ) -> FieldResult:
        component_values = []
        component_notes = []
        missing_codes = []
        review_needed = False
        all_components_blank_or_found = True

        for code in spec.codes:
            occurrences = occurrences_by_code.get(code, [])

            if not occurrences:
                missing_codes.append(code)
                all_components_blank_or_found = False
                component_notes.append(f"{code}: MISSING")
                continue

            selected = next((occ for occ in occurrences if occ.selected), occurrences[0])

            if len(occurrences) > 1:
                review_needed = True

            if selected.nearby_amount_text:
                numeric_value = self.amount_to_number(selected.nearby_amount_text)

                if numeric_value is None:
                    review_needed = True
                    component_notes.append(f"{code}: found '{selected.nearby_amount_text}' but could not convert")
                else:
                    component_values.append(numeric_value)
                    component_notes.append(f"{code}: {selected.nearby_amount_text} ({selected.location_text})")
            else:
                component_notes.append(f"{code}: blank ({selected.location_text})")

        if missing_codes:
            return FieldResult(
                expression=spec.expression,
                label=spec.label,
                excel_cell=spec.excel_cell,
                display_value="",
                numeric_value=None,
                should_write_blank=True,
                status="MISSING",
                difficulty="Hard",
                notes=f"Missing code(s): {', '.join(missing_codes)}",
                component_details=" | ".join(component_notes),
            )

        if len(spec.codes) == 1:
            code = spec.codes[0]
            occurrences = occurrences_by_code.get(code, [])
            selected = next((occ for occ in occurrences if occ.selected), occurrences[0])

            if selected.nearby_amount_text:
                numeric_value = self.amount_to_number(selected.nearby_amount_text)

                status = "READY"
                difficulty = "Easy" if len(occurrences) == 1 else "Medium"
                notes = "Single code found with value."

                if len(occurrences) > 1:
                    notes += " Multiple occurrences exist, so final extraction should use position rules."

                return FieldResult(
                    expression=spec.expression,
                    label=spec.label,
                    excel_cell=spec.excel_cell,
                    display_value=selected.nearby_amount_text,
                    numeric_value=numeric_value,
                    should_write_blank=False,
                    status=status,
                    difficulty=difficulty,
                    notes=notes,
                    component_details=" | ".join(component_notes),
                )

            return FieldResult(
                expression=spec.expression,
                label=spec.label,
                excel_cell=spec.excel_cell,
                display_value="",
                numeric_value=None,
                should_write_blank=True,
                status="VALID BLANK",
                difficulty="Easy" if len(occurrences) == 1 else "Medium",
                notes="Code found with no amount. Per rule, this is a valid blank.",
                component_details=" | ".join(component_notes),
            )

        if component_values:
            total = sum(component_values)
            display_value = self.format_number_for_display(total)

            if review_needed:
                status = "NEEDS REVIEW"
                difficulty = "Medium"
                notes = "Combined field calculated, but one or more component codes had multiple occurrences."
            else:
                status = "READY"
                difficulty = "Medium"
                notes = "Combined field calculated from component helper codes."

            return FieldResult(
                expression=spec.expression,
                label=spec.label,
                excel_cell=spec.excel_cell,
                display_value=display_value,
                numeric_value=total,
                should_write_blank=False,
                status=status,
                difficulty=difficulty,
                notes=notes,
                component_details=" | ".join(component_notes),
            )

        if all_components_blank_or_found:
            return FieldResult(
                expression=spec.expression,
                label=spec.label,
                excel_cell=spec.excel_cell,
                display_value="",
                numeric_value=None,
                should_write_blank=True,
                status="VALID BLANK",
                difficulty="Medium",
                notes="All component codes were found, but all appear blank. Per rule, this is valid blank.",
                component_details=" | ".join(component_notes),
            )

        return FieldResult(
            expression=spec.expression,
            label=spec.label,
            excel_cell=spec.excel_cell,
            display_value="",
            numeric_value=None,
            should_write_blank=True,
            status="NEEDS REVIEW",
            difficulty="Medium/Hard",
            notes="Combined field could not be confidently calculated.",
            component_details=" | ".join(component_notes),
        )

    def build_overall_summary(
        self,
        field_results: list[FieldResult],
        occurrences_by_code: dict[str, list[CodeOccurrence]],
        requested_codes: list[str],
        readability_info: dict,
    ) -> list[str]:
        total_fields = len(field_results)
        ready_count = sum(1 for r in field_results if r.status == "READY")
        blank_count = sum(1 for r in field_results if r.status == "VALID BLANK")
        review_count = sum(1 for r in field_results if r.status == "NEEDS REVIEW")
        missing_count = sum(1 for r in field_results if r.status == "MISSING")

        codes_found = sum(1 for code in requested_codes if occurrences_by_code.get(code))
        multiple_codes = [code for code in requested_codes if len(occurrences_by_code.get(code, [])) > 1]

        lines = [
            f"Required Excel yellow input cells checked: {total_fields}",
            f"Unique helper codes requested: {len(requested_codes)}",
            f"Unique helper codes found: {codes_found}/{len(requested_codes)}",
            f"READY fields: {ready_count}",
            f"VALID BLANK fields: {blank_count}",
            f"NEEDS REVIEW fields: {review_count}",
            f"MISSING fields: {missing_count}",
            f"Codes with multiple occurrences: {', '.join(multiple_codes) if multiple_codes else 'None'}",
            f"PDF readability: {readability_info.get('readability')}",
            f"Extracted word count: {readability_info.get('total_words')}",
            f"Extracted text characters: {readability_info.get('total_text_chars')}",
            "",
        ]

        if missing_count == 0:
            lines.append("Excel generation readiness: GOOD.")
            lines.append("Reason: all requested helper codes were found or valid blanks were detected.")
        else:
            lines.append("Excel generation readiness: BLOCKED.")
            lines.append("Reason: one or more helper codes were missing.")

        lines.append("")
        lines.append("Blank fields will be written as true blank Excel cells.")
        lines.append("The backend Excel template is copied first; the original template is not modified.")

        return lines

    def serialize_occurrences(self, occurrences_by_code: dict[str, list[CodeOccurrence]]) -> list[dict]:
        rows = []

        for code in sorted(occurrences_by_code.keys(), key=lambda value: int(value)):
            for occ in occurrences_by_code[code]:
                item = asdict(occ)
                item["location_text"] = occ.location_text
                rows.append(item)

        return rows

    def serialize_field_results(self, field_results: list[FieldResult]) -> list[dict]:
        return [asdict(result) for result in field_results]

    def create_output_excel_path(self, pdf_path: Path) -> Path:
        """
        Output file uses the PDF's base name:
            Statement ABC.pdf -> Statement ABC.xlsx

        If that file already exists, it creates:
            Statement ABC_2.xlsx
            Statement ABC_3.xlsx
        """
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        safe_stem = re.sub(r"[^A-Za-z0-9_\- .]+", "_", pdf_path.stem).strip(" ._")
        if not safe_stem:
            safe_stem = "Completed_Credit_WS"

        output_path = OUTPUT_DIR / f"{safe_stem}.xlsx"

        if not output_path.exists():
            return output_path

        counter = 2
        while True:
            candidate = OUTPUT_DIR / f"{safe_stem}_{counter}.xlsx"
            if not candidate.exists():
                return candidate
            counter += 1

    def generate_excel_from_pdf(self, pdf_path: Path, recalculate_with_excel: bool = False):
        template_path = detect_backend_template()

        self.log("[TEMPLATE] Auto-detected backend template:")
        self.log(str(template_path))

        self.validate_template(template_path)

        field_results, occurrences_by_code, readability_info, summary = self.extract_pdf_field_results(pdf_path)

        missing = [r for r in field_results if r.status == "MISSING"]

        if missing:
            missing_text = ", ".join(f"{r.expression} ({r.label})" for r in missing)
            raise RuntimeError(f"Excel was not generated because these fields are missing: {missing_text}")

        output_path = self.create_output_excel_path(pdf_path)

        self.log("[EXCEL] Copying backend template to output:")
        self.log(f"  Template: {template_path}")
        self.log(f"  Output:   {output_path}")

        shutil.copy2(template_path, output_path)

        self.write_results_to_excel(output_path, field_results)

        if recalculate_with_excel:
            self.recalculate_workbook_with_excel_com(output_path)
        else:
            self.log("[EXCEL] Skipped Excel COM recalculation. Workbook formulas will calculate when opened in Excel.")

        self.log("[SUCCESS] Completed Excel workbook created.")
        self.log(str(output_path))

        return output_path, field_results, occurrences_by_code, readability_info, summary

    def validate_template(self, template_path: Path):
        self.log("[TEMPLATE] Validating workbook template.")

        try:
            wb = load_workbook(template_path, read_only=True, data_only=False)
            sheetnames = wb.sheetnames
            wb.close()
        except Exception as e:
            raise RuntimeError(f"Template could not be opened as a valid Excel workbook: {e}")

        if SHEET_NAME not in sheetnames:
            raise RuntimeError(f"Expected sheet '{SHEET_NAME}' not found. Found sheets: {sheetnames}")

        self.log(f"[TEMPLATE] Validated. Sheetnames: {sheetnames}")

    def write_results_to_excel(self, output_path: Path, field_results: list[FieldResult]):
        self.log("[EXCEL] Opening copied workbook for writing.")
        wb = load_workbook(output_path)

        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(f"Expected sheet '{SHEET_NAME}' not found. Found sheets: {wb.sheetnames}")

        ws = wb[SHEET_NAME]

        self.log("[EXCEL] Writing extracted values into 24 yellow input cells.")

        for result in field_results:
            cell = result.excel_cell

            if result.status == "MISSING":
                raise RuntimeError(f"Cannot write missing field: {result.expression} {result.label}")

            if result.should_write_blank:
                ws[cell].value = None
                self.log(f"  - {cell}: {result.expression} {result.label} -> BLANK")
            else:
                ws[cell].value = result.numeric_value
                self.log(f"  - {cell}: {result.expression} {result.label} -> {result.display_value}")

        self.force_excel_formula_recalculation_on_open(wb)

        wb.save(output_path)
        self.log(f"[EXCEL] Saved workbook: {output_path}")

    def force_excel_formula_recalculation_on_open(self, wb):
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"
            self.log("[EXCEL] Formula recalculation flags set for next Excel open.")
        except Exception as e:
            self.log(f"[WARN] Could not set formula recalculation flags: {e}")

    def recalculate_workbook_with_excel_com(self, output_path: Path):
        self.log("[EXCEL COM] Attempting to open Excel and force recalculation.")

        try:
            import win32com.client
        except Exception as e:
            self.log(f"[WARN] pywin32/win32com is not available. Skipping Excel COM recalculation. Details: {e}")
            return

        excel = None
        wb = None

        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(str(output_path))
            excel.CalculateFullRebuild()
            wb.Save()

            self.log("[EXCEL COM] Workbook recalculated and saved successfully.")

        except Exception as e:
            self.log(f"[WARN] Excel COM recalculation failed. Workbook will still recalculate when opened. Details: {e}")

        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=True)
            except Exception:
                pass

            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass


# ============================================================
# AUDIT REPORT
# ============================================================

def compute_field_metrics(field_results: list[FieldResult]) -> dict:
    total = len(field_results)
    ready = sum(1 for r in field_results if r.status == "READY")
    valid_blanks = sum(1 for r in field_results if r.status == "VALID BLANK")
    needs_review = sum(1 for r in field_results if r.status == "NEEDS REVIEW")
    missing = sum(1 for r in field_results if r.status == "MISSING")
    found = total - missing

    return {
        "total": total,
        "ready": ready,
        "valid_blanks": valid_blanks,
        "needs_review": needs_review,
        "missing": missing,
        "found": found,
        "fields_found_label": f"{found}/{total}",
    }


def write_audit_report(
    user: str,
    original_filename: str,
    output_path: Path,
    template_name: str,
    field_results: list[FieldResult],
    readability_info: dict,
) -> tuple[Path, dict]:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    metrics = compute_field_metrics(field_results)

    audit_path = AUDIT_DIR / f"{output_path.stem}_audit.txt"
    counter = 2
    while audit_path.exists():
        audit_path = AUDIT_DIR / f"{output_path.stem}_audit_{counter}.txt"
        counter += 1

    lines = [
        "PHILLIPCAPITAL RISK MANAGEMENT - CREDIT WORKSHEET AUDIT",
        "=" * 60,
        f"Source PDF:        {original_filename}",
        f"Template Used:     {template_name}",
        f"Generated Output:  {output_path.name}",
        f"Fields Found:      {metrics['fields_found_label']}",
        f"Ready Fields:      {metrics['ready']}",
        f"Valid Blanks:      {metrics['valid_blanks']}",
        f"Needs Review:      {metrics['needs_review']}",
        f"Missing Fields:    {metrics['missing']}",
        f"PDF Readability:   {readability_info.get('readability', '')}",
        f"Generated By:      {user}",
        f"Machine:           {platform.node()}",
        f"Timestamp:         {datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')}",
        "",
        "FIELD-LEVEL DETAIL",
        "-" * 60,
    ]

    for r in field_results:
        value = r.display_value if r.display_value else "(blank)"
        lines.append(
            f"{r.excel_cell:<5} | code {r.expression:<10} | {r.status:<12} | {value:<18} | {r.label}"
        )

    audit_path.write_text("\n".join(lines), encoding="utf-8")
    return audit_path, metrics


# ============================================================
# CLEANUP
# ============================================================

def cleanup_directory(directory: Path, older_than_days: int, extensions: Optional[list[str]] = None) -> int:
    if older_than_days <= 0 or not directory.exists():
        return 0

    cutoff = time.time() - (older_than_days * 86400)
    deleted = 0

    for path in directory.glob("*"):
        if not path.is_file():
            continue
        if extensions and path.suffix.lower() not in extensions:
            continue
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
                deleted += 1
        except Exception:
            pass

    return deleted


def run_age_based_cleanup() -> dict:
    config = load_config()
    return {
        "uploads": cleanup_directory(UPLOAD_DIR, config["cleanup_uploads_days"]),
        "outputs": cleanup_directory(OUTPUT_DIR, config["cleanup_outputs_days"]),
        "audits": cleanup_directory(AUDIT_DIR, config["cleanup_outputs_days"]),
        "logs": cleanup_directory(LOG_DIR, config["cleanup_logs_days"]),
    }


def force_clear_directory(directory: Path, extensions: Optional[list[str]] = None) -> int:
    deleted = 0
    if not directory.exists():
        return 0
    for path in directory.glob("*"):
        if not path.is_file():
            continue
        if extensions and path.suffix.lower() not in extensions:
            continue
        try:
            path.unlink()
            deleted += 1
        except Exception:
            pass
    return deleted


# ============================================================
# FASTAPI APP
# ============================================================

app = FastAPI(
    title="PhillipCapital Risk Management Credit Worksheet Processor",
    description="FastAPI PDF-to-Excel processor for Risk Management credit worksheet automation.",
    version="7.0.0",
)

app.mount("/assets", StaticFiles(directory=str(ASSETS_DIR)), name="assets")


@app.on_event("startup")
def on_startup():
    try:
        run_age_based_cleanup()
    except Exception:
        pass


# ============================================================
# UTILITIES
# ============================================================

def safe_filename(filename: str) -> str:
    filename = Path(filename).name
    filename = re.sub(r"[^A-Za-z0-9_.\- ]+", "_", filename).strip()
    return filename or "uploaded_file"


def save_json_debug(payload: dict, job_id: str) -> Path:
    path = LOG_DIR / f"credit_ws_payload_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return path


def store_job_result(job_id: str, payload: dict):
    global LATEST_JOB_ID

    JOB_RESULTS[job_id] = payload
    LATEST_JOB_ID = job_id

    if len(JOB_RESULTS) > 25:
        oldest_keys = list(JOB_RESULTS.keys())[:-25]
        for key in oldest_keys:
            JOB_RESULTS.pop(key, None)


def get_job_payload(job_id: Optional[str] = None) -> dict:
    if job_id:
        if job_id not in JOB_RESULTS:
            raise HTTPException(status_code=404, detail="Job result not found in current server memory.")
        return JOB_RESULTS[job_id]

    if LATEST_JOB_ID and LATEST_JOB_ID in JOB_RESULTS:
        return JOB_RESULTS[LATEST_JOB_ID]

    json_files = sorted(LOG_DIR.glob("credit_ws_payload_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)

    if not json_files:
        return {}

    return json.loads(json_files[0].read_text(encoding="utf-8"))


def safe_js_json(payload: dict) -> str:
    return json.dumps(payload, default=str).replace("</", "<\\/")


def build_single_run_console_text(payload: dict) -> str:
    field_rows = [[
        "PDF Code(s)",
        "Field",
        "Excel Cell",
        "Preview Value",
        "Status",
        "Difficulty",
        "Notes",
        "Component Details",
    ]]

    for r in payload.get("field_results", []):
        field_rows.append([
            r.get("expression", ""),
            r.get("label", ""),
            r.get("excel_cell", ""),
            r.get("display_value", ""),
            r.get("status", ""),
            r.get("difficulty", ""),
            r.get("notes", ""),
            r.get("component_details", ""),
        ])

    raw_rows = [[
        "Code",
        "Selected",
        "Page",
        "X0",
        "Y0",
        "X1",
        "Y1",
        "Nearby Amount",
        "Confidence Score",
        "Note",
        "Nearby Context",
    ]]

    for r in payload.get("raw_occurrences", []):
        raw_rows.append([
            r.get("code", ""),
            "YES" if r.get("selected") else "NO",
            r.get("page_number", ""),
            round(float(r.get("x0", 0)), 2),
            round(float(r.get("y0", 0)), 2),
            round(float(r.get("x1", 0)), 2),
            round(float(r.get("y1", 0)), 2),
            r.get("nearby_amount_text", ""),
            r.get("confidence_score", ""),
            r.get("note", ""),
            r.get("nearby_context", ""),
        ])

    def clean(value):
        if value is None:
            return ""
        return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()

    def to_tsv(rows):
        return "\n".join("\t".join(clean(cell) for cell in row) for row in rows)

    return (
        "SUMMARY\n"
        + "\n".join(payload.get("summary", []))
        + "\n\nFIELD RESULTS\n"
        + to_tsv(field_rows)
        + "\n\nRAW HELPER CODE OCCURRENCES\n"
        + to_tsv(raw_rows)
        + "\n\nDEBUG LOG\n"
        + "\n".join(payload.get("debug_log", []))
    )


def build_full_console_text(payload: dict) -> str:
    if not payload:
        return "No console data yet."

    if payload.get("type") == "batch":
        lines = []
        lines.append("BATCH SUMMARY")
        lines.append(f"Batch ID: {payload.get('job_id', '')}")
        lines.append(f"Created: {payload.get('created_at', '')}")
        lines.append(f"PDFs submitted: {payload.get('submitted_count', 0)}")
        lines.append(f"Successful outputs: {payload.get('success_count', 0)}")
        lines.append(f"Failed outputs: {payload.get('failure_count', 0)}")
        lines.append("")

        for index, item in enumerate(payload.get("batch_results", []), start=1):
            lines.append("=" * 100)
            lines.append(f"PDF {index}: {item.get('original_filename', '')}")
            lines.append(f"Status: {item.get('status', '')}")

            if item.get("status") == "SUCCESS":
                lines.append(f"Output: {item.get('output_filename', '')}")
                lines.append("")
                lines.append(build_single_run_console_text(item))
            else:
                lines.append(f"Error: {item.get('error', '')}")
                lines.append("")
                lines.append("DEBUG LOG")
                lines.extend(item.get("debug_log", []))

            lines.append("")

        return "\n".join(lines)

    return build_single_run_console_text(payload)


def create_batch_zip(output_paths: list[Path], batch_id: str) -> Path:
    zip_path = OUTPUT_DIR / f"Credit_Worksheet_Batch_{batch_id[:8]}.zip"

    counter = 2
    while zip_path.exists():
        zip_path = OUTPUT_DIR / f"Credit_Worksheet_Batch_{batch_id[:8]}_{counter}.zip"
        counter += 1

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for output_path in output_paths:
            zip_file.write(output_path, arcname=output_path.name)

    return zip_path


def get_file_size_label(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} bytes"

    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"

    return f"{size_bytes / (1024 * 1024):.1f} MB"


# ============================================================
# CSS / HTML HELPERS
# ============================================================

def base_css() -> str:
    return """
    <style>
        :root {
            --pc-blue: #003b7f;
            --pc-blue-dark: #002d62;
            --pc-blue-soft: #eaf2fb;
            --pc-orange: #f59e0b;
            --pc-orange-dark: #d97706;
            --bg: #f5f7fb;
            --card: #ffffff;
            --text: #172033;
            --muted: #667085;
            --border: #d7dde5;
            --green: #15803d;
            --red: #b91c1c;
            --soft-green: #dcfce7;
            --soft-red: #fee2e2;
        }

        * {
            box-sizing: border-box;
        }

        body {
            margin: 0;
            padding: 22px;
            font-family: Segoe UI, Arial, sans-serif;
            background:
                radial-gradient(circle at 0% 0%, rgba(0, 59, 127, 0.08), transparent 28%),
                linear-gradient(180deg, #f8fafc 0%, var(--bg) 100%);
            color: var(--text);
        }

        .shell {
            max-width: 1180px;
            margin: 0 auto;
        }

        .topbar {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 18px;
            position: relative;
            z-index: 20;
        }

        .brand {
            display: flex;
            align-items: center;
            gap: 16px;
            min-width: 0;
        }

        .brand-logo {
            height: 76px;
            width: auto;
            max-width: 180px;
            object-fit: contain;
            display: block;
            flex-shrink: 0;
        }

        .brand-text-wrap {
            display: flex;
            flex-direction: column;
            justify-content: center;
            line-height: 1.05;
        }

        .brand-title {
            font-weight: 950;
            color: var(--pc-blue);
            font-size: 27px;
            letter-spacing: -.04em;
        }

        .brand-subtitle {
            margin-top: 6px;
            font-size: 19px;
            font-weight: 900;
            color: var(--pc-orange-dark);
            letter-spacing: -.02em;
        }

        .topbar-right {
            display: flex;
            align-items: center;
            gap: 12px;
        }

        .user-pill {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            background: white;
            border: 1px solid rgba(0, 59, 127, .18);
            border-radius: 999px;
            padding: 8px 14px;
            font-size: 13px;
            font-weight: 800;
            color: var(--pc-blue-dark);
            box-shadow: 0 8px 24px rgba(0, 59, 127, .08);
            max-width: 260px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }

        .user-pill .dot {
            width: 9px;
            height: 9px;
            border-radius: 50%;
            background: var(--green);
            flex-shrink: 0;
        }

        .menu-wrap {
            position: relative;
            display: inline-block;
        }

        .menu-button {
            width: 48px;
            height: 48px;
            border-radius: 14px;
            border: 1px solid rgba(0, 59, 127, .18);
            background: white;
            color: var(--pc-blue);
            font-size: 25px;
            font-weight: 900;
            cursor: pointer;
            box-shadow: 0 8px 24px rgba(0, 59, 127, .10);
            margin: 0;
            padding: 0;
        }

        .menu-button:hover {
            background: #f8fafc;
        }

        .menu-dropdown {
            display: none;
            position: absolute;
            top: 58px;
            right: 0;
            min-width: 210px;
            background: white;
            border: 1px solid var(--border);
            border-radius: 16px;
            box-shadow: 0 18px 44px rgba(15, 23, 42, 0.18);
            overflow: hidden;
            z-index: 999;
        }

        .menu-dropdown.show {
            display: block;
        }

        .menu-dropdown a {
            display: block;
            padding: 14px 16px;
            text-decoration: none;
            color: var(--text);
            font-weight: 800;
            border-bottom: 1px solid #edf2f7;
        }

        .menu-dropdown a:last-child {
            border-bottom: none;
        }

        .menu-dropdown a:hover {
            background: #f8fafc;
        }

        .menu-dropdown a.active {
            background: var(--pc-blue);
            color: white;
        }

        .menu-dropdown a.logout {
            color: var(--red);
        }

        .hero {
            background: white;
            border: 1px solid var(--border);
            border-radius: 20px;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.08);
            margin-bottom: 18px;
            overflow: hidden;
        }

        .hero-topline {
            height: 8px;
            background: linear-gradient(90deg, var(--pc-blue) 0%, var(--pc-blue-dark) 60%, var(--pc-orange) 100%);
        }

        .hero-inner {
            padding: 24px 28px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 24px;
        }

        .hero-left {
            max-width: 760px;
        }

        .hero-kicker {
            color: var(--pc-orange-dark);
            font-weight: 900;
            text-transform: uppercase;
            letter-spacing: .12em;
            font-size: 12px;
            margin-bottom: 8px;
        }

        .hero h1 {
            margin: 0 0 8px 0;
            font-size: 34px;
            color: var(--pc-blue-dark);
            line-height: 1.1;
            letter-spacing: -.035em;
        }

        .hero p {
            margin: 0;
            color: var(--muted);
            font-size: 15px;
            line-height: 1.55;
        }

        .hero-right {
            min-width: 310px;
            text-align: right;
            display: flex;
            justify-content: flex-end;
        }

        .hero-right-brand {
            display: inline-flex;
            flex-direction: column;
            align-items: flex-end;
            line-height: 1.04;
        }

        .hero-right-title {
            color: var(--pc-blue);
            font-size: 24px;
            font-weight: 900;
            letter-spacing: -.03em;
        }

        .hero-right-subtitle {
            margin-top: 8px;
            color: #5d6f8b;
            font-size: 19px;
            font-weight: 400;
            letter-spacing: -.02em;
            line-height: 1.08;
        }

        .card {
            background: rgba(255, 255, 255, .96);
            border: 1px solid var(--border);
            border-radius: 18px;
            padding: 22px;
            box-shadow: 0 8px 24px rgba(15, 23, 42, 0.07);
            margin-bottom: 18px;
        }

        .center-card {
            max-width: 800px;
            margin: 0 auto;
        }

        h2, h3 {
            margin-top: 0;
            letter-spacing: -.02em;
        }

        .muted {
            color: var(--muted);
            font-size: 14px;
        }

        .status-pill {
            display: inline-block;
            padding: 6px 11px;
            border-radius: 999px;
            font-weight: 800;
            font-size: 13px;
        }

        .good {
            background: var(--soft-green);
            color: var(--green);
        }

        .bad {
            background: var(--soft-red);
            color: var(--red);
        }

        .upload-zone {
            border: 1px dashed #9fb4cf;
            border-radius: 16px;
            background: linear-gradient(180deg, #fbfdff, #f8fbff);
            padding: 18px;
            margin: 12px 0 16px 0;
        }

        input[type=file] {
            width: 100%;
            font-weight: 650;
        }

        input[type=text], input[type=password], input[type=number], input[type=email] {
            width: 100%;
            padding: 12px 14px;
            border: 1px solid var(--border);
            border-radius: 12px;
            font-size: 15px;
            font-family: inherit;
            background: white;
        }

        input:focus {
            outline: none;
            border-color: var(--pc-blue);
            box-shadow: 0 0 0 3px rgba(0, 59, 127, .12);
        }

        button, .button-link {
            border: 0;
            border-radius: 12px;
            padding: 12px 17px;
            font-weight: 900;
            cursor: pointer;
            background: var(--pc-blue);
            color: white;
            margin: 4px 4px 4px 0;
            text-decoration: none;
            display: inline-block;
            box-shadow: 0 8px 18px rgba(0, 59, 127, .18);
            font-family: inherit;
            font-size: 14px;
        }

        button:hover, .button-link:hover {
            filter: brightness(1.04);
        }

        button.orange, .button-link.orange {
            background: var(--pc-orange-dark);
        }

        button.secondary, .button-link.secondary {
            background: #475467;
        }

        button.danger, .button-link.danger {
            background: var(--red);
            box-shadow: 0 8px 18px rgba(185, 28, 28, .18);
        }

        code {
            background: #eef2f7;
            padding: 3px 6px;
            border-radius: 5px;
            word-break: break-all;
        }

        .output-box {
            border-left: 5px solid var(--green);
            padding: 14px;
            background: #f0fdf4;
            border-radius: 12px;
            margin-top: 14px;
            display: none;
        }

        .error-box {
            border-left: 5px solid var(--red);
            padding: 14px;
            background: #fff1f2;
            border-radius: 12px;
            margin-top: 14px;
            display: none;
            white-space: pre-wrap;
        }

        .console {
            background: #07111f;
            color: #d1e7ff;
            padding: 16px;
            border-radius: 14px;
            min-height: 590px;
            overflow: auto;
            font-family: Consolas, monospace;
            font-size: 12px;
            white-space: pre;
            border: 1px solid rgba(118, 183, 255, .25);
        }

        .summary-list {
            background: #f8fafc;
            padding: 14px;
            border-radius: 14px;
            border: 1px solid var(--border);
            font-family: Consolas, monospace;
            white-space: pre-wrap;
            font-size: 13px;
        }

        .metric-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 12px;
            margin-top: 16px;
        }

        .metric {
            background: #f8fafc;
            border: 1px solid var(--border);
            padding: 14px;
            border-radius: 14px;
        }

        .metric b {
            color: var(--pc-blue);
            display: block;
            font-size: 20px;
            margin-bottom: 4px;
        }

        .outputs-grid {
            display: grid;
            grid-template-columns: 1fr;
            gap: 14px;
        }

        .output-item {
            border: 1px solid var(--border);
            border-radius: 16px;
            background: linear-gradient(180deg, #ffffff 0%, #fbfcfe 100%);
            padding: 18px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 18px;
            box-shadow: 0 6px 18px rgba(15, 23, 42, 0.05);
        }

        .output-left {
            min-width: 0;
            flex: 1;
        }

        .output-title {
            font-size: 16px;
            font-weight: 800;
            color: var(--pc-blue-dark);
            margin-bottom: 6px;
            word-break: break-word;
        }

        .output-meta {
            font-size: 13px;
            color: var(--muted);
            display: flex;
            flex-wrap: wrap;
            gap: 14px;
        }

        .empty-state {
            text-align: center;
            padding: 40px 20px;
            border: 1px dashed var(--border);
            border-radius: 16px;
            background: #fafcff;
            color: var(--muted);
        }

        /* ---- Outputs page (folder/file view) ---- */
        .outputs-list { display: flex; flex-direction: column; }

        .date-divider {
            display: flex;
            align-items: baseline;
            gap: 12px;
            margin: 28px 0 12px 0;
            padding: 0 4px 10px 4px;
            border-bottom: 1px solid var(--border);
        }
        .date-divider:first-child { margin-top: 0; }
        .date-divider .date-label {
            font-size: 18px;
            font-weight: 800;
            color: var(--pc-blue-dark);
            letter-spacing: -0.2px;
        }
        .date-divider .date-meta {
            font-size: 11px;
            color: var(--muted);
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.8px;
        }
        .date-divider .date-count {
            margin-left: auto;
            font-size: 12px;
            color: var(--muted);
            background: var(--pc-blue-soft);
            border-radius: 999px;
            padding: 3px 10px;
            font-weight: 700;
        }

        .folder-card {
            border: 1px solid var(--border);
            border-radius: 14px;
            background: white;
            margin-bottom: 12px;
            overflow: hidden;
            transition: box-shadow 0.15s ease, border-color 0.15s ease;
        }
        .folder-card:hover { border-color: rgba(0, 59, 127, 0.25); }
        .folder-card[open] {
            box-shadow: 0 6px 18px rgba(15, 23, 42, 0.06);
            border-color: rgba(0, 59, 127, 0.18);
        }
        .folder-card > summary {
            display: flex;
            align-items: center;
            gap: 14px;
            padding: 14px 16px;
            cursor: pointer;
            list-style: none;
            background: linear-gradient(180deg, #fcfdfe 0%, #f4f7fc 100%);
        }
        .folder-card > summary::-webkit-details-marker { display: none; }
        .folder-card > summary::before {
            content: "";
            display: inline-block;
            width: 0;
            height: 0;
            border-top: 5px solid transparent;
            border-bottom: 5px solid transparent;
            border-left: 6px solid var(--pc-blue);
            transition: transform 0.15s ease;
            flex-shrink: 0;
        }
        .folder-card[open] > summary::before { transform: rotate(90deg); }

        .file-icon {
            flex-shrink: 0;
            width: 36px;
            height: 36px;
            border-radius: 8px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            font-size: 18px;
            color: white;
        }
        .file-icon.folder { background: var(--pc-orange); }
        .file-icon.xlsx   { background: #16a34a; }
        .file-icon.audit  { background: var(--pc-blue); }
        .file-icon svg { width: 20px; height: 20px; }

        .file-row .file-icon {
            width: 28px;
            height: 28px;
            font-size: 11px;
            font-weight: 800;
            letter-spacing: 0.4px;
            border-radius: 6px;
        }
        .file-row .file-icon svg { width: 16px; height: 16px; }

        .folder-info, .file-info {
            flex: 1;
            min-width: 0;
        }
        .folder-name {
            font-weight: 800;
            font-size: 15px;
            color: var(--pc-blue-dark);
            word-break: break-word;
        }
        .folder-meta, .file-meta {
            font-size: 12px;
            color: var(--muted);
            margin-top: 3px;
            display: flex;
            flex-wrap: wrap;
            gap: 4px 12px;
        }
        .folder-meta .dot, .file-meta .dot {
            color: #c0c8d4;
        }
        .folder-meta code {
            background: #f3f5f9;
            padding: 1px 6px;
            border-radius: 4px;
            font-family: var(--mono, monospace);
            font-size: 11px;
            color: var(--text);
        }
        .folder-actions, .file-actions {
            display: flex;
            gap: 6px;
            flex-shrink: 0;
            align-items: center;
        }

        .folder-contents {
            padding: 4px 16px 12px 56px;
            background: #fafbfd;
            border-top: 1px solid var(--border);
        }

        .file-row {
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 10px 0;
            border-bottom: 1px solid #eef1f7;
        }
        .file-row:last-child { border-bottom: none; }
        .file-row .file-name {
            font-weight: 700;
            color: var(--pc-blue-dark);
            font-size: 13px;
            word-break: break-all;
        }
        .file-row .file-actions a { padding: 6px 10px; font-size: 12px; }

        .single-card {
            display: flex;
            align-items: center;
            gap: 14px;
            padding: 14px 16px;
            border: 1px solid var(--border);
            border-radius: 14px;
            background: white;
            margin-bottom: 12px;
            transition: box-shadow 0.15s ease, border-color 0.15s ease;
        }
        .single-card:hover {
            border-color: rgba(0, 59, 127, 0.25);
            box-shadow: 0 4px 14px rgba(15, 23, 42, 0.05);
        }
        .single-card.audit-only { background: #fafcff; }
        /* ---- /Outputs page ---- */

        .batch-note {
            background: var(--pc-blue-soft);
            border: 1px solid rgba(0, 59, 127, .12);
            color: var(--pc-blue-dark);
            padding: 12px 14px;
            border-radius: 14px;
            font-size: 14px;
            font-weight: 650;
            margin-top: 12px;
        }

        /* ---------- LOGIN ---------- */
        .login-shell {
            min-height: 84vh;
            display: flex;
            align-items: center;
            justify-content: center;
        }

        .login-card {
            width: 420px;
            max-width: 100%;
            background: white;
            border: 1px solid var(--border);
            border-radius: 22px;
            box-shadow: 0 24px 70px rgba(15, 23, 42, .16);
            overflow: hidden;
        }

        .login-topline {
            height: 8px;
            background: linear-gradient(90deg, var(--pc-blue) 0%, var(--pc-blue-dark) 60%, var(--pc-orange) 100%);
        }

        .login-body {
            padding: 30px 30px 34px 30px;
        }

        .login-logo {
            text-align: center;
            margin-bottom: 16px;
        }

        .login-logo img {
            height: 70px;
            max-width: 200px;
            object-fit: contain;
        }

        .login-title {
            text-align: center;
            color: var(--pc-blue-dark);
            font-size: 24px;
            font-weight: 950;
            letter-spacing: -.03em;
            margin-bottom: 4px;
        }

        .login-sub {
            text-align: center;
            color: var(--muted);
            font-size: 14px;
            margin-bottom: 22px;
        }

        .field-group {
            margin-bottom: 14px;
        }

        .field-group label {
            display: block;
            font-weight: 800;
            font-size: 13px;
            color: var(--pc-blue-dark);
            margin-bottom: 6px;
        }

        .login-error {
            background: var(--soft-red);
            color: var(--red);
            border-radius: 12px;
            padding: 11px 14px;
            font-size: 14px;
            font-weight: 700;
            margin-bottom: 16px;
        }

        .full-btn {
            width: 100%;
            text-align: center;
            margin: 8px 0 0 0;
            padding: 14px;
            font-size: 15px;
        }

        /* ---------- PROGRESS ---------- */
        .progress-wrap {
            margin-top: 18px;
            display: none;
        }

        .progress-head {
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 10px;
            margin-bottom: 12px;
        }

        .progress-counts {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }

        .count-chip {
            border-radius: 999px;
            padding: 6px 12px;
            font-weight: 800;
            font-size: 13px;
            border: 1px solid var(--border);
            background: #f8fafc;
        }

        .count-chip.ok { background: var(--soft-green); color: var(--green); border-color: transparent; }
        .count-chip.fail { background: var(--soft-red); color: var(--red); border-color: transparent; }

        .progress-list {
            display: flex;
            flex-direction: column;
            gap: 10px;
        }

        .progress-row {
            border: 1px solid var(--border);
            border-radius: 14px;
            padding: 14px 16px;
            background: white;
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 14px;
        }

        .progress-row .pr-left { min-width: 0; flex: 1; }

        .progress-row .pr-name {
            font-weight: 800;
            color: var(--pc-blue-dark);
            word-break: break-word;
        }

        .progress-row .pr-detail {
            font-size: 12.5px;
            color: var(--muted);
            margin-top: 4px;
            word-break: break-word;
        }

        .progress-row .pr-detail.err { color: var(--red); font-weight: 700; }

        .pstat {
            display: inline-flex;
            align-items: center;
            gap: 7px;
            border-radius: 999px;
            padding: 6px 12px;
            font-weight: 800;
            font-size: 12.5px;
            white-space: nowrap;
        }

        .pstat .pdot { width: 8px; height: 8px; border-radius: 50%; }

        .pstat.queued { background: #eef2f7; color: #475467; }
        .pstat.queued .pdot { background: #94a3b8; }

        .pstat.processing { background: var(--pc-blue-soft); color: var(--pc-blue); }
        .pstat.processing .pdot { background: var(--pc-blue); animation: pulse 1s infinite; }

        .pstat.complete { background: var(--soft-green); color: var(--green); }
        .pstat.complete .pdot { background: var(--green); }

        .pstat.failed { background: var(--soft-red); color: var(--red); }
        .pstat.failed .pdot { background: var(--red); }

        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: .35; }
        }

        .pr-actions { display: flex; flex-direction: column; align-items: flex-end; gap: 6px; }
        .pr-actions a { font-size: 12px; padding: 7px 12px; margin: 0; }

        .download-bar {
            display: none;
            margin-top: 16px;
            padding: 16px;
            border-radius: 14px;
            background: #f0fdf4;
            border: 1px solid #bbf7d0;
        }

        .download-bar.has-fail { background: #fffbeb; border-color: #fde68a; }

        /* ---------- SETTINGS ---------- */
        .settings-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 18px;
        }

        .user-row {
            display: flex;
            flex-direction: column;
            gap: 10px;
            border: 1px solid var(--border);
            border-radius: 12px;
            padding: 12px 14px;
            margin-bottom: 8px;
            background: white;
        }

        .user-row .ur-top {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 12px;
        }

        .user-row .ur-pwform {
            display: flex;
            gap: 6px;
            align-items: stretch;
            flex-wrap: wrap;
        }

        .user-row .ur-pwform input[type="password"] {
            flex: 1;
            min-width: 180px;
            padding: 8px 10px;
            font-size: 13px;
            border: 1px solid var(--border);
            border-radius: 8px;
            font-family: inherit;
            background: #fff;
        }

        .user-row .ur-pwform button {
            padding: 8px 12px;
            font-size: 12px;
            white-space: nowrap;
        }

        .user-row .ur-name { font-weight: 800; color: var(--pc-blue-dark); word-break: break-all; }
        .user-row .ur-meta { font-size: 12px; color: var(--muted); margin-top: 2px; }
        .user-row .ur-self-tag {
            display: inline-block;
            margin-left: 6px;
            padding: 1px 8px;
            background: var(--pc-blue);
            color: white;
            border-radius: 10px;
            font-size: 10px;
            font-weight: 800;
            letter-spacing: 0.5px;
            text-transform: uppercase;
            vertical-align: middle;
        }

        .flash {
            border-radius: 12px;
            padding: 12px 14px;
            font-weight: 700;
            font-size: 14px;
            margin-bottom: 16px;
        }
        .flash.ok { background: var(--soft-green); color: var(--green); }
        .flash.err { background: var(--soft-red); color: var(--red); }

        .limit-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 14px;
        }

        .loading-overlay {
            display: none;
            position: fixed;
            inset: 0;
            background: rgba(15, 23, 42, 0.38);
            z-index: 2000;
            align-items: center;
            justify-content: center;
        }

        .loading-card {
            width: 380px;
            background: white;
            border-radius: 18px;
            padding: 24px;
            text-align: center;
            box-shadow: 0 24px 70px rgba(15, 23, 42, .25);
        }

        .spinner {
            width: 42px;
            height: 42px;
            border: 5px solid #e2e8f0;
            border-top-color: var(--pc-orange);
            border-radius: 50%;
            margin: 0 auto 14px auto;
            animation: spin 1s linear infinite;
        }

        @keyframes spin {
            to { transform: rotate(360deg); }
        }

        @media (max-width: 900px) {
            body { padding: 14px; }
            .brand-logo { height: 58px; max-width: 140px; }
            .brand-title { font-size: 22px; }
            .brand-subtitle { font-size: 16px; }
            .hero-inner { flex-direction: column; align-items: flex-start; }
            .hero-right { min-width: 0; width: 100%; justify-content: flex-start; text-align: left; }
            .hero-right-brand { align-items: flex-start; }
            .hero-right-title { font-size: 22px; }
            .hero-right-subtitle { font-size: 17px; }
            .hero h1 { font-size: 28px; }
            .metric-grid { grid-template-columns: 1fr; }
            .output-item { flex-direction: column; align-items: flex-start; }
            .settings-grid { grid-template-columns: 1fr; }
            .limit-grid { grid-template-columns: 1fr; }
            .user-pill { max-width: 150px; }
            .folder-card > summary { flex-wrap: wrap; gap: 10px; }
            .single-card { flex-wrap: wrap; gap: 10px; }
            .folder-info, .file-info { flex-basis: 100%; min-width: 0; order: 2; }
            .folder-actions, .file-actions { flex-basis: 100%; order: 3; }
            .folder-contents { padding: 4px 12px 12px 12px; }
            .date-divider { flex-wrap: wrap; }
            .date-divider .date-count { margin-left: 0; }
        }
    </style>
    """


def head_html(title: str) -> str:
    return f"""
    <head>
        <title>{html.escape(title)}</title>
        <link rel="icon" type="image/png" href="/favicon.ico">
        {base_css()}
    </head>
    """


def nav_html(active: str) -> str:
    home_active = "active" if active == "home" else ""
    console_active = "active" if active == "console" else ""
    outputs_active = "active" if active == "outputs" else ""
    settings_active = "active" if active == "settings" else ""

    return f"""
    <div class="menu-wrap">
        <button class="menu-button" onclick="toggleMenu(event)" title="Menu">&#9776;</button>
        <div id="pageMenu" class="menu-dropdown">
            <a class="{home_active}" href="/">Home</a>
            <a class="{console_active}" href="/console">Console</a>
            <a class="{outputs_active}" href="/outputs">Outputs</a>
            <a class="{settings_active}" href="/settings">Settings</a>
            <a class="logout" href="/logout">Sign out</a>
        </div>
    </div>

    <script>
        function toggleMenu(event) {{
            event.stopPropagation();
            const menu = document.getElementById("pageMenu");
            menu.classList.toggle("show");
        }}

        document.addEventListener("click", function() {{
            const menu = document.getElementById("pageMenu");
            if (menu) {{
                menu.classList.remove("show");
            }}
        }});
    </script>
    """


def topbar_html(active: str, user: Optional[str] = None) -> str:
    logo_url = get_logo_url()

    if logo_url:
        logo_html = f'<img src="{html.escape(logo_url)}" alt="PhillipCapital" class="brand-logo" />'
    else:
        logo_html = ""

    user_pill = ""
    if user:
        user_pill = f'<div class="user-pill"><span class="dot"></span>{html.escape(user)}</div>'

    return f"""
    <div class="topbar">
        <div class="brand">
            {logo_html}
            <div class="brand-text-wrap">
                <div class="brand-title">Phillip Capital</div>
                <div class="brand-subtitle">Risk Management</div>
            </div>
        </div>
        <div class="topbar-right">
            {user_pill}
            {nav_html(active)}
        </div>
    </div>
    """


def hero_html(title: str, subtitle: str) -> str:
    return f"""
    <div class="hero">
        <div class="hero-topline"></div>
        <div class="hero-inner">
            <div class="hero-left">
                <div class="hero-kicker">Credit Worksheet Automation</div>
                <h1>{html.escape(title)}</h1>
                <p>{html.escape(subtitle)}</p>
            </div>
            <div class="hero-right">
                <div class="hero-right-brand">
                    <div class="hero-right-title">PhillipCapital</div>
                    <div class="hero-right-subtitle">
                        Securities, Fixed Income,<br />
                        Futures, Options.
                    </div>
                </div>
            </div>
        </div>
    </div>
    """


# ============================================================
# PAGES
# ============================================================

def login_page_html(error: str = "") -> str:
    logo_url = get_logo_url()
    logo_block = ""
    if logo_url:
        logo_block = f'<div class="login-logo"><img src="{html.escape(logo_url)}" alt="PhillipCapital" /></div>'

    error_block = ""
    if error:
        error_block = f'<div class="login-error">{html.escape(error)}</div>'

    page = """
<!doctype html>
<html>
__HEAD__
<body>
<div class="shell">
    <div class="login-shell">
        <div class="login-card">
            <div class="login-topline"></div>
            <div class="login-body">
                __LOGO__
                <div class="login-title">Risk Management Portal</div>
                <div class="login-sub">Credit Worksheet Processor &middot; Sign in to continue</div>
                __ERROR__
                <form method="post" action="/login">
                    <div class="field-group">
                        <label for="username">Username</label>
                        <input type="text" id="username" name="username" autocomplete="username" autofocus required />
                    </div>
                    <div class="field-group">
                        <label for="password">Password</label>
                        <input type="password" id="password" name="password" autocomplete="current-password" required />
                    </div>
                    <button type="submit" class="orange full-btn">Sign In</button>
                </form>
            </div>
        </div>
    </div>
</div>
</body>
</html>
    """

    page = page.replace("__HEAD__", head_html("Sign In | Phillip Capital Risk Management"))
    page = page.replace("__LOGO__", logo_block)
    page = page.replace("__ERROR__", error_block)
    return page


def home_page_html(user: str) -> str:
    config = load_config()
    template = get_template_status()
    template_class = "good" if template["exists"] else "bad"
    template_label = "Template detected" if template["exists"] else "Template missing"

    template_note = (
        f"Using backend Excel template: {template['filename']}"
        if template["exists"]
        else f"Place a blank .xlsx template inside: {template['folder']}"
    )

    page = """
<!doctype html>
<html>
__HEAD__
<body>
<div class="shell">
    __TOPBAR__
    __HERO__

    <div class="card center-card">
        <h2>Upload Customer PDF(s)</h2>
        <p>Status: <span class="status-pill __TEMPLATE_CLASS__">__TEMPLATE_LABEL__</span></p>
        <p class="muted">__TEMPLATE_NOTE__</p>

        <form id="generateForm">
            <div class="upload-zone">
                <input type="file" id="generatePdf" name="pdfs" accept="application/pdf" multiple required />
            </div>

            <div class="batch-note">
                Upload one PDF to generate one Excel file, or upload multiple PDFs to generate a ZIP containing one Excel workbook per PDF.
                Limits: up to <strong>__MAX_BATCH__</strong> PDFs per batch, <strong>__MAX_SIZE__ MB</strong> per file.
            </div>

            <label class="muted" style="display:block; margin-top:14px;">
                <input type="checkbox" id="recalculateExcel" name="recalculate_with_excel" value="true" />
                Recalculate formulas with Excel COM if available
            </label>

            <br />
            <button type="submit" class="orange">Generate Completed Excel</button>
        </form>

        <div id="errorBox" class="error-box"></div>

        <div id="progressWrap" class="progress-wrap">
            <div class="progress-head">
                <h3 style="margin:0;">Batch Progress</h3>
                <div id="progressCounts" class="progress-counts"></div>
            </div>
            <div id="progressList" class="progress-list"></div>
            <div id="downloadBar" class="download-bar"></div>
        </div>

        <div class="metric-grid">
            <div class="metric">
                <b>Batch</b>
                Process one or many PDFs
            </div>
            <div class="metric">
                <b>Live</b>
                Per-PDF status updates
            </div>
            <div class="metric">
                <b>Audit</b>
                Report saved per workbook
            </div>
        </div>
    </div>
</div>

<script>
const MAX_BATCH = __MAX_BATCH__;
const MAX_SIZE_BYTES = __MAX_SIZE__ * 1024 * 1024;
let pollTimer = null;

function escapeHtml(value) {
    if (value === null || value === undefined) return "";
    return String(value)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
}

function showError(message) {
    const box = document.getElementById("errorBox");
    box.style.display = "block";
    box.textContent = message;
}

function clearError() {
    const box = document.getElementById("errorBox");
    box.style.display = "none";
    box.textContent = "";
}

function statusBadge(status) {
    const labels = { queued: "Queued", processing: "Processing", complete: "Complete", failed: "Failed" };
    const label = labels[status] || status;
    return `<span class="pstat ${status}"><span class="pdot"></span>${escapeHtml(label)}</span>`;
}

function renderProgress(job) {
    const wrap = document.getElementById("progressWrap");
    wrap.style.display = "block";

    const counts = document.getElementById("progressCounts");
    counts.innerHTML =
        `<span class="count-chip">Total: ${job.total}</span>` +
        `<span class="count-chip ok">Complete: ${job.success_count}</span>` +
        `<span class="count-chip fail">Failed: ${job.failure_count}</span>`;

    const list = document.getElementById("progressList");
    list.innerHTML = (job.files || []).map(function(f) {
        let detail = "";
        if (f.status === "complete") {
            detail = `<div class="pr-detail">Output: ${escapeHtml(f.output_filename || "")} &middot; Fields ${escapeHtml(f.fields_found || "")} &middot; Needs review: ${escapeHtml(String(f.needs_review))}</div>`;
        } else if (f.status === "failed") {
            detail = `<div class="pr-detail err">${escapeHtml(f.error || "Failed")}</div>`;
        } else if (f.status === "processing") {
            detail = `<div class="pr-detail">Extracting values and writing workbook...</div>`;
        } else {
            detail = `<div class="pr-detail">Waiting in queue</div>`;
        }

        let actions = "";
        if (f.status === "complete") {
            if (f.download_url) actions += `<a class="button-link" href="${f.download_url}">Excel</a>`;
            if (f.audit_url) actions += `<a class="button-link secondary" href="${f.audit_url}">Audit</a>`;
        }

        return `<div class="progress-row">
            <div class="pr-left">
                <div class="pr-name">${escapeHtml(f.filename)}</div>
                ${detail}
            </div>
            <div class="pr-actions">
                ${statusBadge(f.status)}
                ${actions}
            </div>
        </div>`;
    }).join("");

    const bar = document.getElementById("downloadBar");
    if (job.status === "complete") {
        bar.style.display = "block";
        bar.className = job.failure_count > 0 ? "download-bar has-fail" : "download-bar";

        let summary = `<strong>${job.success_count}</strong> successful`;
        if (job.failure_count > 0) summary += ` &middot; <strong>${job.failure_count}</strong> failed`;

        let links = "";
        if (job.zip_url) {
            links += `<a class="button-link orange" href="${job.zip_url}">Download All (ZIP)</a>`;
        } else if (job.single_url) {
            links += `<a class="button-link orange" href="${job.single_url}">Download Excel</a>`;
        }
        if (job.console_url) {
            links += `<a class="button-link secondary" href="${job.console_url}">Open Console</a>`;
        }

        bar.innerHTML = `<div style="margin-bottom:10px;font-weight:800;">${summary}</div>${links}`;
    } else {
        bar.style.display = "none";
    }
}

async function poll(jobId) {
    try {
        const res = await fetch(`/batch-status/${jobId}`);
        if (res.status === 401) { window.location = "/login"; return; }
        if (!res.ok) { showError("Lost connection to batch job."); return; }

        const job = await res.json();
        renderProgress(job);

        if (job.status === "complete") {
            if (pollTimer) clearTimeout(pollTimer);
            return;
        }
        pollTimer = setTimeout(() => poll(jobId), 900);
    } catch (err) {
        showError("Status polling failed: " + err.message);
    }
}

document.getElementById("generateForm").addEventListener("submit", async function(e) {
    e.preventDefault();
    clearError();

    const files = document.getElementById("generatePdf").files;

    if (!files || files.length === 0) {
        showError("Please select at least one PDF.");
        return;
    }

    if (files.length > MAX_BATCH) {
        showError(`Too many files. The batch limit is ${MAX_BATCH} PDFs.`);
        return;
    }

    for (let i = 0; i < files.length; i++) {
        if (files[i].size > MAX_SIZE_BYTES) {
            showError(`"${files[i].name}" is larger than the ${MAX_SIZE_BYTES / (1024*1024)} MB limit.`);
            return;
        }
    }

    const formData = new FormData();
    for (let i = 0; i < files.length; i++) {
        formData.append("pdfs", files[i]);
    }
    if (document.getElementById("recalculateExcel").checked) {
        formData.append("recalculate_with_excel", "true");
    }

    try {
        const response = await fetch("/start-batch", { method: "POST", body: formData });
        if (response.status === 401) { window.location = "/login"; return; }

        if (!response.ok) {
            const errText = await response.text();
            throw new Error(errText);
        }

        const data = await response.json();
        if (!data.job_id) throw new Error("Server did not return a job id.");

        document.getElementById("downloadBar").style.display = "none";
        poll(data.job_id);
    } catch (err) {
        showError("Could not start batch:\\n" + err.message);
    }
});
</script>
</body>
</html>
    """

    page = page.replace("__HEAD__", head_html("Phillip Capital Risk Management | Credit Worksheet Processor"))
    page = page.replace("__TOPBAR__", topbar_html("home", user))
    page = page.replace(
        "__HERO__",
        hero_html(
            "Generate Credit Worksheet",
            "Upload customer PDF files and automatically produce completed Excel credit worksheets using the backend template.",
        ),
    )
    page = page.replace("__TEMPLATE_CLASS__", template_class)
    page = page.replace("__TEMPLATE_LABEL__", template_label)
    page = page.replace("__TEMPLATE_NOTE__", html.escape(template_note))
    page = page.replace("__MAX_BATCH__", str(config["max_batch_size"]))
    page = page.replace("__MAX_SIZE__", str(config["max_pdf_size_mb"]))

    return page


def settings_page_html(user: str, flash: str = "", flash_type: str = "ok") -> str:
    config = load_config()
    users = load_users()

    flash_block = ""
    if flash:
        flash_block = f'<div class="flash {html.escape(flash_type)}">{html.escape(flash)}</div>'

    user_rows = []
    for record in sorted(users.values(), key=lambda r: r.get("username", "").lower()):
        uname = record.get("username", "")
        created_at = record.get("created_at", "")
        created_by = record.get("created_by", "")
        is_self = uname.strip().lower() == user.strip().lower()

        delete_control = (
            '<span class="muted" style="font-size:12px;">Current session</span>'
            if is_self
            else f"""
            <form method="post" action="/settings/delete-user" onsubmit="return confirm('Remove {html.escape(uname)}?');" style="margin:0;">
                <input type="hidden" name="username" value="{html.escape(uname)}" />
                <button type="submit" class="danger" style="padding:8px 12px; font-size:12px;">Remove</button>
            </form>
            """
        )

        pw_updated_at = record.get("password_updated_at", "")
        pw_updated_by = record.get("password_updated_by", "")
        if pw_updated_at:
            meta_pw_line = (
                f'<div class="ur-meta">Password updated {html.escape(pw_updated_at)}'
                f' by {html.escape(pw_updated_by)}</div>'
            )
        else:
            meta_pw_line = ""

        self_tag = '<span class="ur-self-tag">You</span>' if is_self else ""
        confirm_msg = (
            "Change your own password?"
            if is_self
            else f"Update password for {uname}? Any existing sessions for this user will be signed out."
        )

        user_rows.append(
            f"""
            <div class="user-row">
                <div class="ur-top">
                    <div style="min-width:0;">
                        <div class="ur-name">{html.escape(uname)}{self_tag}</div>
                        <div class="ur-meta">Added {html.escape(created_at)} by {html.escape(created_by)}</div>
                        {meta_pw_line}
                    </div>
                    <div>{delete_control}</div>
                </div>
                <form method="post" action="/settings/update-password" class="ur-pwform"
                      onsubmit="return confirm('{html.escape(confirm_msg)}');">
                    <input type="hidden" name="username" value="{html.escape(uname)}" />
                    <input type="password" name="new_password" placeholder="New password for {html.escape(uname)}"
                           required minlength="1" autocomplete="new-password" />
                    <button type="submit" class="primary">Update password</button>
                </form>
            </div>
            """
        )

    users_block = "".join(user_rows) if user_rows else '<p class="muted">No users found.</p>'

    page = """
<!doctype html>
<html>
__HEAD__
<body>
<div class="shell">
    __TOPBAR__
    __HERO__

    __FLASH__

    <div class="settings-grid">
        <div class="card">
            <h2>Login Credentials</h2>
            <p class="muted">Existing users who can sign in to the portal.</p>
            __USERS_BLOCK__
        </div>

        <div class="card">
            <h2>Add New Login</h2>
            <p class="muted">Create a username and password. Passwords are stored hashed.</p>
            <form method="post" action="/settings/add-user">
                <div class="field-group">
                    <label for="new_username">Username</label>
                    <input type="text" id="new_username" name="new_username" placeholder="name@phillipcapital.com" required />
                </div>
                <div class="field-group">
                    <label for="new_password">Password</label>
                    <input type="text" id="new_password" name="new_password" placeholder="Set a password" required />
                </div>
                <button type="submit" class="orange">Add User</button>
            </form>
        </div>
    </div>

    <div class="card">
        <h2>Upload Limits</h2>
        <p class="muted">Controls applied to every batch. Session timeout also signs users out after inactivity.</p>
        <form method="post" action="/settings/limits">
            <div class="limit-grid">
                <div class="field-group">
                    <label for="max_pdf_size_mb">Max PDF size (MB)</label>
                    <input type="number" min="1" max="500" id="max_pdf_size_mb" name="max_pdf_size_mb" value="__MAX_SIZE__" />
                </div>
                <div class="field-group">
                    <label for="max_batch_size">Max PDFs per batch</label>
                    <input type="number" min="1" max="200" id="max_batch_size" name="max_batch_size" value="__MAX_BATCH__" />
                </div>
                <div class="field-group">
                    <label for="session_timeout_minutes">Session timeout (minutes)</label>
                    <input type="number" min="5" max="1440" id="session_timeout_minutes" name="session_timeout_minutes" value="__TIMEOUT__" />
                </div>
            </div>
            <button type="submit">Save Limits</button>
        </form>
    </div>

    <div class="card">
        <h2>Cleanup</h2>
        <p class="muted">
            Automatic cleanup runs on startup. Uploads older than <strong>__CU_UPLOADS__</strong> days,
            outputs/audits older than <strong>__CU_OUTPUTS__</strong> days, and logs older than <strong>__CU_LOGS__</strong> days
            are removed. You can also change retention or clear folders now.
        </p>

        <form method="post" action="/settings/retention">
            <div class="limit-grid">
                <div class="field-group">
                    <label for="cleanup_uploads_days">Keep uploads (days)</label>
                    <input type="number" min="0" max="365" id="cleanup_uploads_days" name="cleanup_uploads_days" value="__CU_UPLOADS__" />
                </div>
                <div class="field-group">
                    <label for="cleanup_outputs_days">Keep outputs/audits (days)</label>
                    <input type="number" min="0" max="365" id="cleanup_outputs_days" name="cleanup_outputs_days" value="__CU_OUTPUTS__" />
                </div>
                <div class="field-group">
                    <label for="cleanup_logs_days">Keep logs (days)</label>
                    <input type="number" min="0" max="365" id="cleanup_logs_days" name="cleanup_logs_days" value="__CU_LOGS__" />
                </div>
            </div>
            <button type="submit">Save Retention</button>
        </form>

        <hr style="border:none; border-top:1px solid var(--border); margin:18px 0;" />

        <p class="muted">Run cleanup immediately:</p>
        <form method="post" action="/settings/cleanup" style="display:inline;">
            <input type="hidden" name="target" value="age" />
            <button type="submit" class="secondary">Run Age-Based Cleanup</button>
        </form>
        <form method="post" action="/settings/cleanup" style="display:inline;" onsubmit="return confirm('Clear ALL uploaded PDFs now?');">
            <input type="hidden" name="target" value="uploads" />
            <button type="submit" class="danger">Clear All Uploads</button>
        </form>
        <form method="post" action="/settings/cleanup" style="display:inline;" onsubmit="return confirm('Clear ALL logs now?');">
            <input type="hidden" name="target" value="logs" />
            <button type="submit" class="danger">Clear All Logs</button>
        </form>
        <form method="post" action="/settings/cleanup" style="display:inline;" onsubmit="return confirm('Clear ALL generated outputs and audits now?');">
            <input type="hidden" name="target" value="outputs" />
            <button type="submit" class="danger">Clear All Outputs</button>
        </form>
    </div>
</div>
</body>
</html>
    """

    page = page.replace("__HEAD__", head_html("Settings | Phillip Capital Risk Management"))
    page = page.replace("__TOPBAR__", topbar_html("settings", user))
    page = page.replace(
        "__HERO__",
        hero_html(
            "Settings",
            "Manage portal logins, upload limits, session timeout, and folder cleanup.",
        ),
    )
    page = page.replace("__FLASH__", flash_block)
    page = page.replace("__USERS_BLOCK__", users_block)
    page = page.replace("__MAX_SIZE__", str(config["max_pdf_size_mb"]))
    page = page.replace("__MAX_BATCH__", str(config["max_batch_size"]))
    page = page.replace("__TIMEOUT__", str(config["session_timeout_minutes"]))
    page = page.replace("__CU_UPLOADS__", str(config["cleanup_uploads_days"]))
    page = page.replace("__CU_OUTPUTS__", str(config["cleanup_outputs_days"]))
    page = page.replace("__CU_LOGS__", str(config["cleanup_logs_days"]))

    return page


def console_page_html(user: str, job_id: Optional[str] = None) -> str:
    try:
        payload = get_job_payload(job_id)
    except Exception:
        payload = {}

    console_text = build_full_console_text(payload) if payload else "No console data yet."
    payload_json = safe_js_json(payload)

    page = """
<!doctype html>
<html>
__HEAD__
<body>
<div class="shell">
    __TOPBAR__
    __HERO__

    <div class="card">
        <h2>Run Summary</h2>
        <div id="summaryBox" class="summary-list">No run found yet.</div>
        <br />
        <button onclick="copyEverything()">Copy Full Console</button>
        <button onclick="copyFieldResults()" class="secondary">Copy Field Results</button>
        <button onclick="copyRawOccurrences()" class="secondary">Copy Raw Occurrences</button>
        <button onclick="copyDebugLog()" class="secondary">Copy Debug Log</button>
    </div>

    <div class="card">
        <h2>Combined Console</h2>
        <div id="consoleBox" class="console">__CONSOLE_TEXT__</div>
    </div>
</div>

<script>
const payload = __PAYLOAD__;

let latestFieldResults = payload.field_results || [];
let latestRawOccurrences = payload.raw_occurrences || [];
let latestDebugLog = payload.debug_log || [];
let latestSummary = payload.summary || [];

if (payload.type === "batch" && payload.batch_results && payload.batch_results.length > 0) {
    const firstSuccess = payload.batch_results.find(item => item.status === "SUCCESS");
    if (firstSuccess) {
        latestFieldResults = firstSuccess.field_results || [];
        latestRawOccurrences = firstSuccess.raw_occurrences || [];
        latestDebugLog = firstSuccess.debug_log || [];
        latestSummary = payload.batch_summary || [];
    }
}

function rowsToTsv(rows) {
    return rows.map(row => row.map(cell => {
        if (cell === null || cell === undefined) return "";
        return String(cell).replaceAll("\\t", " ").replaceAll("\\n", " ").replaceAll("\\r", " ").trim();
    }).join("\\t")).join("\\n");
}

function buildFieldResultsTsv() {
    const rows = [[
        "PDF Code(s)", "Field", "Excel Cell", "Preview Value", "Status", "Difficulty", "Notes", "Component Details"
    ]];
    latestFieldResults.forEach(r => rows.push([
        r.expression, r.label, r.excel_cell, r.display_value, r.status, r.difficulty, r.notes, r.component_details
    ]));
    return rowsToTsv(rows);
}

function buildRawOccurrencesTsv() {
    const rows = [[
        "Code", "Selected", "Page", "X0", "Y0", "X1", "Y1", "Nearby Amount", "Confidence Score", "Note", "Nearby Context"
    ]];
    latestRawOccurrences.forEach(r => rows.push([
        r.code, r.selected ? "YES" : "NO", r.page_number,
        Number(r.x0 || 0).toFixed(2), Number(r.y0 || 0).toFixed(2),
        Number(r.x1 || 0).toFixed(2), Number(r.y1 || 0).toFixed(2),
        r.nearby_amount_text, r.confidence_score, r.note, r.nearby_context
    ]));
    return rowsToTsv(rows);
}

function buildFullConsoleText() {
    return document.getElementById("consoleBox").textContent || "";
}

async function copyText(text, label) {
    if (!text || !text.trim()) { alert("No " + label + " available to copy."); return; }
    await navigator.clipboard.writeText(text);
    alert(label + " copied to clipboard.");
}

function copyEverything() { copyText(buildFullConsoleText(), "Full Console"); }
function copyFieldResults() { copyText(buildFieldResultsTsv(), "Field Results"); }
function copyRawOccurrences() { copyText(buildRawOccurrencesTsv(), "Raw Occurrences"); }
function copyDebugLog() { copyText(latestDebugLog.join("\\n"), "Debug Log"); }

function render() {
    if (payload.type === "batch") {
        document.getElementById("summaryBox").textContent = (payload.batch_summary || []).join("\\n") || "No run found yet.";
    } else {
        document.getElementById("summaryBox").textContent = latestSummary.join("\\n") || "No run found yet.";
    }
}

render();
</script>
</body>
</html>
    """

    page = page.replace("__HEAD__", head_html("Console | Phillip Capital Risk Management"))
    page = page.replace("__TOPBAR__", topbar_html("console", user))
    page = page.replace(
        "__HERO__",
        hero_html(
            "Processing Console",
            "Review field results, raw helper-code occurrences, and debug logs from the latest workbook generation.",
        ),
    )
    page = page.replace("__PAYLOAD__", payload_json)
    page = page.replace("__CONSOLE_TEXT__", html.escape(console_text))

    return page


def outputs_page_html(user: str) -> str:
    # ---- SVG icons (inline so they inherit currentColor) ----
    folder_svg = (
        '<svg viewBox="0 0 20 16" fill="currentColor" aria-hidden="true">'
        '<path d="M0 2.5C0 1.12 1.12 0 2.5 0H7l2 2h8.5C18.88 2 20 3.12 20 4.5v9c0 1.38-1.12 2.5-2.5 2.5h-15C1.12 16 0 14.88 0 13.5v-11z"/>'
        '</svg>'
    )
    doc_svg = (
        '<svg viewBox="0 0 16 20" fill="currentColor" aria-hidden="true">'
        '<path d="M2 0C0.9 0 0 0.9 0 2v16c0 1.1 0.9 2 2 2h12c1.1 0 2-0.9 2-2V6L10 0H2zm8 1.5L14.5 6H10V1.5z"/>'
        '</svg>'
    )
    audit_svg = (
        '<svg viewBox="0 0 16 20" fill="currentColor" aria-hidden="true">'
        '<path d="M2 0C0.9 0 0 0.9 0 2v16c0 1.1 0.9 2 2 2h12c1.1 0 2-0.9 2-2V2c0-1.1-0.9-2-2-2H2zm2 6h8v1.5H4V6zm0 3h8v1.5H4V9zm0 3h5v1.5H4V12z"/>'
        '</svg>'
    )

    def fmt_time(mtime: float) -> str:
        return datetime.fromtimestamp(mtime).strftime("%I:%M %p").lstrip("0")

    def fmt_size(n: int) -> str:
        return get_file_size_label(n)

    # ---- Discover ZIPs (each = one batch) and read members ----
    batches = []
    for zip_path in OUTPUT_DIR.glob("*.zip"):
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                members = {n for n in zf.namelist() if not n.endswith("/")}
        except Exception:
            members = set()
        batches.append({
            "kind": "batch",
            "zip_path": zip_path,
            "members": members,
            "mtime": zip_path.stat().st_mtime,
            "files": [],   # populated below: list of {"xlsx_path", "audit_paths"}
        })

    # Map xlsx filename -> batch index (so xlsx files can be sorted into their folder)
    xlsx_to_batch = {}
    for idx, b in enumerate(batches):
        for name in b["members"]:
            xlsx_to_batch[name] = idx

    # ---- Index audits by xlsx stem (handles both "_audit" and "_audit_2" naming) ----
    audit_pattern = re.compile(r"^(.+?)_audit(?:_\d+)?$")
    audits_by_xlsx_stem: dict[str, list[Path]] = {}
    for audit_path in AUDIT_DIR.glob("*.txt"):
        m = audit_pattern.match(audit_path.stem)
        if m:
            audits_by_xlsx_stem.setdefault(m.group(1), []).append(audit_path)

    matched_audits: set[Path] = set()
    singles = []

    # ---- Walk xlsx files; classify each as batch member or single ----
    for xlsx_path in OUTPUT_DIR.glob("*.xlsx"):
        these_audits = audits_by_xlsx_stem.get(xlsx_path.stem, [])
        for a in these_audits:
            matched_audits.add(a)

        entry = {"xlsx_path": xlsx_path, "audit_paths": these_audits}

        if xlsx_path.name in xlsx_to_batch:
            batches[xlsx_to_batch[xlsx_path.name]]["files"].append(entry)
        else:
            singles.append({
                "kind": "single",
                "xlsx_path": xlsx_path,
                "audit_paths": these_audits,
                "mtime": xlsx_path.stat().st_mtime,
            })

    # ---- Orphan audits (xlsx no longer on disk) ----
    orphan_audits = []
    for audits in audits_by_xlsx_stem.values():
        for a in audits:
            if a not in matched_audits:
                orphan_audits.append({
                    "kind": "audit_only",
                    "audit_path": a,
                    "mtime": a.stat().st_mtime,
                })

    # ---- Combine + sort by mtime desc ----
    items = batches + singles + orphan_audits
    items.sort(key=lambda it: it["mtime"], reverse=True)

    # ---- Group by date ----
    today = date.today()
    yesterday = today - timedelta(days=1)

    def date_labels(d):
        if d == today:
            return "Today", d.strftime("%A, %B %d")
        if d == yesterday:
            return "Yesterday", d.strftime("%A, %B %d")
        if d.year == today.year:
            return d.strftime("%A, %B %d"), ""
        return d.strftime("%A, %B %d, %Y"), ""

    date_groups: list[tuple[date, list[dict]]] = []
    for it in items:
        d = datetime.fromtimestamp(it["mtime"]).date()
        if not date_groups or date_groups[-1][0] != d:
            date_groups.append((d, []))
        date_groups[-1][1].append(it)

    # ---- Renderers ----
    def render_file_row(entry):
        xlsx_path = entry["xlsx_path"]
        audit_paths = entry["audit_paths"]
        if not xlsx_path.exists():
            return ""
        stat = xlsx_path.stat()
        time_label = fmt_time(stat.st_mtime)
        size_label = fmt_size(stat.st_size)
        audit_btns = "".join(
            f'<a class="button-link secondary" href="/download-audit/{html.escape(a.name)}">Audit</a>'
            for a in audit_paths
        )
        return f"""
        <div class="file-row">
            <span class="file-icon xlsx">{doc_svg}</span>
            <div class="file-info">
                <div class="file-name">{html.escape(xlsx_path.name)}</div>
                <div class="file-meta">
                    <span>{html.escape(size_label)}</span>
                    <span class="dot">·</span>
                    <span>{html.escape(time_label)}</span>
                </div>
            </div>
            <div class="file-actions">
                <a class="button-link" href="/download-output/{html.escape(xlsx_path.name)}">Download</a>
                {audit_btns}
            </div>
        </div>
        """

    def render_batch(batch):
        zip_path = batch["zip_path"]
        files = batch["files"]
        n_files = len(files)
        n_audits = sum(len(e["audit_paths"]) for e in files)
        total_size = zip_path.stat().st_size
        for e in files:
            if e["xlsx_path"].exists():
                total_size += e["xlsx_path"].stat().st_size
            for a in e["audit_paths"]:
                if a.exists():
                    total_size += a.stat().st_size
        time_label = fmt_time(batch["mtime"])

        if not files:
            # ZIP with no surviving members
            inner = (
                '<p class="muted" style="margin:8px 0;">'
                'The individual workbooks for this batch are no longer on disk. '
                'You can still download the ZIP below.'
                '</p>'
            )
        else:
            inner = "".join(
                render_file_row(e)
                for e in sorted(files, key=lambda e: e["xlsx_path"].name.lower())
            )

        files_word = "workbook" if n_files == 1 else "workbooks"
        audits_word = "audit" if n_audits == 1 else "audits"

        return f"""
        <details class="folder-card" open>
            <summary>
                <span class="file-icon folder">{folder_svg}</span>
                <div class="folder-info">
                    <div class="folder-name">Batch &middot; {n_files} {files_word} &middot; {html.escape(time_label)}</div>
                    <div class="folder-meta">
                        <span>{n_audits} {audits_word}</span>
                        <span class="dot">·</span>
                        <span>{html.escape(fmt_size(total_size))}</span>
                        <span class="dot">·</span>
                        <code>{html.escape(zip_path.name)}</code>
                    </div>
                </div>
                <div class="folder-actions">
                    <a class="button-link orange" href="/download-output/{html.escape(zip_path.name)}">Download all (ZIP)</a>
                </div>
            </summary>
            <div class="folder-contents">
                {inner}
            </div>
        </details>
        """

    def render_single(item):
        xlsx_path = item["xlsx_path"]
        audit_paths = item["audit_paths"]
        if not xlsx_path.exists():
            return ""
        stat = xlsx_path.stat()
        time_label = fmt_time(stat.st_mtime)
        size_label = fmt_size(stat.st_size)
        audit_btns = "".join(
            f'<a class="button-link secondary" href="/download-audit/{html.escape(a.name)}">Audit</a>'
            for a in audit_paths
        )
        audit_word = "audit attached" if audit_paths else "no audit on file"
        return f"""
        <div class="single-card">
            <span class="file-icon xlsx">{doc_svg}</span>
            <div class="folder-info">
                <div class="folder-name">{html.escape(xlsx_path.name)}</div>
                <div class="folder-meta">
                    <span>Single workbook</span>
                    <span class="dot">·</span>
                    <span>{html.escape(size_label)}</span>
                    <span class="dot">·</span>
                    <span>{html.escape(time_label)}</span>
                    <span class="dot">·</span>
                    <span>{audit_word}</span>
                </div>
            </div>
            <div class="folder-actions">
                <a class="button-link orange" href="/download-output/{html.escape(xlsx_path.name)}">Download</a>
                {audit_btns}
            </div>
        </div>
        """

    def render_audit_only(item):
        audit_path = item["audit_path"]
        time_label = fmt_time(item["mtime"])
        size_label = fmt_size(audit_path.stat().st_size)
        return f"""
        <div class="single-card audit-only">
            <span class="file-icon audit">{audit_svg}</span>
            <div class="folder-info">
                <div class="folder-name">{html.escape(audit_path.name)}</div>
                <div class="folder-meta">
                    <span>Audit report only</span>
                    <span class="dot">·</span>
                    <span>workbook no longer on disk</span>
                    <span class="dot">·</span>
                    <span>{html.escape(size_label)}</span>
                    <span class="dot">·</span>
                    <span>{html.escape(time_label)}</span>
                </div>
            </div>
            <div class="folder-actions">
                <a class="button-link secondary" href="/download-audit/{html.escape(audit_path.name)}">Download audit</a>
            </div>
        </div>
        """

    # ---- Build body ----
    if not date_groups:
        outputs_html = """
        <div class="empty-state">
            <h3 style="margin-bottom:8px;">No outputs yet</h3>
            <p style="margin:0;">Generate a workbook from the Home page and it will appear here.</p>
        </div>
        """
    else:
        sections = []
        for grp_date, grp_items in date_groups:
            primary, secondary = date_labels(grp_date)
            secondary_html = (
                f'<span class="date-meta">{html.escape(secondary)}</span>'
                if secondary else ""
            )
            item_word = "item" if len(grp_items) == 1 else "items"
            divider = f"""
            <div class="date-divider">
                <span class="date-label">{html.escape(primary)}</span>
                {secondary_html}
                <span class="date-count">{len(grp_items)} {item_word}</span>
            </div>
            """
            cards = []
            for it in grp_items:
                if it["kind"] == "batch":
                    cards.append(render_batch(it))
                elif it["kind"] == "single":
                    cards.append(render_single(it))
                else:
                    cards.append(render_audit_only(it))
            sections.append(divider + "".join(cards))
        outputs_html = '<div class="outputs-list">' + "".join(sections) + "</div>"

    page = """
<!doctype html>
<html>
__HEAD__
<body>
<div class="shell">
    __TOPBAR__
    __HERO__

    <div class="card">
        <h2>Completed Workbooks &amp; Audits</h2>
        <p class="muted">Batches are shown as folders containing each generated workbook. Single runs appear as individual cards. Everything is grouped by the day it was produced.</p>
        __OUTPUTS_HTML__
    </div>
</div>
</body>
</html>
    """

    page = page.replace("__HEAD__", head_html("Outputs | Phillip Capital Risk Management"))
    page = page.replace("__TOPBAR__", topbar_html("outputs", user))
    page = page.replace(
        "__HERO__",
        hero_html(
            "Completed Outputs",
            "View and download generated Excel workbooks, batch ZIP files, and audit reports.",
        ),
    )
    page = page.replace("__OUTPUTS_HTML__", outputs_html)

    return page


# ============================================================
# BATCH WORKER (live progress)
# ============================================================

async def stream_save_pdf(pdf: UploadFile, batch_id: str, index: int, max_bytes: int) -> tuple[Optional[Path], Optional[str]]:
    if not pdf.filename or not pdf.filename.lower().endswith(".pdf"):
        return None, "Not a PDF file (only .pdf is accepted)."

    upload_name = safe_filename(pdf.filename)
    upload_path = UPLOAD_DIR / f"{batch_id}_{index}_{upload_name}"

    size = 0
    try:
        with open(upload_path, "wb") as f:
            while True:
                chunk = await pdf.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > max_bytes:
                    f.close()
                    upload_path.unlink(missing_ok=True)
                    return None, f"Exceeds the {max_bytes // (1024 * 1024)} MB size limit."
                f.write(chunk)
    except Exception as e:
        upload_path.unlink(missing_ok=True)
        return None, f"Could not save upload: {e}"

    if size == 0:
        upload_path.unlink(missing_ok=True)
        return None, "File is empty."

    return upload_path, None


def update_file_entry(job_id: str, index: int, **changes):
    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        if not job:
            return
        for entry in job["files"]:
            if entry["index"] == index:
                entry.update(changes)
                break


def run_batch_job(job_id: str):
    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        if not job:
            return
        queue = list(job.get("_queue", []))
        recalc = job.get("_recalc", False)
        user = job.get("user", "unknown")

    batch_results = []
    successful_output_paths: list[Path] = []
    combined_debug_lines = []
    template_status = get_template_status()
    template_name = template_status.get("filename", "")

    # Pre-record validation failures already present in files list as batch_results
    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        for entry in job["files"]:
            if entry["status"] == "failed":
                batch_results.append({
                    "status": "FAILED",
                    "job_id": f"{job_id}_{entry['index']}",
                    "original_filename": entry["filename"],
                    "error": entry.get("error", "Validation failed"),
                    "debug_log": [f"Validation failed: {entry.get('error', '')}"],
                })

    for index, original_filename, upload_path in queue:
        update_file_entry(job_id, index, status="processing")
        engine = CreditWorksheetEngine()

        try:
            output_path, field_results, occurrences_by_code, readability_info, summary = engine.generate_excel_from_pdf(
                pdf_path=upload_path,
                recalculate_with_excel=recalc,
            )

            audit_path, metrics = write_audit_report(
                user=user,
                original_filename=original_filename,
                output_path=output_path,
                template_name=template_name,
                field_results=field_results,
                readability_info=readability_info,
            )

            log_path = engine.save_debug_log(job_id, suffix=f"pdf_{index}")

            item_payload = {
                "status": "SUCCESS",
                "job_id": f"{job_id}_{index}",
                "original_filename": original_filename,
                "uploaded_pdf": str(upload_path),
                "output_path": str(output_path),
                "output_filename": output_path.name,
                "audit_filename": audit_path.name,
                "template": template_status,
                "readability": readability_info,
                "summary": summary,
                "field_results": engine.serialize_field_results(field_results),
                "raw_occurrences": engine.serialize_occurrences(occurrences_by_code),
                "metrics": metrics,
                "debug_log_filename": log_path.name,
                "debug_log_download_url": f"/download-log/{log_path.name}",
                "debug_log": engine.debug_lines,
            }

            successful_output_paths.append(output_path)
            batch_results.append(item_payload)

            update_file_entry(
                job_id,
                index,
                status="complete",
                output_filename=output_path.name,
                download_url=f"/download-output/{output_path.name}",
                audit_url=f"/download-audit/{audit_path.name}",
                fields_found=metrics["fields_found_label"],
                needs_review=metrics["needs_review"],
                valid_blanks=metrics["valid_blanks"],
            )

        except Exception as e:
            engine.log("[FATAL ERROR]")
            engine.log(traceback.format_exc())
            engine.save_debug_log(job_id, suffix=f"pdf_{index}_failed")

            batch_results.append({
                "status": "FAILED",
                "job_id": f"{job_id}_{index}",
                "original_filename": original_filename,
                "error": str(e),
                "debug_log": engine.debug_lines,
            })

            update_file_entry(job_id, index, status="failed", error=str(e))

        combined_debug_lines.extend([f"===== {original_filename} ====="])
        combined_debug_lines.extend(engine.debug_lines)
        combined_debug_lines.append("")

        with BATCH_JOBS_LOCK:
            job = BATCH_JOBS.get(job_id)
            if job:
                job["success_count"] = sum(1 for f in job["files"] if f["status"] == "complete")
                job["failure_count"] = sum(1 for f in job["files"] if f["status"] == "failed")

    success_count = sum(1 for item in batch_results if item.get("status") == "SUCCESS")
    failure_count = sum(1 for item in batch_results if item.get("status") == "FAILED")
    submitted_count = len(batch_results)

    batch_summary = [
        f"Batch ID: {job_id}",
        f"Created by: {user}",
        f"PDFs submitted: {submitted_count}",
        f"Successful outputs: {success_count}",
        f"Failed outputs: {failure_count}",
        "",
        "Output naming rule: each Excel file is named after its source PDF.",
        "An audit report (.txt) is written alongside each successful workbook.",
    ]

    payload = {
        "type": "batch" if submitted_count > 1 else "single",
        "job_id": job_id,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "submitted_count": submitted_count,
        "success_count": success_count,
        "failure_count": failure_count,
        "batch_summary": batch_summary,
        "batch_results": batch_results,
        "debug_log": combined_debug_lines,
    }

    if submitted_count == 1 and batch_results and batch_results[0].get("status") == "SUCCESS":
        payload.update(batch_results[0])
        payload["type"] = "single"

    try:
        save_json_debug(payload, job_id)
        store_job_result(job_id, payload)
        combined_log_path = LOG_DIR / f"credit_ws_batch_debug_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        combined_log_path.write_text(build_full_console_text(payload), encoding="utf-8")
    except Exception:
        pass

    zip_url = None
    single_url = None
    if len(successful_output_paths) > 1:
        zip_path = create_batch_zip(successful_output_paths, job_id)
        zip_url = f"/download-output/{zip_path.name}"
    elif len(successful_output_paths) == 1:
        single_url = f"/download-output/{successful_output_paths[0].name}"

    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        if job:
            job["status"] = "complete"
            job["success_count"] = success_count
            job["failure_count"] = failure_count
            job["zip_url"] = zip_url
            job["single_url"] = single_url
            job["console_url"] = f"/console?job_id={job_id}"
            job.pop("_queue", None)
            job.pop("_recalc", None)


# ============================================================
# ROUTES: AUTH
# ============================================================

@app.get("/login", response_class=HTMLResponse)
def login_get(request: Request):
    if current_user_or_none(request):
        return RedirectResponse("/", status_code=303)
    return HTMLResponse(login_page_html())


@app.post("/login")
def login_post(request: Request, username: str = Form(...), password: str = Form(...)):
    resolved = authenticate(username, password)
    if not resolved:
        return HTMLResponse(login_page_html("Invalid username or password."), status_code=401)

    token = create_session(resolved)
    timeout_seconds = load_config()["session_timeout_minutes"] * 60

    response = RedirectResponse("/", status_code=303)
    response.set_cookie(
        SESSION_COOKIE,
        token,
        httponly=True,
        samesite="lax",
        max_age=timeout_seconds,
        path="/",
    )
    return response


@app.get("/logout")
def logout(request: Request):
    destroy_session(request.cookies.get(SESSION_COOKIE))
    response = RedirectResponse("/login", status_code=303)
    response.delete_cookie(SESSION_COOKIE, path="/")
    return response


# ============================================================
# ROUTES: PAGES (protected)
# ============================================================

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return HTMLResponse(home_page_html(user))


@app.get("/home", response_class=HTMLResponse)
def home_alias(request: Request):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return HTMLResponse(home_page_html(user))


@app.get("/console", response_class=HTMLResponse)
def console(request: Request, job_id: Optional[str] = None):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return HTMLResponse(console_page_html(user, job_id))


@app.get("/outputs", response_class=HTMLResponse)
def outputs(request: Request):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return HTMLResponse(outputs_page_html(user))


@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request, msg: Optional[str] = None, t: Optional[str] = None):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return HTMLResponse(settings_page_html(user, flash=msg or "", flash_type=t or "ok"))


# ============================================================
# ROUTES: SETTINGS ACTIONS (protected)
# ============================================================

def settings_redirect(message: str, flash_type: str = "ok") -> RedirectResponse:
    from urllib.parse import quote
    return RedirectResponse(f"/settings?msg={quote(message)}&t={flash_type}", status_code=303)


@app.post("/settings/add-user")
def settings_add_user(request: Request, new_username: str = Form(...), new_password: str = Form(...)):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    ok, message = add_user(new_username, new_password, created_by=user)
    return settings_redirect(message, "ok" if ok else "err")


@app.post("/settings/delete-user")
def settings_delete_user(request: Request, username: str = Form(...)):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    ok, message = delete_user(username, current_user=user)
    return settings_redirect(message, "ok" if ok else "err")


@app.post("/settings/update-password")
def settings_update_password(
    request: Request,
    username: str = Form(...),
    new_password: str = Form(...),
):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    ok, message = update_password(username, new_password, actor=user)
    return settings_redirect(message, "ok" if ok else "err")


@app.post("/settings/limits")
def settings_limits(
    request: Request,
    max_pdf_size_mb: int = Form(...),
    max_batch_size: int = Form(...),
    session_timeout_minutes: int = Form(...),
):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    config = load_config()
    config["max_pdf_size_mb"] = max(1, min(int(max_pdf_size_mb), 500))
    config["max_batch_size"] = max(1, min(int(max_batch_size), 200))
    config["session_timeout_minutes"] = max(5, min(int(session_timeout_minutes), 1440))
    save_config(config)

    return settings_redirect("Upload limits saved.")


@app.post("/settings/retention")
def settings_retention(
    request: Request,
    cleanup_uploads_days: int = Form(...),
    cleanup_outputs_days: int = Form(...),
    cleanup_logs_days: int = Form(...),
):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    config = load_config()
    config["cleanup_uploads_days"] = max(0, min(int(cleanup_uploads_days), 365))
    config["cleanup_outputs_days"] = max(0, min(int(cleanup_outputs_days), 365))
    config["cleanup_logs_days"] = max(0, min(int(cleanup_logs_days), 365))
    save_config(config)

    return settings_redirect("Retention settings saved.")


@app.post("/settings/cleanup")
def settings_cleanup(request: Request, target: str = Form(...)):
    user = current_user_or_none(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    if target == "age":
        result = run_age_based_cleanup()
        total = sum(result.values())
        return settings_redirect(f"Age-based cleanup removed {total} file(s).")

    if target == "uploads":
        deleted = force_clear_directory(UPLOAD_DIR)
        return settings_redirect(f"Cleared {deleted} uploaded file(s).")

    if target == "logs":
        deleted = force_clear_directory(LOG_DIR)
        return settings_redirect(f"Cleared {deleted} log file(s).")

    if target == "outputs":
        deleted = force_clear_directory(OUTPUT_DIR) + force_clear_directory(AUDIT_DIR)
        return settings_redirect(f"Cleared {deleted} output/audit file(s).")

    return settings_redirect("Unknown cleanup target.", "err")


# ============================================================
# ROUTES: STATUS / API
# ============================================================

@app.get("/favicon.ico")
def favicon():
    logo_path = get_logo_path()
    if not logo_path:
        raise HTTPException(status_code=404, detail="Favicon not found.")
    return FileResponse(path=str(logo_path), media_type="image/png", filename=logo_path.name)


@app.get("/health")
def health(request: Request):
    require_api_user(request)
    return {
        "status": "running",
        "template": get_template_status(),
        "logo_url": get_logo_url(),
        "field_count": len(FIELD_DEFINITIONS),
    }


@app.get("/api/status")
def api_status(request: Request):
    require_api_user(request)
    return {
        "status": "running",
        "template": get_template_status(),
        "field_count": len(FIELD_DEFINITIONS),
        "expected_sheet": SHEET_NAME,
    }


@app.get("/api/latest-result")
def api_latest_result(request: Request):
    require_api_user(request)
    return get_job_payload()


@app.get("/api/result/{job_id}")
def api_result(request: Request, job_id: str):
    require_api_user(request)
    return get_job_payload(job_id)


# ============================================================
# ROUTES: BATCH PROCESSING (protected)
# ============================================================

@app.post("/start-batch")
async def start_batch(
    request: Request,
    pdfs: list[UploadFile] = File(...),
    recalculate_with_excel: bool = Form(False),
):
    user = require_api_user(request)
    config = load_config()

    if not pdfs:
        raise HTTPException(status_code=400, detail="Please upload at least one PDF.")

    if len(pdfs) > config["max_batch_size"]:
        raise HTTPException(
            status_code=400,
            detail=f"Too many files. The batch limit is {config['max_batch_size']} PDFs.",
        )

    batch_id = str(uuid.uuid4())
    max_bytes = config["max_pdf_size_mb"] * 1024 * 1024

    file_entries = []
    queue = []

    for index, pdf in enumerate(pdfs, start=1):
        original_filename = pdf.filename or f"file_{index}.pdf"
        upload_path, error = await stream_save_pdf(pdf, batch_id, index, max_bytes)

        if error:
            file_entries.append({
                "index": index,
                "filename": original_filename,
                "status": "failed",
                "error": error,
            })
        else:
            file_entries.append({
                "index": index,
                "filename": original_filename,
                "status": "queued",
            })
            queue.append((index, original_filename, upload_path))

    job = {
        "job_id": batch_id,
        "user": user,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "status": "running",
        "total": len(pdfs),
        "success_count": 0,
        "failure_count": sum(1 for f in file_entries if f["status"] == "failed"),
        "files": file_entries,
        "zip_url": None,
        "single_url": None,
        "console_url": f"/console?job_id={batch_id}",
        "_queue": queue,
        "_recalc": bool(recalculate_with_excel),
    }

    with BATCH_JOBS_LOCK:
        BATCH_JOBS[batch_id] = job
        if len(BATCH_JOBS) > 40:
            for key in list(BATCH_JOBS.keys())[:-40]:
                BATCH_JOBS.pop(key, None)

    if queue:
        thread = threading.Thread(target=run_batch_job, args=(batch_id,), daemon=True)
        thread.start()
    else:
        # Nothing valid to process — mark complete immediately.
        with BATCH_JOBS_LOCK:
            job["status"] = "complete"
            job.pop("_queue", None)
            job.pop("_recalc", None)

    return JSONResponse({"job_id": batch_id})


@app.get("/batch-status/{job_id}")
def batch_status(request: Request, job_id: str):
    require_api_user(request)

    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Batch job not found.")
        public = {k: v for k, v in job.items() if not k.startswith("_")}

    return JSONResponse(public)


# Backward-compatible synchronous endpoint (still protected).
@app.post("/generate-excel")
async def generate_excel(
    request: Request,
    pdfs: list[UploadFile] = File(...),
    recalculate_with_excel: bool = Form(False),
):
    user = require_api_user(request)
    config = load_config()

    batch_id = str(uuid.uuid4())
    batch_created_at = datetime.now().isoformat(timespec="seconds")

    if not pdfs:
        raise HTTPException(status_code=400, detail="Please upload at least one PDF.")

    if len(pdfs) > config["max_batch_size"]:
        raise HTTPException(status_code=400, detail=f"Batch limit is {config['max_batch_size']} PDFs.")

    max_bytes = config["max_pdf_size_mb"] * 1024 * 1024

    batch_results = []
    successful_output_paths: list[Path] = []
    combined_debug_lines = []
    template_status = get_template_status()
    template_name = template_status.get("filename", "")

    for index, pdf in enumerate(pdfs, start=1):
        engine = CreditWorksheetEngine()
        original_filename = pdf.filename or f"file_{index}.pdf"

        upload_path, error = await stream_save_pdf(pdf, batch_id, index, max_bytes)

        if error:
            batch_results.append({
                "status": "FAILED",
                "original_filename": original_filename,
                "error": error,
                "debug_log": [f"Validation failed: {error}"],
            })
            combined_debug_lines.extend([f"===== {original_filename} =====", f"Validation failed: {error}", ""])
            continue

        try:
            output_path, field_results, occurrences_by_code, readability_info, summary = engine.generate_excel_from_pdf(
                pdf_path=upload_path,
                recalculate_with_excel=recalculate_with_excel,
            )

            audit_path, metrics = write_audit_report(
                user=user,
                original_filename=original_filename,
                output_path=output_path,
                template_name=template_name,
                field_results=field_results,
                readability_info=readability_info,
            )

            log_path = engine.save_debug_log(batch_id, suffix=f"pdf_{index}")

            batch_results.append({
                "status": "SUCCESS",
                "original_filename": original_filename,
                "output_filename": output_path.name,
                "audit_filename": audit_path.name,
                "template": template_status,
                "readability": readability_info,
                "summary": summary,
                "field_results": engine.serialize_field_results(field_results),
                "raw_occurrences": engine.serialize_occurrences(occurrences_by_code),
                "metrics": metrics,
                "debug_log": engine.debug_lines,
            })
            successful_output_paths.append(output_path)

        except Exception as e:
            engine.log("[FATAL ERROR]")
            engine.log(traceback.format_exc())
            engine.save_debug_log(batch_id, suffix=f"pdf_{index}_failed")
            batch_results.append({
                "status": "FAILED",
                "original_filename": original_filename,
                "error": str(e),
                "debug_log": engine.debug_lines,
            })

        combined_debug_lines.extend([f"===== {original_filename} ====="])
        combined_debug_lines.extend(engine.debug_lines)
        combined_debug_lines.append("")

    success_count = sum(1 for item in batch_results if item.get("status") == "SUCCESS")
    failure_count = sum(1 for item in batch_results if item.get("status") == "FAILED")

    payload = {
        "type": "batch" if len(pdfs) > 1 else "single",
        "job_id": batch_id,
        "created_at": batch_created_at,
        "submitted_count": len(pdfs),
        "success_count": success_count,
        "failure_count": failure_count,
        "batch_summary": [
            f"Batch ID: {batch_id}",
            f"Created by: {user}",
            f"PDFs submitted: {len(pdfs)}",
            f"Successful outputs: {success_count}",
            f"Failed outputs: {failure_count}",
        ],
        "batch_results": batch_results,
        "debug_log": combined_debug_lines,
    }

    if len(pdfs) == 1 and batch_results and batch_results[0].get("status") == "SUCCESS":
        payload.update(batch_results[0])
        payload["type"] = "single"

    save_json_debug(payload, batch_id)
    store_job_result(batch_id, payload)

    if success_count == 0:
        raise HTTPException(status_code=500, detail="No PDFs could be processed.")

    if len(successful_output_paths) == 1:
        output_path = successful_output_paths[0]
        return FileResponse(
            path=str(output_path),
            filename=output_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"X-Job-ID": batch_id, "X-Success-Count": str(success_count), "X-Failure-Count": str(failure_count)},
        )

    zip_path = create_batch_zip(successful_output_paths, batch_id)
    return FileResponse(
        path=str(zip_path),
        filename=zip_path.name,
        media_type="application/zip",
        headers={"X-Job-ID": batch_id, "X-Success-Count": str(success_count), "X-Failure-Count": str(failure_count)},
    )


# ============================================================
# ROUTES: DOWNLOADS (protected)
# ============================================================

@app.get("/download-output/{filename}")
def download_output(request: Request, filename: str):
    require_api_user(request)

    safe_name = safe_filename(filename)
    path = OUTPUT_DIR / safe_name

    if not path.exists():
        raise HTTPException(status_code=404, detail="Output file not found.")

    if path.suffix.lower() == ".zip":
        media_type = "application/zip"
    elif path.suffix.lower() == ".xlsx":
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        media_type = "application/octet-stream"

    return FileResponse(path=str(path), filename=path.name, media_type=media_type)


@app.get("/download-audit/{filename}")
def download_audit(request: Request, filename: str):
    require_api_user(request)

    safe_name = safe_filename(filename)
    path = AUDIT_DIR / safe_name

    if not path.exists():
        raise HTTPException(status_code=404, detail="Audit file not found.")

    return FileResponse(path=str(path), filename=path.name, media_type="text/plain")


@app.get("/download-log/{filename}")
def download_log(request: Request, filename: str):
    require_api_user(request)

    safe_name = safe_filename(filename)
    path = LOG_DIR / safe_name

    if not path.exists():
        raise HTTPException(status_code=404, detail="Debug log not found.")

    return FileResponse(path=str(path), filename=path.name, media_type="text/plain")