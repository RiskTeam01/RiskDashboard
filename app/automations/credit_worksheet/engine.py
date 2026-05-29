import re
import shutil
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from app.config import SHEET_NAME, LOG_DIR, detect_backend_template
from app.automations.credit_worksheet.models import (
    WordItem, CodeOccurrence, FieldSpec, FieldResult,
)
from app.automations.credit_worksheet.fields import FIELD_DEFINITIONS

AMOUNT_PATTERN = re.compile(
    r"""
    ^\$?
    \(?
    -?
    (?:
        \d{1,3}(?:,\d{3})+ |
        \d+
    )
    (?:\.\d+)?
    \)?
    $
    """,
    re.VERBOSE,
)


class CreditWorksheetEngine:
    def __init__(self):
        self.debug_lines: list[str] = []

    def log(self, message: str):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.debug_lines.append(f"[{timestamp}] {message}")

    def save_debug_log(self, job_id: str, suffix: str = "") -> Path:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        safe_suffix = f"_{suffix}" if suffix else ""
        log_path = LOG_DIR / f"credit_ws_debug_{job_id}{safe_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        log_path.write_text("\n".join(self.debug_lines), encoding="utf-8")
        return log_path

    def get_field_specs(self) -> list[FieldSpec]:
        specs = []
        for item in FIELD_DEFINITIONS:
            expression = item["expression"].replace(" ", "")
            codes = expression.split("+")
            specs.append(
                FieldSpec(
                    expression=expression,
                    label=item["label"],
                    excel_cell=item["excel_cell"],
                    codes=codes,
                )
            )
        return specs

    def get_requested_codes(self, field_specs: list[FieldSpec]) -> list[str]:
        return sorted(
            set(code for spec in field_specs for code in spec.codes),
            key=lambda value: int(value),
        )

    def extract_pdf_field_results(self, pdf_path: Path):
        try:
            import fitz
        except ImportError:
            raise RuntimeError("PyMuPDF is not installed. Run: python -m pip install pymupdf")

        field_specs = self.get_field_specs()
        requested_codes = self.get_requested_codes(field_specs)

        self.log("=" * 100)
        self.log("Starting PDF extraction")
        self.log(f"PDF: {pdf_path}")
        self.log(f"Field definitions: {len(field_specs)}")
        self.log(f"Requested helper codes: {', '.join(requested_codes)}")
        self.log("=" * 100)

        doc = fitz.open(pdf_path)

        all_words: list[WordItem] = []
        page_count = len(doc)
        total_text_chars = 0
        pages_with_text = 0
        total_images = 0

        self.log(f"[PDF] Page count: {page_count}")

        for page_index in range(page_count):
            page = doc[page_index]
            text = page.get_text("text") or ""
            words_raw = page.get_text("words") or []
            images = page.get_images(full=True) or []

            text_chars = len(text.strip())
            total_text_chars += text_chars
            total_images += len(images)

            if text.strip():
                pages_with_text += 1

            self.log(
                f"[PAGE {page_index + 1}] "
                f"text chars={text_chars}, "
                f"words={len(words_raw)}, "
                f"images={len(images)}, "
                f"size={round(page.rect.width, 2)}x{round(page.rect.height, 2)}"
            )

            if page_index == 0:
                preview = text.strip().replace("\n", " ")
                preview = preview[:900] + ("..." if len(preview) > 900 else "")
                self.log(f"[PAGE 1 TEXT PREVIEW] {preview}")

            for word in words_raw:
                try:
                    x0, y0, x1, y1, word_text, block_no, line_no, word_no = word
                    clean_text = str(word_text).strip()
                    if clean_text:
                        all_words.append(
                            WordItem(
                                page_index=page_index,
                                page_number=page_index + 1,
                                x0=x0, y0=y0, x1=x1, y1=y1,
                                text=clean_text,
                                block_no=block_no,
                                line_no=line_no,
                                word_no=word_no,
                            )
                        )
                except Exception as e:
                    self.log(f"[WARN] Failed parsing word on page {page_index + 1}: {e}")

        doc.close()

        readability = self.calculate_readability(
            page_count=page_count,
            pages_with_text=pages_with_text,
            total_text_chars=total_text_chars,
            total_words=len(all_words),
            total_images=total_images,
        )

        readability_info = {
            "page_count": page_count,
            "pages_with_text": pages_with_text,
            "total_text_chars": total_text_chars,
            "total_words": len(all_words),
            "total_images": total_images,
            "readability": readability,
        }

        self.log("-" * 100)
        self.log("[READABILITY SUMMARY]")
        self.log(f"Pages with selectable text: {pages_with_text}/{page_count}")
        self.log(f"Total extracted text characters: {total_text_chars}")
        self.log(f"Total extracted word items: {len(all_words)}")
        self.log(f"Total embedded images detected: {total_images}")
        self.log(f"Readability result: {readability}")
        self.log("-" * 100)

        occurrences_by_code = self.find_requested_code_occurrences(
            all_words=all_words,
            requested_codes=requested_codes,
        )

        self.select_best_occurrences(occurrences_by_code)

        for code in requested_codes:
            occurrences = occurrences_by_code.get(code, [])
            self.log("")
            self.log(f"[CODE {code}] occurrences found: {len(occurrences)}")
            if not occurrences:
                self.log("  - Missing from selectable text.")
                continue
            for occ in occurrences:
                selected_text = "YES" if occ.selected else "NO"
                self.log(
                    f"  - selected={selected_text}, "
                    f"page={occ.page_number}, "
                    f"x={round(occ.x0, 2)}-{round(occ.x1, 2)}, "
                    f"y={round(occ.y0, 2)}-{round(occ.y1, 2)}, "
                    f"amount='{occ.nearby_amount_text}', "
                    f"score={occ.confidence_score}, "
                    f"note='{occ.note}', "
                    f"context='{occ.nearby_context}'"
                )

        field_results: list[FieldResult] = []
        for spec in field_specs:
            result = self.evaluate_field(spec, occurrences_by_code)
            field_results.append(result)
            self.log("")
            self.log(f"[FIELD] {spec.expression} = {spec.label} -> {spec.excel_cell}")
            self.log(f"  Preview value: {result.display_value}")
            self.log(f"  Status: {result.status}")
            self.log(f"  Difficulty: {result.difficulty}")
            self.log(f"  Notes: {result.notes}")
            self.log(f"  Details: {result.component_details}")

        summary = self.build_overall_summary(
            field_results=field_results,
            occurrences_by_code=occurrences_by_code,
            requested_codes=requested_codes,
            readability_info=readability_info,
        )

        self.log("")
        self.log("=" * 100)
        self.log("[OVERALL PDF AUTOMATION READINESS]")
        for line in summary:
            self.log(line)
        self.log("=" * 100)

        return field_results, occurrences_by_code, readability_info, summary

    def calculate_readability(
        self,
        page_count: int,
        pages_with_text: int,
        total_text_chars: int,
        total_words: int,
        total_images: int,
    ) -> str:
        if total_words == 0 and total_text_chars == 0:
            if total_images > 0:
                return "HARD - likely scanned/image PDF; OCR would be needed"
            return "HARD - no selectable text detected"
        avg_words_per_page = total_words / max(page_count, 1)
        if pages_with_text == page_count and avg_words_per_page >= 30:
            return "EASY - selectable text detected on all pages"
        if pages_with_text > 0 and total_words >= 50:
            return "MEDIUM - some selectable text detected; layout needs testing"
        return "HARD - very limited selectable text detected"

    def normalize_code_token(self, text: str) -> str:
        cleaned = str(text).strip()
        for ch in "[](){}":
            cleaned = cleaned.replace(ch, "")
        cleaned = cleaned.replace(".", "").strip()
        return cleaned

    def is_amount_like(self, text: str) -> bool:
        cleaned = str(text).strip()
        if not cleaned:
            return False
        possible_code = (
            cleaned.replace("$", "").replace("-", "").replace("(", "").replace(")", "").strip()
        )
        if "," not in cleaned and possible_code.isdigit() and len(possible_code) <= 5:
            return False
        return bool(AMOUNT_PATTERN.match(cleaned))

    def amount_to_number(self, text: str):
        if not text:
            return None
        cleaned = str(text).strip()
        negative = cleaned.startswith("(") and cleaned.endswith(")")
        cleaned = cleaned.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
        try:
            value = float(cleaned) if "." in cleaned else int(cleaned)
            return -value if negative else value
        except Exception:
            return None

    def format_number_for_display(self, value: Optional[float]) -> str:
        if value is None:
            return ""
        if abs(value - int(value)) < 0.000001:
            return f"{int(value):,}"
        return f"{value:,.2f}"

    def find_requested_code_occurrences(
        self,
        all_words: list[WordItem],
        requested_codes: list[str],
    ) -> dict[str, list[CodeOccurrence]]:
        requested_set = set(requested_codes)
        occurrences_by_code = {code: [] for code in requested_codes}
        for word in all_words:
            normalized = self.normalize_code_token(word.text)
            if normalized not in requested_set:
                continue
            nearby_amount, nearby_context, note, score = self.find_left_side_amount_preview(
                target_word=word,
                all_words=all_words,
            )
            occurrence = CodeOccurrence(
                code=normalized,
                page_number=word.page_number,
                x0=word.x0, y0=word.y0, x1=word.x1, y1=word.y1,
                nearby_amount_text=nearby_amount,
                nearby_context=nearby_context,
                note=note,
                confidence_score=score,
            )
            occurrences_by_code[normalized].append(occurrence)
        return occurrences_by_code

    def find_left_side_amount_preview(
        self,
        target_word: WordItem,
        all_words: list[WordItem],
    ):
        y_tolerance = 5.75
        max_left_distance = 260
        same_page_words = [w for w in all_words if w.page_index == target_word.page_index]
        same_row_left_words = [
            w for w in same_page_words
            if w.x1 < target_word.x0
            and abs(w.y_center - target_word.y_center) <= y_tolerance
            and (target_word.x0 - w.x1) <= max_left_distance
        ]
        same_row_left_words.sort(key=lambda w: w.x1, reverse=True)
        nearby_context = " ".join(w.text for w in same_row_left_words[:12])
        for candidate in same_row_left_words:
            candidate_text = candidate.text.strip()
            if self.is_amount_like(candidate_text):
                score = 80
                distance = target_word.x0 - candidate.x1
                if distance <= 80:
                    score += 10
                elif distance <= 150:
                    score += 5
                if "," in candidate_text:
                    score += 10
                return (
                    candidate_text,
                    nearby_context,
                    "Nearby amount-like value found to the left",
                    score,
                )
        if nearby_context:
            return (
                "",
                nearby_context,
                "Code found, but no amount-like value was found on the same row. Treat as valid blank unless later rules say otherwise.",
                55,
            )
        return (
            "",
            "",
            "Code found, but left-side row area appears blank or not selectable. Treat as valid blank unless later rules say otherwise.",
            50,
        )

    def select_best_occurrences(self, occurrences_by_code: dict[str, list[CodeOccurrence]]):
        for code, occurrences in occurrences_by_code.items():
            if not occurrences:
                continue
            with_amount = [occ for occ in occurrences if occ.nearby_amount_text]
            if len(occurrences) == 1:
                occurrences[0].selected = True
                continue
            if len(with_amount) == 1:
                with_amount[0].selected = True
                with_amount[0].note += " | Auto-selected because it is the only occurrence with a nearby amount."
                continue
            if len(with_amount) > 1:
                with_amount.sort(key=lambda occ: occ.confidence_score, reverse=True)
                with_amount[0].selected = True
                with_amount[0].note += " | Auto-selected by highest score, but multiple amount matches exist. Needs review."
                continue
            occurrences.sort(key=lambda occ: occ.confidence_score, reverse=True)
            occurrences[0].selected = True
            occurrences[0].note += " | Auto-selected blank candidate, but multiple blank/code-reference matches exist. Needs review."

    def evaluate_field(
        self,
        spec: FieldSpec,
        occurrences_by_code: dict[str, list[CodeOccurrence]],
    ) -> FieldResult:
        component_values = []
        component_notes = []
        missing_codes = []
        review_needed = False
        all_components_blank_or_found = True

        for code in spec.codes:
            occurrences = occurrences_by_code.get(code, [])
            if not occurrences:
                missing_codes.append(code)
                all_components_blank_or_found = False
                component_notes.append(f"{code}: MISSING")
                continue
            selected = next((occ for occ in occurrences if occ.selected), occurrences[0])
            if len(occurrences) > 1:
                review_needed = True
            if selected.nearby_amount_text:
                numeric_value = self.amount_to_number(selected.nearby_amount_text)
                if numeric_value is None:
                    review_needed = True
                    component_notes.append(f"{code}: found '{selected.nearby_amount_text}' but could not convert")
                else:
                    component_values.append(numeric_value)
                    component_notes.append(f"{code}: {selected.nearby_amount_text} ({selected.location_text})")
            else:
                component_notes.append(f"{code}: blank ({selected.location_text})")

        if missing_codes:
            return FieldResult(
                expression=spec.expression, label=spec.label, excel_cell=spec.excel_cell,
                display_value="", numeric_value=None, should_write_blank=True,
                status="MISSING", difficulty="Hard",
                notes=f"Missing code(s): {', '.join(missing_codes)}",
                component_details=" | ".join(component_notes),
            )

        if len(spec.codes) == 1:
            code = spec.codes[0]
            occurrences = occurrences_by_code.get(code, [])
            selected = next((occ for occ in occurrences if occ.selected), occurrences[0])
            if selected.nearby_amount_text:
                numeric_value = self.amount_to_number(selected.nearby_amount_text)
                status = "READY"
                difficulty = "Easy" if len(occurrences) == 1 else "Medium"
                notes = "Single code found with value."
                if len(occurrences) > 1:
                    notes += " Multiple occurrences exist, so final extraction should use position rules."
                return FieldResult(
                    expression=spec.expression, label=spec.label, excel_cell=spec.excel_cell,
                    display_value=selected.nearby_amount_text, numeric_value=numeric_value,
                    should_write_blank=False, status=status, difficulty=difficulty,
                    notes=notes, component_details=" | ".join(component_notes),
                )
            return FieldResult(
                expression=spec.expression, label=spec.label, excel_cell=spec.excel_cell,
                display_value="", numeric_value=None, should_write_blank=True,
                status="VALID BLANK",
                difficulty="Easy" if len(occurrences) == 1 else "Medium",
                notes="Code found with no amount. Per rule, this is a valid blank.",
                component_details=" | ".join(component_notes),
            )

        if component_values:
            total = sum(component_values)
            display_value = self.format_number_for_display(total)
            if review_needed:
                status = "NEEDS REVIEW"
                difficulty = "Medium"
                notes = "Combined field calculated, but one or more component codes had multiple occurrences."
            else:
                status = "READY"
                difficulty = "Medium"
                notes = "Combined field calculated from component helper codes."
            return FieldResult(
                expression=spec.expression, label=spec.label, excel_cell=spec.excel_cell,
                display_value=display_value, numeric_value=total, should_write_blank=False,
                status=status, difficulty=difficulty, notes=notes,
                component_details=" | ".join(component_notes),
            )

        if all_components_blank_or_found:
            return FieldResult(
                expression=spec.expression, label=spec.label, excel_cell=spec.excel_cell,
                display_value="", numeric_value=None, should_write_blank=True,
                status="VALID BLANK", difficulty="Medium",
                notes="All component codes were found, but all appear blank. Per rule, this is valid blank.",
                component_details=" | ".join(component_notes),
            )

        return FieldResult(
            expression=spec.expression, label=spec.label, excel_cell=spec.excel_cell,
            display_value="", numeric_value=None, should_write_blank=True,
            status="NEEDS REVIEW", difficulty="Medium/Hard",
            notes="Combined field could not be confidently calculated.",
            component_details=" | ".join(component_notes),
        )

    def build_overall_summary(
        self,
        field_results: list[FieldResult],
        occurrences_by_code: dict[str, list[CodeOccurrence]],
        requested_codes: list[str],
        readability_info: dict,
    ) -> list[str]:
        total_fields = len(field_results)
        ready_count = sum(1 for r in field_results if r.status == "READY")
        blank_count = sum(1 for r in field_results if r.status == "VALID BLANK")
        review_count = sum(1 for r in field_results if r.status == "NEEDS REVIEW")
        missing_count = sum(1 for r in field_results if r.status == "MISSING")
        codes_found = sum(1 for code in requested_codes if occurrences_by_code.get(code))
        multiple_codes = [code for code in requested_codes if len(occurrences_by_code.get(code, [])) > 1]

        lines = [
            f"Required Excel yellow input cells checked: {total_fields}",
            f"Unique helper codes requested: {len(requested_codes)}",
            f"Unique helper codes found: {codes_found}/{len(requested_codes)}",
            f"READY fields: {ready_count}",
            f"VALID BLANK fields: {blank_count}",
            f"NEEDS REVIEW fields: {review_count}",
            f"MISSING fields: {missing_count}",
            f"Codes with multiple occurrences: {', '.join(multiple_codes) if multiple_codes else 'None'}",
            f"PDF readability: {readability_info.get('readability')}",
            f"Extracted word count: {readability_info.get('total_words')}",
            f"Extracted text characters: {readability_info.get('total_text_chars')}",
            "",
        ]

        if missing_count == 0:
            lines.append("Excel generation readiness: GOOD.")
            lines.append("Reason: all requested helper codes were found or valid blanks were detected.")
        else:
            lines.append("Excel generation readiness: BLOCKED.")
            lines.append("Reason: one or more helper codes were missing.")

        lines.append("")
        lines.append("Blank fields will be written as true blank Excel cells.")
        lines.append("The backend Excel template is copied first; the original template is not modified.")
        return lines

    def serialize_occurrences(self, occurrences_by_code: dict[str, list[CodeOccurrence]]) -> list[dict]:
        rows = []
        for code in sorted(occurrences_by_code.keys(), key=lambda value: int(value)):
            for occ in occurrences_by_code[code]:
                item = asdict(occ)
                item["location_text"] = occ.location_text
                rows.append(item)
        return rows

    def serialize_field_results(self, field_results: list[FieldResult]) -> list[dict]:
        return [asdict(result) for result in field_results]

    def create_output_excel_path(self, pdf_path: Path) -> Path:
        from app.config import OUTPUT_DIR
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        safe_stem = re.sub(r"[^A-Za-z0-9_\- .]+", "_", pdf_path.stem).strip(" ._")
        if not safe_stem:
            safe_stem = "Completed_Credit_WS"
        output_path = OUTPUT_DIR / f"{safe_stem}.xlsx"
        if not output_path.exists():
            return output_path
        counter = 2
        while True:
            candidate = OUTPUT_DIR / f"{safe_stem}_{counter}.xlsx"
            if not candidate.exists():
                return candidate
            counter += 1

    def generate_excel_from_pdf(self, pdf_path: Path, recalculate_with_excel: bool = False):
        template_path = detect_backend_template()
        self.log("[TEMPLATE] Auto-detected backend template:")
        self.log(str(template_path))
        self.validate_template(template_path)

        field_results, occurrences_by_code, readability_info, summary = self.extract_pdf_field_results(pdf_path)

        missing = [r for r in field_results if r.status == "MISSING"]
        if missing:
            missing_text = ", ".join(f"{r.expression} ({r.label})" for r in missing)
            raise RuntimeError(f"Excel was not generated because these fields are missing: {missing_text}")

        output_path = self.create_output_excel_path(pdf_path)

        self.log("[EXCEL] Copying backend template to output:")
        self.log(f"  Template: {template_path}")
        self.log(f"  Output:   {output_path}")

        shutil.copy2(template_path, output_path)
        self.write_results_to_excel(output_path, field_results)

        if recalculate_with_excel:
            self.recalculate_workbook_with_excel_com(output_path)
        else:
            self.log("[EXCEL] Skipped Excel COM recalculation. Workbook formulas will calculate when opened in Excel.")

        self.log("[SUCCESS] Completed Excel workbook created.")
        self.log(str(output_path))

        return output_path, field_results, occurrences_by_code, readability_info, summary

    def validate_template(self, template_path: Path):
        self.log("[TEMPLATE] Validating workbook template.")
        try:
            wb = load_workbook(template_path, read_only=True, data_only=False)
            sheetnames = wb.sheetnames
            wb.close()
        except Exception as e:
            raise RuntimeError(f"Template could not be opened as a valid Excel workbook: {e}")
        if SHEET_NAME not in sheetnames:
            raise RuntimeError(f"Expected sheet '{SHEET_NAME}' not found. Found sheets: {sheetnames}")
        self.log(f"[TEMPLATE] Validated. Sheetnames: {sheetnames}")

    def write_results_to_excel(self, output_path: Path, field_results: list[FieldResult]):
        self.log("[EXCEL] Opening copied workbook for writing.")
        wb = load_workbook(output_path)
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(f"Expected sheet '{SHEET_NAME}' not found. Found sheets: {wb.sheetnames}")
        ws = wb[SHEET_NAME]
        self.log("[EXCEL] Writing extracted values into 24 yellow input cells.")
        for result in field_results:
            cell = result.excel_cell
            if result.status == "MISSING":
                raise RuntimeError(f"Cannot write missing field: {result.expression} {result.label}")
            if result.should_write_blank:
                ws[cell].value = None
                self.log(f"  - {cell}: {result.expression} {result.label} -> BLANK")
            else:
                ws[cell].value = result.numeric_value
                self.log(f"  - {cell}: {result.expression} {result.label} -> {result.display_value}")
        self.force_excel_formula_recalculation_on_open(wb)
        wb.save(output_path)
        self.log(f"[EXCEL] Saved workbook: {output_path}")

    def force_excel_formula_recalculation_on_open(self, wb):
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"
            self.log("[EXCEL] Formula recalculation flags set for next Excel open.")
        except Exception as e:
            self.log(f"[WARN] Could not set formula recalculation flags: {e}")

    def recalculate_workbook_with_excel_com(self, output_path: Path):
        self.log("[EXCEL COM] Attempting to open Excel and force recalculation.")
        try:
            import win32com.client
        except Exception as e:
            self.log(f"[WARN] pywin32/win32com is not available. Skipping Excel COM recalculation. Details: {e}")
            return
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(output_path))
            excel.CalculateFullRebuild()
            wb.Save()
            self.log("[EXCEL COM] Workbook recalculated and saved successfully.")
        except Exception as e:
            self.log(f"[WARN] Excel COM recalculation failed. Workbook will still recalculate when opened. Details: {e}")
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=True)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
