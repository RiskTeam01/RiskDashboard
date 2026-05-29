"""
Manages per-customer-per-year workbooks that combine monthly credit worksheet
sheets with an accumulating Net Capital sheet.

Workbook stored at: NET_CAPITAL_DIR/{safe_name}_{year}.xlsx
Sheets inside:
  - "March 2026", "July 2026", ...   (one per credit worksheet run)
  - "Net Capital 2026"               (accumulating month columns)
"""
import re
from copy import copy
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import NET_CAPITAL_DIR, NET_CAPITAL_SHEET_NAME


# ── helpers ──────────────────────────────────────────────────────────────────

def _safe_customer_name(name: str) -> str:
    name = re.sub(r"[^A-Za-z0-9 ]+", "", name).strip()
    name = re.sub(r"\s+", "_", name)
    return name[:40] or "Customer"


def customer_workbook_path(customer_name: str, year: int) -> Path:
    return NET_CAPITAL_DIR / f"{_safe_customer_name(customer_name)}_{year}.xlsx"


def open_or_create(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    return wb


def _parse_amount(text: str):
    if not text:
        return None
    s = str(text).strip()
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    try:
        val = float(s) if "." in s else int(s)
        return -val if neg else val
    except Exception:
        return None


def _copy_sheet(src_ws, target_wb: Workbook, new_name: str) -> Worksheet:
    """Deep-copy a worksheet from one workbook into another."""
    tgt = target_wb.create_sheet(title=new_name)
    for row in src_ws.iter_rows():
        for cell in row:
            tc = tgt.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                tc.font = copy(cell.font)
                tc.border = copy(cell.border)
                tc.fill = copy(cell.fill)
                tc.number_format = cell.number_format
                tc.protection = copy(cell.protection)
                tc.alignment = copy(cell.alignment)
    for mr in src_ws.merged_cells.ranges:
        tgt.merge_cells(str(mr))
    for col, dim in src_ws.column_dimensions.items():
        tgt.column_dimensions[col].width = dim.width
    for row, dim in src_ws.row_dimensions.items():
        tgt.row_dimensions[row].height = dim.height
    return tgt


MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
MONTH_COLUMNS = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]


def credit_sheet_name(month: int, year: int) -> str:
    return f"{MONTH_NAMES[month - 1]} {year}"


def net_capital_sheet_name(year: int) -> str:
    return f"Net Capital {year}"


def _clear_net_capital_placeholders(ws, row_map: dict) -> int:
    """Blank the FOCUS-code placeholder values in the data region of a freshly
    created Net Capital sheet. Clears all month columns (C–N) for each mapped
    row, the period-end date row (5), and the C1 company-name cell."""
    cleared = 0
    rows_to_clear = set(row_map.keys()) | {5}
    for row_num in rows_to_clear:
        for col in MONTH_COLUMNS:
            cell = ws[f"{col}{row_num}"]
            if cell.value is not None:
                cell.value = None
                cleared += 1
    if ws["C1"].value is not None:
        ws["C1"].value = None
        cleared += 1
    return cleared


# ── credit worksheet sheet ────────────────────────────────────────────────────

def add_credit_month_sheet(
    wb: Workbook,
    month: int,
    year: int,
    credit_template_path: Path,
    credit_template_sheet: str,
    field_results: list,
    log,
) -> str:
    """Copy credit WS template sheet into wb as 'March 2026' etc., write values."""
    sheet_name = credit_sheet_name(month, year)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        log(f"[CW] Replaced existing sheet '{sheet_name}'")

    src_wb = load_workbook(credit_template_path)
    src_ws = src_wb[credit_template_sheet]
    ws = _copy_sheet(src_ws, wb, sheet_name)
    src_wb.close()
    log(f"[CW] Created sheet '{sheet_name}' from credit template")

    for result in field_results:
        if result.status == "MISSING":
            raise RuntimeError(f"Cannot write missing field: {result.expression} {result.label}")
        if result.should_write_blank:
            ws[result.excel_cell].value = None
            log(f"[CW]   {result.excel_cell}: {result.expression} -> BLANK")
        else:
            ws[result.excel_cell].value = result.numeric_value
            log(f"[CW]   {result.excel_cell}: {result.expression} -> {result.display_value}")

    return sheet_name


# ── net capital sheet ─────────────────────────────────────────────────────────

def add_or_update_net_capital_sheet(
    wb: Workbook,
    month: int,
    year: int,
    company_name: str,
    date_text: Optional[str],
    net_capital_template_path: Optional[Path],
    occurrences_by_code: dict,
    row_map: dict,
    log,
) -> str:
    """Add or update the 'Net Capital YEAR' sheet with the new month column."""
    nc_name = net_capital_sheet_name(year)
    col = MONTH_COLUMNS[month - 1]
    log(f"[NC] Filling '{nc_name}' column {col} ({MONTH_NAMES[month-1]})")

    if nc_name not in wb.sheetnames:
        if net_capital_template_path and net_capital_template_path.exists():
            src_wb = load_workbook(net_capital_template_path)
            src_sheet = next(
                (n for n in src_wb.sheetnames if "net capital" in n.lower()),
                src_wb.sheetnames[0],
            )
            ws = _copy_sheet(src_wb[src_sheet], wb, nc_name)
            src_wb.close()
            # The template carries FOCUS code numbers in the data cells as backend
            # reference markers. Blank them out on the output so only real extracted
            # values appear. Done once, at sheet creation, so later monthly updates
            # never wipe previously-filled columns.
            cleared = _clear_net_capital_placeholders(ws, row_map)
            log(f"[NC] Created '{nc_name}' from template; cleared {cleared} placeholder cell(s)")
        else:
            ws = wb.create_sheet(title=nc_name)
            log(f"[NC] No Net Capital template — created blank '{nc_name}'")
    else:
        ws = wb[nc_name]
        log(f"[NC] Updating existing '{nc_name}'")

    # Company name into C1 (replace literal "13" placeholder once)
    existing_c1 = ws["C1"].value
    if company_name and (existing_c1 is None or str(existing_c1).strip() in ("", "13")):
        ws["C1"] = company_name
        log(f"[NC] Set C1 = '{company_name}'")

    # Period end date into row 5
    if date_text:
        ws[f"{col}5"] = date_text

    written = 0
    for row_num, code in row_map.items():
        occs = occurrences_by_code.get(code, [])
        if not occs:
            continue
        sel = next((o for o in occs if o.selected), occs[0])
        val = _parse_amount(sel.nearby_amount_text) if sel.nearby_amount_text else None
        ws[f"{col}{row_num}"] = val
        if val is not None:
            written += 1

    log(f"[NC] Wrote {written} values into column {col}")
    _reorder_sheets(wb, year)
    return nc_name


def _reorder_sheets(wb: Workbook, year: int):
    """Put 'Net Capital YEAR' first, then credit month sheets in calendar order."""
    nc = net_capital_sheet_name(year)
    month_order = {credit_sheet_name(m, year): m for m in range(1, 13)}

    desired: list[str] = []
    if nc in wb.sheetnames:
        desired.append(nc)
    for m in range(1, 13):
        name = credit_sheet_name(m, year)
        if name in wb.sheetnames:
            desired.append(name)
    # Any sheets for other years: keep them at the end, sorted
    others = [s for s in wb.sheetnames if s not in desired]
    desired.extend(others)

    wb._sheets.sort(key=lambda ws: desired.index(ws.title) if ws.title in desired else len(desired))


def set_recalc_on_open(wb: Workbook):
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass


def autofit_columns(wb: Workbook, min_width: float = 10.0, max_width: float = 50.0):
    """Widen any column that is too narrow to display its contents."""
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            best = ws.column_dimensions[col_letter].width or 0
            for cell in col_cells:
                if cell.value is None:
                    continue
                # Estimate rendered width: numeric values need room for formatting
                val = cell.value
                if isinstance(val, (int, float)):
                    fmt = cell.number_format or ""
                    # Use the formatted string length as a proxy
                    try:
                        if "%" in fmt:
                            text = f"{val * 100:.2f}%"
                        elif "," in fmt or "#" in fmt:
                            text = f"{val:,.2f}"
                        else:
                            text = str(val)
                    except Exception:
                        text = str(val)
                else:
                    text = str(val)
                best = max(best, len(text) + 2)
            # Only widen, never shrink — keeps intentionally narrow label columns tidy
            current = ws.column_dimensions[col_letter].width or 0
            if best > current:
                ws.column_dimensions[col_letter].width = min(max(best, min_width), max_width)
