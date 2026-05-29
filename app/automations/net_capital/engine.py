import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from app.config import (
    TEMPLATE_DIR, NET_CAPITAL_DIR, NET_CAPITAL_TEMPLATE_FILENAME,
    NET_CAPITAL_SHEET_NAME,
)
from app.automations.net_capital.fields import (
    NET_CAPITAL_ROW_MAP, MONTH_COLUMNS, MONTH_NAMES,
)

DATE_PATTERNS = [
    re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})"),   # MM/DD/YY or MM/DD/YYYY
    re.compile(r"(\d{4})-(\d{2})-(\d{2})"),           # YYYY-MM-DD
]


def _extract_month_from_date_text(text: str) -> Optional[int]:
    """Return 1-12 month number from a date string, or None."""
    for pat in DATE_PATTERNS:
        m = pat.search(text)
        if m:
            groups = m.groups()
            if len(groups[0]) == 4:
                # YYYY-MM-DD
                return int(groups[1])
            else:
                # MM/DD/YY or MM/DD/YYYY — month is first group
                month = int(groups[0])
                if 1 <= month <= 12:
                    return month
    return None


def _extract_text_near_code(
    all_words: list,
    code: str,
    y_tolerance: float = 5.75,
    max_left_distance: float = 260,
) -> str:
    """Extract the nearby text to the RIGHT of a code token (for text fields like names/dates)."""
    results = []
    for word in all_words:
        normalized = str(word.text).strip().replace(".", "").replace("(", "").replace(")", "")
        if normalized != code:
            continue
        same_page = [w for w in all_words if w.page_index == word.page_index]
        # Grab text to the right on the same row
        row_right = [
            w for w in same_page
            if w.x0 > word.x1
            and abs(w.y_center - word.y_center) <= y_tolerance
        ]
        row_right.sort(key=lambda w: w.x0)
        if row_right:
            results.append(" ".join(w.text for w in row_right[:10]))
    return results[0] if results else ""


def detect_net_capital_template() -> Optional[Path]:
    preferred = TEMPLATE_DIR / NET_CAPITAL_TEMPLATE_FILENAME
    if preferred.exists():
        return preferred
    # fallback: any xlsx whose name contains "net capital" (any spacing/underscore/case)
    for path in TEMPLATE_DIR.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        normalized = path.name.lower().replace("_", " ").replace("-", " ")
        if "net capital" in normalized:
            return path
    return None


def get_or_create_net_capital_workbook(customer_id: str, year: int) -> Path:
    NET_CAPITAL_DIR.mkdir(parents=True, exist_ok=True)
    path = NET_CAPITAL_DIR / f"NetCapital_{customer_id[:8]}_{year}.xlsx"
    if path.exists():
        return path
    template = detect_net_capital_template()
    if template is None:
        raise FileNotFoundError(
            f"Net Capital template not found in {TEMPLATE_DIR}. "
            f"Place '{NET_CAPITAL_TEMPLATE_FILENAME}' in the templates folder."
        )
    shutil.copy2(template, path)
    return path


class NetCapitalEngine:
    def __init__(self):
        self.debug_lines: list[str] = []

    def log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.debug_lines.append(f"[{ts}] {msg}")

    def extract_company_name(self, all_words: list) -> str:
        """Extract text near code 13 (company name) to the RIGHT of the code."""
        text = _extract_text_near_code(all_words, "13")
        self.log(f"[CODE 13] Company name raw: '{text}'")
        return text.strip()

    def extract_period_end_date(self, all_words: list) -> Optional[str]:
        """Extract text near code 25 (period ending date)."""
        text = _extract_text_near_code(all_words, "25")
        self.log(f"[CODE 25] Period end date raw: '{text}'")
        if not text:
            return None
        m = re.search(r"\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2}", text)
        if m:
            return m.group(0)
        return text.strip() or None

    def determine_month(self, date_text: str) -> Optional[int]:
        month = _extract_month_from_date_text(date_text)
        self.log(f"[MONTH] Determined month={month} from '{date_text}'")
        return month

    def fill_net_capital_workbook(
        self,
        workbook_path: Path,
        month: int,
        company_name: str,
        occurrences_by_code: dict,
        date_text: Optional[str] = None,
    ) -> bool:
        """Fill the month column in the workbook. Returns True on success."""
        col_letter = MONTH_COLUMNS[month - 1]
        self.log(f"[NC] Filling month={month} ({MONTH_NAMES[month-1]}) column={col_letter}")

        wb = load_workbook(workbook_path)
        if NET_CAPITAL_SHEET_NAME not in wb.sheetnames:
            self.log(f"[NC ERROR] Sheet '{NET_CAPITAL_SHEET_NAME}' not found. Available: {wb.sheetnames}")
            wb.close()
            return False

        ws = wb[NET_CAPITAL_SHEET_NAME]

        # Write company name in C1 if blank
        if company_name and not ws["C1"].value:
            ws["C1"] = company_name
            self.log(f"[NC] Set C1 = '{company_name}'")

        # Write period end date in the date row (row 5) for this month column
        if date_text:
            ws[f"{col_letter}5"] = date_text
            self.log(f"[NC] Set {col_letter}5 = '{date_text}'")

        fields_written = 0
        for row_num, code in NET_CAPITAL_ROW_MAP.items():
            cell_addr = f"{col_letter}{row_num}"
            occs = occurrences_by_code.get(code, [])
            if not occs:
                self.log(f"[NC] {cell_addr} code={code}: MISSING")
                continue
            selected = next((o for o in occs if o.selected), occs[0])
            if selected.nearby_amount_text:
                value = _parse_amount(selected.nearby_amount_text)
                ws[cell_addr] = value
                self.log(f"[NC] {cell_addr} code={code} -> {value}")
                fields_written += 1
            else:
                ws[cell_addr] = None
                self.log(f"[NC] {cell_addr} code={code} -> BLANK")

        wb.save(workbook_path)
        wb.close()
        self.log(f"[NC] Saved. Fields written: {fields_written}")
        return True

    def build_occurrences(self, all_words: list) -> dict:
        """Scan all_words for the Net Capital row codes using the shared left-side
        amount logic from the credit worksheet engine."""
        from app.automations.credit_worksheet.engine import CreditWorksheetEngine

        cw = CreditWorksheetEngine()
        nc_codes = sorted(set(NET_CAPITAL_ROW_MAP.values()), key=lambda v: int(v))
        occurrences = cw.find_requested_code_occurrences(
            all_words=all_words,
            requested_codes=nc_codes,
        )
        cw.select_best_occurrences(occurrences)
        found = sum(1 for c in nc_codes if occurrences.get(c))
        self.log(f"[NC] Scanned {len(nc_codes)} net-capital row codes; found {found} in PDF.")
        return occurrences

    def run(
        self,
        all_words: list,
        occurrences_by_code: Optional[dict],
        customer_id: str,
        customer_name: str,
    ) -> Optional[Path]:
        """Run Net Capital extraction and return the workbook path, or None on failure."""
        self.log("[NC] Starting Net Capital extraction")

        # The credit worksheet only scans its own codes, so build a fresh
        # occurrence map for the Net Capital row codes from all_words.
        occurrences_by_code = self.build_occurrences(all_words)

        date_text = self.extract_period_end_date(all_words)
        if not date_text:
            self.log("[NC] Could not extract period end date (code 25). Skipping Net Capital.")
            return None

        month = self.determine_month(date_text)
        if not month:
            self.log(f"[NC] Could not parse month from '{date_text}'. Skipping.")
            return None

        year_match = re.search(r"\d{4}", date_text)
        year = int(year_match.group(0)) if year_match else datetime.now().year
        if year < 100:
            year += 2000

        workbook_path = get_or_create_net_capital_workbook(customer_id, year)
        self.log(f"[NC] Workbook path: {workbook_path}")

        success = self.fill_net_capital_workbook(
            workbook_path=workbook_path,
            month=month,
            company_name=customer_name,
            occurrences_by_code=occurrences_by_code,
            date_text=date_text,
        )

        if success:
            self.log(f"[NC] Complete. File: {workbook_path.name}")
            return workbook_path
        return None


def _parse_amount(text: str):
    if not text:
        return None
    cleaned = str(text).strip()
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = cleaned.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    try:
        value = float(cleaned) if "." in cleaned else int(cleaned)
        return -value if negative else value
    except Exception:
        return None
