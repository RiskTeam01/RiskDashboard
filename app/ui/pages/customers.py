import html
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.customers import load_customers, get_customer
from app.config import NET_CAPITAL_DIR, AUDIT_DIR
from app.ui.components import head_html, topbar_html, hero_html


# Net Capital rows we surface in the per-year financial summary.
SUMMARY_METRICS = [
    {
        "row": 7, "label": "Total Equity", "key": "total_equity",
        "desc": "Ownership equity reported on the FOCUS filing (line 3500).",
        "noun": "the firm's capital base",
        "up_msg": "is growing, strengthening the firm's financial foundation and its "
                  "capacity to absorb losses.",
        "down_msg": "is shrinking, which reduces the firm's cushion against losses and "
                    "can be an early warning sign worth watching.",
        "flat_msg": "is holding steady, indicating a stable capital position with little "
                    "movement between filings.",
    },
    {
        "row": 47, "label": "Net Capital", "key": "net_capital",
        "desc": "Regulatory net capital after all required deductions (line 3750).",
        "noun": "regulatory net capital",
        "up_msg": "is improving, moving the firm further above its required regulatory "
                  "minimum — a healthy sign for compliance.",
        "down_msg": "is declining; continued erosion would pressure the firm's compliance "
                    "buffer and should be monitored closely.",
        "flat_msg": "is stable, keeping the firm's regulatory liquidity roughly level "
                    "across these filings.",
    },
    {
        "row": 53, "label": "Excess Net Capital", "key": "excess_net_capital",
        "desc": "Net capital above the required minimum — the cushion (line 3910).",
        "noun": "the safety cushion above the regulatory minimum",
        "up_msg": "is widening, giving the firm more room before it risks a net-capital "
                  "deficiency — a clearly positive trend.",
        "down_msg": "is thinning; if this continues the firm moves closer to its "
                    "net-capital floor, the most important early-warning signal here.",
        "flat_msg": "is steady, holding the firm's safety margin roughly constant.",
    },
]
_MONTH_COLUMNS = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
QUARTERLY_MONTHS = [("March", 3), ("June", 6), ("September", 9), ("December", 12)]


def _abbrev(val) -> str:
    """Compact money label: 1,250,000 -> 1.25M."""
    if val is None:
        return "—"
    sign = "-" if val < 0 else ""
    a = abs(val)
    if a >= 1_000_000_000:
        return f"{sign}{a/1_000_000_000:.2f}B"
    if a >= 1_000_000:
        return f"{sign}{a/1_000_000:.2f}M"
    if a >= 1_000:
        return f"{sign}{a/1_000:.1f}K"
    return f"{sign}{a:,.0f}"


def _full_money(val) -> str:
    if val is None:
        return "—"
    return f"${val:,.0f}"


def _year_from_wb_name(wb_name: str) -> int:
    m = re.search(r"_(\d{4})\.xlsx$", wb_name or "")
    return int(m.group(1)) if m else 0


def _build_metric_analysis_html(metric, series) -> str:
    """Build a data-driven analysis of how this metric has moved over time."""
    label = metric["label"]

    if not series:
        return (f'<h3 class="modal-title">{html.escape(label)}</h3>'
                f'<div class="mm-empty">No data has been recorded for {html.escape(label)} '
                f'yet. Once filings are processed, this view will break down how it is trending.</div>')

    first_m, first_v = series[0]
    last_m, last_v = series[-1]
    n = len(series)
    vals = [v for _, v in series]
    high_v = max(vals); high_m = series[vals.index(high_v)][0]
    low_v = min(vals); low_m = series[vals.index(low_v)][0]
    avg_v = sum(vals) / n

    net = last_v - first_v
    pct = (net / abs(first_v) * 100) if first_v else 0.0
    direction = "up" if net > 0.0001 else ("down" if net < -0.0001 else "flat")

    # Verdict bucket
    if direction == "flat" or abs(pct) < 0.5:
        verdict, vcls = "Stable", "verdict-neutral"
        msg = metric["flat_msg"]
    elif direction == "up":
        verdict = "Positive" if pct >= 5 else "Improving"
        vcls = "verdict-good"
        msg = metric["up_msg"]
    else:
        verdict = "Negative" if pct <= -5 else "Caution"
        vcls = "verdict-bad" if pct <= -5 else "verdict-warn"
        msg = metric["down_msg"]

    # Headline
    arrow = "&#9650;" if net > 0 else ("&#9660;" if net < 0 else "&#8211;")
    sign = "+" if net >= 0 else ""
    if n == 1:
        headline_note = f'Only one reading on file ({_MONTH_ABBR[first_m-1]}). A trend will appear once more filings are processed.'
    else:
        headline_note = (f'{sign}{_full_money(net)} ({sign}{pct:.1f}%) from {_MONTH_ABBR[first_m-1]} '
                         f'to {_MONTH_ABBR[last_m-1]} across {n} readings.')

    # Narrative
    narrative = f'Over the tracked period, {metric["noun"]} {msg}'

    # Period-over-period movement rows
    move_rows = ""
    if n >= 2:
        rows = []
        for i in range(1, n):
            pm, pv = series[i-1]
            cm, cv = series[i]
            d = cv - pv
            dp = (d / abs(pv) * 100) if pv else 0.0
            up = d >= 0
            a = "&#9650;" if up else "&#9660;"
            cls = "mv-up" if up else "mv-down"
            s = "+" if d >= 0 else ""
            rows.append(f"""
            <div class="mm-qrow">
                <span class="mm-qlabel">{_MONTH_ABBR[pm-1]} &rarr; {_MONTH_ABBR[cm-1]}</span>
                <span class="mm-qval">{_full_money(cv)}</span>
                <span class="mm-qdelta {cls}">{a} {s}{_abbrev(d)} ({s}{dp:.1f}%)</span>
            </div>
            """)
        move_rows = f"""
        <div class="modal-section">
            <div class="modal-h">Movement period over period</div>
            <div class="mm-quarters">{"".join(rows)}</div>
        </div>
        """

    # Stats grid
    stats = f"""
    <div class="mm-stats">
        <div class="mm-stat"><div class="mm-stat-label">Latest</div><div class="mm-stat-val">{_abbrev(last_v)}</div><div class="mm-stat-sub">{_MONTH_ABBR[last_m-1]}</div></div>
        <div class="mm-stat"><div class="mm-stat-label">Period High</div><div class="mm-stat-val">{_abbrev(high_v)}</div><div class="mm-stat-sub">{_MONTH_ABBR[high_m-1]}</div></div>
        <div class="mm-stat"><div class="mm-stat-label">Period Low</div><div class="mm-stat-val">{_abbrev(low_v)}</div><div class="mm-stat-sub">{_MONTH_ABBR[low_m-1]}</div></div>
        <div class="mm-stat"><div class="mm-stat-label">Average</div><div class="mm-stat-val">{_abbrev(avg_v)}</div><div class="mm-stat-sub">{n} readings</div></div>
    </div>
    """

    return f"""
    <div class="mm-head">
        <h3 class="modal-title">{html.escape(label)}</h3>
        <span class="mm-verdict {vcls}">{verdict}</span>
    </div>
    <div class="mm-headline">
        <span class="mm-current">{_full_money(last_v)}</span>
        <span class="mm-change {('mv-up' if net>=0 else 'mv-down')}">{arrow} {sign}{pct:.1f}%</span>
    </div>
    <div class="mm-subnote">{headline_note}</div>
    <div class="modal-section">
        <div class="modal-h">What this means</div>
        <p>{narrative}</p>
    </div>
    {move_rows}
    {stats}
    <div class="modal-foot">Figures read live from the Net Capital sheet (FOCUS {metric['desc'].split('(')[-1].rstrip(').')}). Quarter-end months are shown darker on the card chart.</div>
    """


def _metric_analysis_json(year_metrics: dict) -> str:
    """Build {f'{year}|{key}': html} for every year/metric pairing."""
    import json
    payload = {}
    for yr, metrics in year_metrics.items():
        for m in SUMMARY_METRICS:
            payload[f"{yr}|{m['key']}"] = _build_metric_analysis_html(m, metrics[m["key"]])
    return json.dumps(payload)


def read_year_metrics(wb_path: Path, year: int) -> dict:
    """Read the three summary rows from a workbook's Net Capital sheet.
    Returns {metric_key: [(month_num, value), ...]} for months that have data."""
    result = {m["key"]: [] for m in SUMMARY_METRICS}
    if not wb_path or not wb_path.exists():
        return result
    try:
        wb = load_workbook(wb_path, data_only=True)
    except Exception:
        return result
    nc_name = f"Net Capital {year}"
    if nc_name not in wb.sheetnames:
        wb.close()
        return result
    ws = wb[nc_name]
    for metric in SUMMARY_METRICS:
        for idx, col in enumerate(_MONTH_COLUMNS):
            v = ws[f"{col}{metric['row']}"].value
            if isinstance(v, (int, float)):
                result[metric["key"]].append((idx + 1, float(v)))
    wb.close()
    return result


# ── list page ─────────────────────────────────────────────────────────────────

def customers_list_page_html(user: str) -> str:
    data = load_customers()
    customers = sorted(
        [{"id": cid, **cust} for cid, cust in data.items()],
        key=lambda c: c.get("name", "").lower(),
    )

    if not customers:
        body = """
        <div class="empty-state">
            <h3 style="margin-bottom:8px;">No customers yet</h3>
            <p style="margin:0;">Customer accounts are created automatically when you process a PDF.
            The company name is detected from helper code 13 on the cover page.</p>
        </div>
        """
    else:
        rows = []
        for c in customers:
            reports = c.get("reports", [])
            last_run = ""
            if reports:
                latest = max(reports, key=lambda r: r.get("created_at", ""))
                try:
                    dt = datetime.fromisoformat(latest["created_at"])
                    last_run = dt.strftime("%b %d, %Y %I:%M %p").lstrip("0")
                except Exception:
                    last_run = latest.get("created_at", "")
            report_count = len(reports)
            word = "report" if report_count == 1 else "reports"
            rows.append(f"""
            <a class="customer-card" href="/customers/{html.escape(c['id'])}">
                <div class="customer-name">{html.escape(c['name'])}</div>
                <div class="customer-meta">
                    <span>{report_count} {word}</span>
                    {f'<span class="dot">·</span><span>Last run: {html.escape(last_run)}</span>' if last_run else ''}
                </div>
            </a>
            """)
        body = '<div class="customer-list">' + "".join(rows) + "</div>"

    return f"""
<!doctype html>
<html>
{head_html("Customers | Phillip Capital Risk Management")}
<body>
<div class="shell">
    {topbar_html("customers", user)}
    {hero_html("Customer Accounts", "Each customer account collects all reports processed for that firm.")}
    <div class="card">
        <h2>All Customers</h2>
        <p class="muted">Accounts are created automatically from helper code 13 (company name) on each PDF run.</p>
        {body}
    </div>
</div>
</body>
</html>
    """


# ── detail page builders ──────────────────────────────────────────────────────

def _latest_value(series):
    return series[-1][1] if series else None


def _first_value(series):
    return series[0][1] if series else None


def _delta_badge(series):
    """Returns (html_badge, since_note) comparing latest vs first reading."""
    if not series or len(series) < 2:
        if series:
            m = series[0][0]
            return ('<span class="metric-delta delta-flat">first reading</span>',
                    f'as of {_MONTH_ABBR[m-1]}')
        return ("", "")
    first_m, first_v = series[0]
    last_m, last_v = series[-1]
    delta = last_v - first_v
    pct = (delta / abs(first_v) * 100) if first_v else 0.0
    up = delta >= 0
    arrow = "&#9650;" if up else "&#9660;"
    cls = "delta-up" if up else "delta-down"
    badge = f'<span class="metric-delta {cls}">{arrow} {abs(pct):.1f}%</span>'
    sign = "+" if delta >= 0 else ""
    note = f'{sign}{_abbrev(delta)} since {_MONTH_ABBR[first_m-1]}'
    return badge, note


def _quarter_stepper_html(sheets_received, yr, current_year, current_month):
    steps = []
    for i, (mname, mnum) in enumerate(QUARTERLY_MONTHS):
        received = f"{mname} {yr}" in sheets_received
        is_future = (yr > current_year) or (yr == current_year and mnum > current_month)
        if received:
            state, icon, qlabel = "step-done", "&#10003;", "Filed"
        elif is_future:
            state, icon, qlabel = "step-future", "&middot;", "Upcoming"
        else:
            state, icon, qlabel = "step-missing", "!", "Missing"
        connector = '<div class="step-line"></div>' if i < len(QUARTERLY_MONTHS) - 1 else ""
        steps.append(f"""
        <div class="step {state}">
            <div class="step-dot">{icon}</div>
            <div class="step-name">Q{i+1} · {mname}</div>
            <div class="step-status">{qlabel}</div>
        </div>
        {connector}
        """)
    return f'<div class="stepper">{"".join(steps)}</div>'


def _metric_card_html(metric, series, yr):
    last_v = _latest_value(series)
    modal_id = f"{yr}|{metric['key']}"
    if not series:
        return f"""
        <div class="metric-card empty">
            <div class="metric-top">
                <span class="metric-label">{metric['label']}</span>
                <button class="metric-info" onclick="openMetricModal('{modal_id}')" title="Analyze {html.escape(metric['label'])}">?</button>
            </div>
            <div class="metric-current muted">No data yet</div>
            <div class="metric-desc">{html.escape(metric['desc'])}</div>
        </div>
        """
    badge, note = _delta_badge(series)

    vals = [v for _, v in series]
    vmax, vmin = max(vals), min(vals)
    span = (vmax - vmin) or abs(vmax) or 1
    bars = []
    for m_num, v in series:
        h = 20 + 80 * ((v - vmin) / span) if span else 60
        bcls = "bar-pos" if v >= 0 else "bar-neg"
        is_q = m_num in (3, 6, 9, 12)
        bars.append(f"""
        <div class="bar-col" title="{_MONTH_ABBR[m_num-1]}: {_full_money(v)}">
            <div class="bar-val">{_abbrev(v)}</div>
            <div class="bar {bcls} {'bar-q' if is_q else ''}" style="height:{h:.0f}%;"></div>
            <div class="bar-month">{_MONTH_ABBR[m_num-1]}</div>
        </div>
        """)

    return f"""
    <div class="metric-card">
        <div class="metric-top">
            <span class="metric-label">{metric['label']}</span>
            <button class="metric-info" onclick="openMetricModal('{modal_id}')" title="Analyze {html.escape(metric['label'])}">?</button>
        </div>
        <div class="metric-current">{_full_money(last_v)}</div>
        <div class="metric-deltarow">{badge}<span class="metric-since">{note}</span></div>
        <div class="bar-chart">{"".join(bars)}</div>
    </div>
    """


def customer_detail_page_html(user: str, customer_id: str) -> str:
    customer = get_customer(customer_id)
    if not customer:
        return f"""
<!doctype html><html>
{head_html("Not Found | Phillip Capital Risk Management")}
<body><div class="shell">
{topbar_html("customers", user)}
<div class="card"><h2>Customer not found</h2>
<p><a href="/customers">Back to Customers</a></p></div>
</div></body></html>
        """

    reports = sorted(
        customer.get("reports", []),
        key=lambda r: r.get("created_at", ""),
        reverse=True,
    )

    # Group by year
    years: dict[int, dict] = {}
    for r in reports:
        yr = _year_from_wb_name(r.get("output_filename", ""))
        if yr not in years:
            years[yr] = {"workbook_filename": r.get("output_filename", ""), "runs": []}
        years[yr]["runs"].append(r)

    all_years = sorted(years.keys(), reverse=True)
    current_year = datetime.now().year
    current_month = datetime.now().month

    # Pre-read metrics for each year (also used for account overview)
    year_metrics = {yr: read_year_metrics(NET_CAPITAL_DIR / years[yr]["workbook_filename"], yr)
                    for yr in all_years}

    # ── Account overview KPIs ────────────────────────────────────────────────
    total_filings = len(reports)
    years_tracked = len(all_years)
    latest_filing_label = reports[0].get("period_label", "—") if reports else "—"

    newest_year = all_years[0] if all_years else None
    latest_nc = None
    latest_equity = None
    if newest_year is not None:
        latest_nc = _latest_value(year_metrics[newest_year]["net_capital"])
        latest_equity = _latest_value(year_metrics[newest_year]["total_equity"])

    # Compliance for newest year
    compliance_txt = "—"
    compliance_cls = "kpi-neutral"
    if newest_year is not None:
        sheets_recv = {r.get("credit_sheet", "") for r in years[newest_year]["runs"]}
        due = [m for (mn, m) in QUARTERLY_MONTHS
               if not ((newest_year > current_year) or (newest_year == current_year and m > current_month))]
        filed = [m for (mn, m) in QUARTERLY_MONTHS if f"{mn} {newest_year}" in sheets_recv]
        if due:
            pct = int(round(len(filed) / len(due) * 100))
            compliance_txt = f"{pct}%"
            compliance_cls = "kpi-good" if pct == 100 else ("kpi-warn" if pct >= 50 else "kpi-bad")

    overview_html = f"""
    <div class="overview">
        <div class="kpi">
            <div class="kpi-label">Total Filings</div>
            <div class="kpi-value">{total_filings}</div>
            <div class="kpi-sub">across {years_tracked} year{'s' if years_tracked != 1 else ''}</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Latest Filing</div>
            <div class="kpi-value sm">{html.escape(latest_filing_label)}</div>
            <div class="kpi-sub">most recent period</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Net Capital</div>
            <div class="kpi-value sm">{_full_money(latest_nc)}</div>
            <div class="kpi-sub">latest reported</div>
        </div>
        <div class="kpi">
            <div class="kpi-label">Total Equity</div>
            <div class="kpi-value sm">{_full_money(latest_equity)}</div>
            <div class="kpi-sub">latest reported</div>
        </div>
        <div class="kpi {compliance_cls}">
            <div class="kpi-label">Quarterly Compliance</div>
            <div class="kpi-value">{compliance_txt}</div>
            <div class="kpi-sub">{newest_year if newest_year else ''} filings on time</div>
        </div>
    </div>
    """

    # ── Year panels ──────────────────────────────────────────────────────────
    year_cards_html = ""
    for idx, yr in enumerate(all_years):
        info = years[yr]
        wb_name = info["workbook_filename"]
        wb_path = NET_CAPITAL_DIR / wb_name if wb_name else None
        wb_exists = wb_path and wb_path.exists()
        dl_btn = (
            f'<a class="button-link orange" href="/download-net-capital/{html.escape(wb_name)}">Download Workbook</a>'
            if wb_exists else
            '<span class="muted" style="font-size:12px;">File missing</span>'
        )

        sheets_received = {r.get("credit_sheet", "") for r in info["runs"]}
        metrics = year_metrics[yr]

        # Missing quarters check
        missing_count = 0
        for mname, mnum in QUARTERLY_MONTHS:
            is_future = (yr > current_year) or (yr == current_year and mnum > current_month)
            if not is_future and f"{mname} {yr}" not in sheets_received:
                missing_count += 1
        any_missing = missing_count > 0

        stepper = _quarter_stepper_html(sheets_received, yr, current_year, current_month)
        metric_cards = "".join(_metric_card_html(m, metrics[m["key"]], yr) for m in SUMMARY_METRICS)

        # Inline header KPI
        hdr_nc = _latest_value(metrics["net_capital"])
        hdr_kpi = (
            f'<span class="header-kpi">Net Capital <b>{_abbrev(hdr_nc)}</b></span>'
            if hdr_nc is not None else ''
        )

        # Run timeline
        run_items = []
        for r in info["runs"]:
            try:
                dt = datetime.fromisoformat(r["created_at"])
                date_str = dt.strftime("%b %d, %Y · %I:%M %p").replace("· 0", "· ")
            except Exception:
                date_str = r.get("created_at", "")
            orig = r.get("original_filename", "")
            period = r.get("period_label", "")
            audit = r.get("audit_filename", "")
            audit_btn = (
                f'<a class="timeline-audit" href="/download-audit/{html.escape(audit)}">Audit</a>'
                if audit and (AUDIT_DIR / audit).exists() else ""
            )
            run_items.append(f"""
            <div class="timeline-item run-row" data-period="{html.escape(period.lower())}" data-file="{html.escape(orig.lower())}">
                <div class="timeline-dot"></div>
                <div class="timeline-body">
                    <div class="timeline-head">
                        <span class="timeline-period">{html.escape(period) or 'Unknown period'}</span>
                        {audit_btn}
                    </div>
                    <div class="timeline-file">{html.escape(orig)}</div>
                    <div class="timeline-date">{html.escape(date_str)}</div>
                </div>
            </div>
            """)

        collapsed = "" if idx == 0 else "collapsed"  # newest year open by default
        alert_chip = (
            f'<span class="header-alert">! {missing_count} quarter{"s" if missing_count != 1 else ""} missing</span>'
            if any_missing else
            '<span class="header-ok">&#10003; On track</span>'
        )

        year_cards_html += f"""
        <div class="year-card {collapsed}" data-year="{yr}">
            <div class="year-header" onclick="toggleYear(this)">
                <div class="year-header-left">
                    <span class="chevron">&#9656;</span>
                    <span class="year-num">{yr}</span>
                    <span class="year-runs-count">{len(info['runs'])} filing{'s' if len(info['runs']) != 1 else ''}</span>
                    {alert_chip}
                    {hdr_kpi}
                </div>
                <span onclick="event.stopPropagation();">{dl_btn}</span>
            </div>
            <div class="year-body">
                <div class="section-block">
                    <div class="section-title">Quarterly Filing Tracker
                        <span class="section-hint">FOCUS reports are due for Q1–Q4. Missing past quarters are flagged.</span>
                    </div>
                    {stepper}
                </div>
                <div class="section-block">
                    <div class="section-title">Financial Position
                        <span class="section-hint">Pulled live from the Net Capital sheet. Bars highlight quarter-end months.</span>
                    </div>
                    <div class="metrics-summary">{metric_cards}</div>
                </div>
                <div class="section-block">
                    <div class="section-title">Filing History
                        <span class="section-hint">{len(info['runs'])} PDF{'s' if len(info['runs']) != 1 else ''} processed into this workbook.</span>
                    </div>
                    <div class="timeline">{"".join(run_items)}</div>
                </div>
            </div>
        </div>
        """

    if not year_cards_html:
        year_cards_html = '<div class="empty-state"><p style="margin:0;">No reports on file yet.</p></div>'
        overview_html = ""

    return f"""
<!doctype html>
<html>
{head_html(f"{html.escape(customer['name'])} | Phillip Capital Risk Management")}
{_detail_styles()}
<body>
<div class="shell">
    {topbar_html("customers", user)}
    {hero_html(customer['name'], "Customer account — quarterly filings, net capital position, and full processing history.")}
    <div class="card">
        <div style="margin-bottom:16px;">
            <a href="/customers" style="color:var(--pc-blue);text-decoration:none;font-size:0.875rem;font-weight:600;">&larr; All Customers</a>
        </div>

        {overview_html}

        <div class="controls-bar">
            <input type="text" id="searchInput" placeholder="Search filings by name or period (e.g. March, 2025)&hellip;" oninput="applyFilters()">
            <div class="filter-pills" id="yearPills">
                <span class="pill active" data-year="all" onclick="setPill(this)">All Years</span>
                {"".join(f'<span class="pill" data-year="{yr}" onclick="setPill(this)">{yr}</span>' for yr in all_years)}
            </div>
            <span class="pill expand-toggle" onclick="toggleAll(this)" data-state="mixed">Expand all</span>
        </div>

        <div id="yearList">
            {year_cards_html}
        </div>
    </div>
</div>

<div id="metricModal" class="modal-overlay" onclick="closeMetricModal(event)">
    <div class="modal-box" onclick="event.stopPropagation();">
        <button class="modal-close" onclick="closeMetricModal(event)" title="Close">&times;</button>
        <div class="modal-kicker">Metric Analysis</div>
        <div id="mmBody"></div>
    </div>
</div>

<script>
const METRIC_ANALYSIS = {_metric_analysis_json(year_metrics)};

let activeYear = "all";

function openMetricModal(id) {{
    const html = METRIC_ANALYSIS[id];
    if (!html) return;
    document.getElementById("mmBody").innerHTML = html;
    document.getElementById("metricModal").classList.add("show");
    document.body.style.overflow = "hidden";
}}

function closeMetricModal(event) {{
    if (event) event.stopPropagation();
    document.getElementById("metricModal").classList.remove("show");
    document.body.style.overflow = "";
}}

document.addEventListener("keydown", function(e) {{
    if (e.key === "Escape") closeMetricModal();
}});

function toggleYear(el) {{
    el.closest(".year-card").classList.toggle("collapsed");
}}

function toggleAll(el) {{
    const expanding = el.dataset.state !== "expanded";
    document.querySelectorAll("#yearList .year-card").forEach(c => c.classList.toggle("collapsed", !expanding));
    el.dataset.state = expanding ? "expanded" : "collapsed";
    el.textContent = expanding ? "Collapse all" : "Expand all";
}}

function setPill(el) {{
    document.querySelectorAll("#yearPills .pill").forEach(p => p.classList.remove("active"));
    el.classList.add("active");
    activeYear = el.dataset.year;
    applyFilters();
}}

function applyFilters() {{
    const q = document.getElementById("searchInput").value.trim().toLowerCase();
    document.querySelectorAll("#yearList .year-card").forEach(card => {{
        const yr = card.dataset.year;
        if (activeYear !== "all" && activeYear !== yr) {{ card.classList.add("hidden"); return; }}
        if (!q) {{
            card.classList.remove("hidden");
            card.querySelectorAll(".run-row").forEach(r => r.classList.remove("hidden"));
            return;
        }}
        let any = false;
        card.querySelectorAll(".run-row").forEach(row => {{
            const text = row.dataset.period + " " + row.dataset.file + " " + yr;
            if (text.includes(q)) {{ row.classList.remove("hidden"); any = true; }}
            else {{ row.classList.add("hidden"); }}
        }});
        if (any) {{ card.classList.remove("hidden", "collapsed"); }}
        else {{ card.classList.add("hidden"); }}
    }});
}}
</script>
</body>
</html>
    """


def _detail_styles() -> str:
    return """
<style>
  /* Account overview */
  .overview {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
    gap: 14px;
    margin-bottom: 24px;
  }
  .kpi {
    background: linear-gradient(165deg, #fff 0%, var(--pc-blue-soft) 140%);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 16px 18px;
    position: relative;
    overflow: hidden;
  }
  .kpi::before {
    content: "";
    position: absolute; top: 0; left: 0;
    width: 4px; height: 100%;
    background: var(--pc-blue);
  }
  .kpi.kpi-good::before { background: #10b981; }
  .kpi.kpi-warn::before { background: var(--pc-orange); }
  .kpi.kpi-bad::before  { background: #ef4444; }
  .kpi-label {
    font-size: 0.7rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.05em; color: var(--muted); margin-bottom: 8px;
  }
  .kpi-value { font-size: 1.9rem; font-weight: 850; color: var(--pc-blue-dark); line-height: 1; letter-spacing: -0.02em; }
  .kpi-value.sm { font-size: 1.25rem; }
  .kpi-sub { font-size: 0.72rem; color: var(--muted); margin-top: 6px; }
  .kpi.kpi-good .kpi-value { color: #047857; }
  .kpi.kpi-warn .kpi-value { color: var(--pc-orange-dark); }
  .kpi.kpi-bad .kpi-value  { color: #b91c1c; }

  /* Controls */
  .controls-bar {
    display: flex; gap: 10px; margin-bottom: 20px; flex-wrap: wrap; align-items: center;
  }
  .controls-bar input {
    flex: 1; min-width: 220px; padding: 9px 14px;
    border: 1px solid var(--border); border-radius: 9px;
    font-size: 0.9rem; outline: none; background: var(--bg); color: var(--text);
  }
  .controls-bar input:focus { border-color: var(--pc-blue); box-shadow: 0 0 0 3px rgba(0,59,127,.1); }
  .filter-pills { display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }
  .pill {
    padding: 6px 15px; border-radius: 20px; border: 1px solid var(--border);
    font-size: 0.8rem; cursor: pointer; background: var(--bg); color: var(--text);
    transition: all 0.15s; user-select: none; font-weight: 600;
  }
  .pill:hover { border-color: var(--pc-blue); }
  .pill.active { background: var(--pc-blue); color: #fff; border-color: var(--pc-blue); }
  .expand-toggle { margin-left: auto; }

  /* Year panels */
  .year-card {
    border: 1px solid var(--border); border-radius: 14px;
    margin-bottom: 18px; overflow: hidden; background: #fff;
    box-shadow: 0 1px 4px rgba(0,0,0,0.04);
  }
  .year-card.hidden { display: none; }
  .year-header {
    display: flex; align-items: center; justify-content: space-between;
    padding: 16px 22px; cursor: pointer; user-select: none;
    background: linear-gradient(90deg, var(--pc-blue-soft) 0%, #fff 100%);
    border-bottom: 1px solid var(--border);
  }
  .year-card.collapsed .year-header { border-bottom: none; }
  .year-header:hover { filter: brightness(0.985); }
  .year-header-left { display: flex; align-items: center; gap: 14px; flex-wrap: wrap; }
  .chevron { display: inline-block; transition: transform 0.2s ease; color: var(--pc-blue); font-size: 0.95rem; }
  .year-card:not(.collapsed) .chevron { transform: rotate(90deg); }
  .year-num { font-size: 1.5rem; font-weight: 850; color: var(--pc-blue-dark); letter-spacing: -0.02em; }
  .year-runs-count { font-size: 0.8rem; color: var(--muted); font-weight: 600; }
  .header-kpi { font-size: 0.8rem; color: var(--muted); }
  .header-kpi b { color: var(--pc-blue-dark); font-weight: 800; }
  .header-alert { font-size: 0.7rem; font-weight: 700; color: #b91c1c; background: #fee2e2; padding: 3px 10px; border-radius: 11px; }
  .header-ok { font-size: 0.7rem; font-weight: 700; color: #047857; background: #d1fae5; padding: 3px 10px; border-radius: 11px; }

  .year-body { max-height: 6000px; overflow: hidden; transition: max-height 0.35s ease; }
  .year-card.collapsed .year-body { max-height: 0; }

  .section-block { padding: 20px 22px; border-bottom: 1px solid var(--border); }
  .section-block:last-child { border-bottom: none; }
  .section-title {
    font-size: 0.95rem; font-weight: 800; color: var(--pc-blue-dark);
    margin-bottom: 16px; display: flex; align-items: baseline; gap: 10px; flex-wrap: wrap;
  }
  .section-hint { font-size: 0.74rem; font-weight: 500; color: var(--muted); }

  /* Quarter stepper */
  .stepper { display: flex; align-items: flex-start; gap: 0; }
  .step { display: flex; flex-direction: column; align-items: center; text-align: center; min-width: 80px; }
  .step-dot {
    width: 38px; height: 38px; border-radius: 50%; display: flex;
    align-items: center; justify-content: center; font-weight: 800;
    font-size: 1.05rem; color: #fff; border: 3px solid #fff;
    box-shadow: 0 0 0 1px var(--border);
  }
  .step-done .step-dot    { background: #10b981; box-shadow: 0 0 0 1px #10b981; }
  .step-missing .step-dot { background: #ef4444; box-shadow: 0 0 0 1px #ef4444; }
  .step-future .step-dot  { background: #cbd5e1; color: #64748b; box-shadow: 0 0 0 1px #cbd5e1; }
  .step-name { font-size: 0.78rem; font-weight: 700; color: var(--text); margin-top: 8px; }
  .step-status { font-size: 0.68rem; color: var(--muted); margin-top: 2px; }
  .step-done .step-status    { color: #047857; }
  .step-missing .step-status { color: #b91c1c; }
  .step-line { flex: 1; height: 3px; background: var(--border); margin-top: 18px; border-radius: 2px; min-width: 20px; }

  /* Metric cards */
  .metrics-summary {
    display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 16px;
  }
  .metric-card {
    background: #f8fafc; border: 1px solid var(--border); border-radius: 14px; padding: 16px 18px;
  }
  .metric-card.empty { background: #fafbfc; }
  .metric-top { display: flex; align-items: center; justify-content: space-between; margin-bottom: 8px; }
  .metric-label { font-size: 0.72rem; font-weight: 800; text-transform: uppercase; letter-spacing: 0.05em; color: var(--muted); }
  .metric-info {
    width: 20px; height: 20px; border-radius: 50%; background: var(--pc-blue); color: #fff;
    font-size: 0.72rem; font-weight: 800; display: flex; align-items: center; justify-content: center;
    cursor: pointer; border: none; padding: 0; line-height: 1; transition: transform 0.12s, background 0.12s;
  }
  .metric-info:hover { background: var(--pc-blue-dark); transform: scale(1.12); }

  /* Metric explainer modal */
  .modal-overlay {
    position: fixed; inset: 0; background: rgba(15, 23, 42, 0.55);
    backdrop-filter: blur(2px); display: none; align-items: center; justify-content: center;
    z-index: 1000; padding: 20px;
  }
  .modal-overlay.show { display: flex; animation: mmFade 0.15s ease; }
  @keyframes mmFade { from { opacity: 0; } to { opacity: 1; } }
  .modal-box {
    background: #fff; border-radius: 18px; max-width: 540px; width: 100%;
    max-height: 88vh; overflow-y: auto; padding: 28px 30px 24px;
    box-shadow: 0 24px 60px rgba(0,0,0,0.3); position: relative;
    animation: mmSlide 0.2s ease;
  }
  @keyframes mmSlide { from { transform: translateY(14px); opacity: 0; } to { transform: translateY(0); opacity: 1; } }
  .modal-close {
    position: absolute; top: 16px; right: 18px; background: var(--bg); border: 1px solid var(--border);
    width: 32px; height: 32px; border-radius: 50%; font-size: 1.2rem; line-height: 1; color: var(--muted);
    cursor: pointer; transition: all 0.12s;
  }
  .modal-close:hover { background: #fee2e2; color: #b91c1c; border-color: #fecaca; }
  .modal-kicker {
    font-size: 0.68rem; font-weight: 800; text-transform: uppercase; letter-spacing: 0.08em;
    color: var(--pc-orange-dark); margin-bottom: 4px;
  }
  .modal-title { margin: 0; font-size: 1.5rem; color: var(--pc-blue-dark); letter-spacing: -0.02em; }
  .mm-head { display: flex; align-items: center; justify-content: space-between; gap: 12px; margin-bottom: 14px; }
  .mm-verdict {
    font-size: 0.72rem; font-weight: 800; text-transform: uppercase; letter-spacing: 0.04em;
    padding: 4px 12px; border-radius: 20px; white-space: nowrap;
  }
  .verdict-good { color: #047857; background: #d1fae5; }
  .verdict-warn { color: var(--pc-orange-dark); background: #fef3c7; }
  .verdict-bad  { color: #b91c1c; background: #fee2e2; }
  .verdict-neutral { color: #475569; background: #eef0f3; }

  .mm-headline {
    display: flex; align-items: baseline; gap: 12px; flex-wrap: wrap;
    background: var(--pc-blue-soft); border-radius: 12px; padding: 14px 18px; margin-bottom: 8px;
  }
  .mm-current { font-size: 1.8rem; font-weight: 850; color: var(--pc-blue-dark); letter-spacing: -0.02em; }
  .mm-change { font-size: 0.9rem; font-weight: 800; padding: 2px 10px; border-radius: 11px; }
  .mv-up { color: #047857; background: #d1fae5; }
  .mv-down { color: #b91c1c; background: #fee2e2; }
  .mm-subnote { font-size: 0.82rem; color: var(--muted); margin-bottom: 18px; line-height: 1.5; }

  .modal-section { margin-bottom: 18px; }
  .modal-h {
    font-size: 0.72rem; font-weight: 800; text-transform: uppercase; letter-spacing: 0.05em;
    color: var(--pc-blue); margin-bottom: 8px;
  }
  .modal-section p { margin: 0; font-size: 0.92rem; line-height: 1.6; color: var(--text); }

  .mm-quarters { display: flex; flex-direction: column; gap: 2px; }
  .mm-qrow {
    display: grid; grid-template-columns: 1fr auto auto; align-items: center; gap: 12px;
    padding: 9px 12px; border-radius: 9px;
  }
  .mm-qrow:nth-child(odd) { background: #f8fafc; }
  .mm-qlabel { font-size: 0.82rem; font-weight: 700; color: var(--text); }
  .mm-qval { font-size: 0.82rem; color: var(--muted); font-variant-numeric: tabular-nums; }
  .mm-qdelta { font-size: 0.78rem; font-weight: 800; padding: 1px 8px; border-radius: 9px; white-space: nowrap; }

  .mm-stats { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin-bottom: 4px; }
  .mm-stat { background: #f8fafc; border: 1px solid var(--border); border-radius: 11px; padding: 10px 12px; text-align: center; }
  .mm-stat-label { font-size: 0.62rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.04em; color: var(--muted); }
  .mm-stat-val { font-size: 1.1rem; font-weight: 850; color: var(--pc-blue-dark); margin: 3px 0 1px; }
  .mm-stat-sub { font-size: 0.62rem; color: var(--muted); }

  .mm-empty { font-size: 0.92rem; color: var(--muted); line-height: 1.6; padding: 10px 0 4px; }
  .modal-foot {
    margin-top: 18px; padding-top: 14px; border-top: 1px solid var(--border);
    font-size: 0.74rem; color: var(--muted); line-height: 1.5;
  }
  @media (max-width: 480px) {
    .mm-stats { grid-template-columns: repeat(2, 1fr); }
  }
  .metric-current { font-size: 1.5rem; font-weight: 850; color: var(--pc-blue-dark); letter-spacing: -0.02em; line-height: 1.1; }
  .metric-deltarow { display: flex; align-items: center; gap: 8px; margin: 6px 0 14px; flex-wrap: wrap; }
  .metric-delta { font-size: 0.76rem; font-weight: 800; padding: 2px 9px; border-radius: 11px; }
  .delta-up { color: #047857; background: #d1fae5; }
  .delta-down { color: #b91c1c; background: #fee2e2; }
  .delta-flat { color: #6b7280; background: #eef0f3; }
  .metric-since { font-size: 0.74rem; color: var(--muted); }
  .metric-desc { font-size: 0.72rem; color: var(--muted); margin-top: 10px; line-height: 1.4; }

  .bar-chart { display: flex; align-items: flex-end; gap: 7px; height: 96px; }
  .bar-col { flex: 1; display: flex; flex-direction: column; align-items: center; justify-content: flex-end; height: 100%; min-width: 0; }
  .bar-val { font-size: 0.6rem; font-weight: 800; color: var(--muted); margin-bottom: 4px; white-space: nowrap; }
  .bar { width: 100%; max-width: 30px; border-radius: 5px 5px 0 0; transition: opacity 0.15s; }
  .bar:hover { opacity: 0.82; }
  .bar-pos { background: linear-gradient(180deg, #3b82f6 0%, var(--pc-blue) 100%); }
  .bar-pos.bar-q { background: linear-gradient(180deg, var(--pc-blue) 0%, var(--pc-blue-dark) 100%); }
  .bar-neg { background: linear-gradient(180deg, #f87171 0%, #b91c1c 100%); }
  .bar-month { font-size: 0.63rem; color: var(--muted); margin-top: 5px; font-weight: 600; }

  /* Timeline */
  .timeline { position: relative; padding-left: 6px; }
  .timeline-item { position: relative; padding: 0 0 18px 24px; border-left: 2px solid var(--border); }
  .timeline-item:last-child { border-left-color: transparent; padding-bottom: 0; }
  .timeline-item.hidden { display: none; }
  .timeline-dot {
    position: absolute; left: -7px; top: 2px; width: 12px; height: 12px;
    border-radius: 50%; background: var(--pc-blue); border: 2px solid #fff; box-shadow: 0 0 0 1px var(--pc-blue);
  }
  .timeline-head { display: flex; align-items: center; gap: 12px; margin-bottom: 3px; }
  .timeline-period { font-size: 0.9rem; font-weight: 800; color: var(--pc-blue-dark); }
  .timeline-audit { font-size: 0.7rem; font-weight: 700; color: var(--pc-blue); text-decoration: none; border: 1px solid var(--border); padding: 1px 9px; border-radius: 10px; }
  .timeline-audit:hover { background: var(--pc-blue-soft); }
  .timeline-file { font-size: 0.82rem; color: var(--text); word-break: break-word; }
  .timeline-date { font-size: 0.72rem; color: var(--muted); margin-top: 2px; }
</style>
    """
