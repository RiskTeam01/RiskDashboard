import os
import re
import csv
import shutil
import queue
import threading
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from tkinter import (
    Tk, Frame, Label, Button, Text, Scrollbar, filedialog, messagebox,
    StringVar, BooleanVar, END, BOTH, LEFT, RIGHT, Y, X, BOTTOM,
    VERTICAL, HORIZONTAL, Checkbutton
)
from tkinter import ttk

from openpyxl import load_workbook


# ============================================================
# PDF TO EXCEL CREDIT WORKSHEET APP - V4
# ------------------------------------------------------------
# PURPOSE:
#   - User selects only the customer PDF
#   - App uses a backend Excel template stored in /templates
#   - Extracts helper-code values from the PDF
#   - Copies the Excel template into /outputs
#   - Writes the 24 yellow input cells
#   - Preserves formulas, formatting, layout
#   - Blank PDF fields stay blank in Excel
#   - Does not modify the original template
#
# INSTALL:
#   python -m pip install pymupdf openpyxl
#
# OPTIONAL:
#   python -m pip install pywin32
#
# RUN:
#   python pdf_to_excel_credit_ws_v4.py
# ============================================================


APP_ROOT = Path(__file__).resolve().parent
TEMPLATE_DIR = APP_ROOT / "templates"
OUTPUT_DIR = APP_ROOT / "outputs"
LOG_DIR = APP_ROOT / "logs"

TEMPLATE_FILENAME = "Buckler Excel Credit WS Template.xlsx"
BACKEND_TEMPLATE_PATH = TEMPLATE_DIR / TEMPLATE_FILENAME

SHEET_NAME = "Sheet1"


# ============================================================
# 24 EXCEL TARGET FIELDS
# ------------------------------------------------------------
# key:
#   - field_key/expression is what the app extracts from PDF
#   - excel_cell is where it writes in the copied Excel template
#   - duplicates are allowed, e.g. 1480 writes to D22 and D30
# ============================================================

FIELD_DEFINITIONS = [
    # LEFT SIDE INPUTS
    {"expression": "940", "label": "Total Assets", "excel_cell": "B4"},
    {"expression": "1800", "label": "Total Equity", "excel_cell": "B6"},
    {"expression": "750", "label": "Cash and cash equivalents", "excel_cell": "B8"},
    {"expression": "760", "label": "Cash segregated under federal", "excel_cell": "B10"},
    {"expression": "770", "label": "Fails to Deliver", "excel_cell": "B12"},
    {"expression": "780", "label": "Stocks Borrowed", "excel_cell": "B14"},
    {"expression": "800", "label": "Clearing Org receivables", "excel_cell": "B16"},
    {"expression": "810", "label": "Others", "excel_cell": "B18"},
    {"expression": "820", "label": "Customer Receivables", "excel_cell": "B20"},
    {"expression": "840", "label": "Reverse Repos", "excel_cell": "B22"},
    {"expression": "292", "label": "Trade Date Receivable", "excel_cell": "B24"},
    {"expression": "12019", "label": "Marketable securities", "excel_cell": "B26"},
    {"expression": "740", "label": "Non-allowable assets", "excel_cell": "B28"},
    {"expression": "890", "label": "Secured Demand Notes", "excel_cell": "B32"},
    {"expression": "1760", "label": "Total Liabilities", "excel_cell": "B34"},

    # RIGHT SIDE YELLOW FIELDS
    {"expression": "1490+1500", "label": "Fails to Receive", "excel_cell": "D12"},
    {"expression": "1510+1520", "label": "Stocks Loaned", "excel_cell": "D14"},
    {"expression": "1550+1560", "label": "Clearing Org payables", "excel_cell": "D16"},
    {"expression": "1570", "label": "Other", "excel_cell": "D18"},
    {"expression": "1580+1590", "label": "Customer payables", "excel_cell": "D20"},
    {"expression": "1480", "label": "Repos", "excel_cell": "D22"},
    {"expression": "1686", "label": "Obligation to rtn Securities Collateral", "excel_cell": "D26"},
    {"expression": "1480", "label": "All other liabilities", "excel_cell": "D30"},
    {"expression": "1730", "label": "Securities borrowings", "excel_cell": "D32"},
]


AMOUNT_PATTERN = re.compile(
    r"""
    ^\$?
    \(?
    -?
    (?:
        \d{1,3}(?:,\d{3})+ |
        \d+
    )
    (?:\.\d+)?
    \)?
    $
    """,
    re.VERBOSE,
)


@dataclass
class WordItem:
    page_index: int
    page_number: int
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    block_no: int
    line_no: int
    word_no: int

    @property
    def y_center(self) -> float:
        return (self.y0 + self.y1) / 2

    @property
    def x_center(self) -> float:
        return (self.x0 + self.x1) / 2


@dataclass
class CodeOccurrence:
    code: str
    page_number: int
    x0: float
    y0: float
    x1: float
    y1: float
    nearby_amount_text: str = ""
    nearby_context: str = ""
    note: str = ""
    confidence_score: int = 0
    selected: bool = False

    @property
    def location_text(self) -> str:
        return f"Page {self.page_number}, x={round(self.x0, 2)}, y={round(self.y0, 2)}"


@dataclass
class FieldSpec:
    expression: str
    label: str
    excel_cell: str
    codes: list[str] = field(default_factory=list)


@dataclass
class FieldResult:
    expression: str
    label: str
    excel_cell: str
    display_value: str
    numeric_value: float | None
    should_write_blank: bool
    status: str
    difficulty: str
    notes: str
    component_details: str


class CreditWorksheetApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF to Excel Credit Worksheet Generator V4")
        self.root.geometry("1560x920")

        self.pdf_path = StringVar(value="No PDF selected")
        self.template_status = StringVar(value="")
        self.status_text = StringVar(value="Ready")
        self.recalculate_with_excel = BooleanVar(value=False)

        self.queue = queue.Queue()

        self.latest_report_lines = []
        self.latest_field_results = []
        self.latest_occurrence_rows = []
        self.latest_code_occurrences = {}
        self.latest_output_excel_path = ""

        self.ensure_folder_structure()
        self._build_ui()
        self.refresh_template_status()
        self._poll_queue()

    # ============================================================
    # FOLDER SETUP
    # ============================================================

    def ensure_folder_structure(self):
        for folder in [TEMPLATE_DIR, OUTPUT_DIR, LOG_DIR]:
            folder.mkdir(parents=True, exist_ok=True)

    def refresh_template_status(self):
        if BACKEND_TEMPLATE_PATH.exists():
            self.template_status.set(f"Backend template found: {BACKEND_TEMPLATE_PATH}")
        else:
            self.template_status.set(f"Missing backend template: {BACKEND_TEMPLATE_PATH}")

    # ============================================================
    # UI
    # ============================================================

    def _build_ui(self):
        top = Frame(self.root, padx=10, pady=8)
        top.pack(fill=X)

        Label(
            top,
            text="PDF to Excel Credit Worksheet Generator V4",
            font=("Segoe UI", 16, "bold"),
        ).pack(anchor="w")

        Label(
            top,
            text=(
                "User selects a customer PDF. The app copies the backend Excel template, "
                "fills the 24 yellow input cells, and saves a completed workbook."
            ),
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(2, 8))

        pdf_frame = Frame(top)
        pdf_frame.pack(fill=X)

        Button(pdf_frame, text="Select PDF", command=self.select_pdf, width=16).pack(side=LEFT)

        Label(
            pdf_frame,
            textvariable=self.pdf_path,
            font=("Segoe UI", 9),
            anchor="w",
        ).pack(side=LEFT, padx=10, fill=X, expand=True)

        template_frame = Frame(top)
        template_frame.pack(fill=X, pady=(6, 0))

        Button(
            template_frame,
            text="Install/Replace Backend Template",
            command=self.install_backend_template,
            width=30,
        ).pack(side=LEFT)

        Label(
            template_frame,
            textvariable=self.template_status,
            font=("Segoe UI", 9),
            anchor="w",
        ).pack(side=LEFT, padx=10, fill=X, expand=True)

        button_frame = Frame(top)
        button_frame.pack(fill=X, pady=(8, 0))

        Button(button_frame, text="Run PDF Preview", command=self.run_pdf_preview, width=18).pack(side=LEFT)

        Button(
            button_frame,
            text="Generate Excel",
            command=self.generate_excel,
            width=18,
        ).pack(side=LEFT, padx=(6, 0))

        Button(
            button_frame,
            text="Open Outputs Folder",
            command=self.open_outputs_folder,
            width=18,
        ).pack(side=LEFT, padx=(6, 0))

        Button(
            button_frame,
            text="Copy Field Results",
            command=self.copy_field_results_to_clipboard,
            width=20,
        ).pack(side=LEFT, padx=(6, 0))

        Button(
            button_frame,
            text="Copy Raw Occurrences",
            command=self.copy_raw_occurrences_to_clipboard,
            width=22,
        ).pack(side=LEFT, padx=(6, 0))

        Button(
            button_frame,
            text="Copy Everything",
            command=self.copy_all_results_to_clipboard,
            width=18,
        ).pack(side=LEFT, padx=(6, 0))

        Button(
            button_frame,
            text="Save Debug Report",
            command=self.save_report,
            width=18,
        ).pack(side=LEFT, padx=(6, 0))

        Button(
            button_frame,
            text="Clear",
            command=self.clear_all,
            width=10,
        ).pack(side=LEFT, padx=(6, 0))

        Checkbutton(
            button_frame,
            text="Recalculate with Excel COM",
            variable=self.recalculate_with_excel,
            font=("Segoe UI", 9),
        ).pack(side=LEFT, padx=(12, 0))

        Label(
            button_frame,
            textvariable=self.status_text,
            font=("Segoe UI", 9, "bold"),
        ).pack(side=LEFT, padx=12)

        self.progress = ttk.Progressbar(button_frame, mode="indeterminate", length=160)
        self.progress.pack(side=LEFT, padx=8)

        main = Frame(self.root, padx=10, pady=8)
        main.pack(fill=BOTH, expand=True)

        left = Frame(main)
        left.pack(side=LEFT, fill=BOTH, expand=False)

        Label(
            left,
            text="Built-In 24 Field Map",
            font=("Segoe UI", 11, "bold"),
        ).pack(anchor="w")

        Label(
            left,
            text="These PDF helper code values write into the listed Excel cells.",
            font=("Segoe UI", 9),
        ).pack(anchor="w")

        self.field_map_text = Text(left, width=48, height=30, font=("Consolas", 10), wrap="none")
        self.field_map_text.pack(fill=BOTH, expand=True, pady=(4, 0))

        self.populate_field_map_text()

        right = Frame(main)
        right.pack(side=RIGHT, fill=BOTH, expand=True, padx=(12, 0))

        # Field Results
        field_header = Frame(right)
        field_header.pack(fill=X)

        Label(field_header, text="Field Results", font=("Segoe UI", 11, "bold")).pack(side=LEFT)

        Button(
            field_header,
            text="Copy Field Results",
            command=self.copy_field_results_to_clipboard,
            width=20,
        ).pack(side=RIGHT)

        field_frame = Frame(right)
        field_frame.pack(fill=BOTH, expand=True, pady=(4, 8))

        field_columns = (
            "expression",
            "label",
            "excel_cell",
            "value",
            "status",
            "difficulty",
            "notes",
            "details",
        )

        self.field_tree = ttk.Treeview(field_frame, columns=field_columns, show="headings", height=14)

        self.field_tree.heading("expression", text="PDF Code(s)")
        self.field_tree.heading("label", text="Field")
        self.field_tree.heading("excel_cell", text="Excel Cell")
        self.field_tree.heading("value", text="Preview Value")
        self.field_tree.heading("status", text="Status")
        self.field_tree.heading("difficulty", text="Difficulty")
        self.field_tree.heading("notes", text="Notes")
        self.field_tree.heading("details", text="Component Details")

        self.field_tree.column("expression", width=110, anchor="center")
        self.field_tree.column("label", width=260)
        self.field_tree.column("excel_cell", width=90, anchor="center")
        self.field_tree.column("value", width=150)
        self.field_tree.column("status", width=120, anchor="center")
        self.field_tree.column("difficulty", width=120, anchor="center")
        self.field_tree.column("notes", width=360)
        self.field_tree.column("details", width=560)

        field_y_scroll = Scrollbar(field_frame, orient=VERTICAL, command=self.field_tree.yview)
        field_x_scroll = Scrollbar(field_frame, orient=HORIZONTAL, command=self.field_tree.xview)
        self.field_tree.configure(yscrollcommand=field_y_scroll.set, xscrollcommand=field_x_scroll.set)

        self.field_tree.grid(row=0, column=0, sticky="nsew")
        field_y_scroll.grid(row=0, column=1, sticky="ns")
        field_x_scroll.grid(row=1, column=0, sticky="ew")
        field_frame.grid_rowconfigure(0, weight=1)
        field_frame.grid_columnconfigure(0, weight=1)

        self.field_tree.tag_configure("ready", background="#dff5df")
        self.field_tree.tag_configure("blank", background="#fff6cc")
        self.field_tree.tag_configure("review", background="#ffe0b3")
        self.field_tree.tag_configure("missing", background="#ffd6d6")

        # Raw Occurrences
        occ_header = Frame(right)
        occ_header.pack(fill=X)

        Label(occ_header, text="Raw Helper Code Occurrences", font=("Segoe UI", 11, "bold")).pack(side=LEFT)

        Button(
            occ_header,
            text="Copy Raw Occurrences",
            command=self.copy_raw_occurrences_to_clipboard,
            width=22,
        ).pack(side=RIGHT)

        occ_frame = Frame(right)
        occ_frame.pack(fill=BOTH, expand=True, pady=(4, 8))

        occ_columns = (
            "code",
            "selected",
            "page",
            "x0",
            "y0",
            "x1",
            "y1",
            "amount",
            "confidence",
            "note",
            "context",
        )

        self.occ_tree = ttk.Treeview(occ_frame, columns=occ_columns, show="headings", height=8)

        self.occ_tree.heading("code", text="Code")
        self.occ_tree.heading("selected", text="Selected")
        self.occ_tree.heading("page", text="Page")
        self.occ_tree.heading("x0", text="X0")
        self.occ_tree.heading("y0", text="Y0")
        self.occ_tree.heading("x1", text="X1")
        self.occ_tree.heading("y1", text="Y1")
        self.occ_tree.heading("amount", text="Nearby Amount")
        self.occ_tree.heading("confidence", text="Score")
        self.occ_tree.heading("note", text="Note")
        self.occ_tree.heading("context", text="Nearby Context")

        self.occ_tree.column("code", width=70, anchor="center")
        self.occ_tree.column("selected", width=80, anchor="center")
        self.occ_tree.column("page", width=60, anchor="center")
        self.occ_tree.column("x0", width=80, anchor="center")
        self.occ_tree.column("y0", width=80, anchor="center")
        self.occ_tree.column("x1", width=80, anchor="center")
        self.occ_tree.column("y1", width=80, anchor="center")
        self.occ_tree.column("amount", width=150)
        self.occ_tree.column("confidence", width=70, anchor="center")
        self.occ_tree.column("note", width=360)
        self.occ_tree.column("context", width=650)

        occ_y_scroll = Scrollbar(occ_frame, orient=VERTICAL, command=self.occ_tree.yview)
        occ_x_scroll = Scrollbar(occ_frame, orient=HORIZONTAL, command=self.occ_tree.xview)
        self.occ_tree.configure(yscrollcommand=occ_y_scroll.set, xscrollcommand=occ_x_scroll.set)

        self.occ_tree.grid(row=0, column=0, sticky="nsew")
        occ_y_scroll.grid(row=0, column=1, sticky="ns")
        occ_x_scroll.grid(row=1, column=0, sticky="ew")
        occ_frame.grid_rowconfigure(0, weight=1)
        occ_frame.grid_columnconfigure(0, weight=1)

        self.occ_tree.tag_configure("selected", background="#dff5df")
        self.occ_tree.tag_configure("unselected", background="#f4f4f4")

        # Debug Console
        console_header = Frame(right)
        console_header.pack(fill=X)

        Label(console_header, text="Debug Console", font=("Segoe UI", 11, "bold")).pack(side=LEFT)

        Button(
            console_header,
            text="Copy Console",
            command=self.copy_console_to_clipboard,
            width=16,
        ).pack(side=RIGHT)

        console_frame = Frame(right)
        console_frame.pack(fill=BOTH, expand=True)

        self.console = Text(console_frame, height=10, font=("Consolas", 9), wrap="none")
        self.console.pack(side=LEFT, fill=BOTH, expand=True)

        console_y_scroll = Scrollbar(console_frame, orient=VERTICAL, command=self.console.yview)
        console_y_scroll.pack(side=RIGHT, fill=Y)
        self.console.configure(yscrollcommand=console_y_scroll.set)

        bottom = Frame(self.root, padx=10, pady=6)
        bottom.pack(fill=X, side=BOTTOM)

        Label(
            bottom,
            text=(
                "Rules: blank PDF fields stay blank in Excel. "
                "1480 is intentionally used twice. "
                "Template is copied before writing. Original backend template is never modified."
            ),
            font=("Segoe UI", 9, "italic"),
        ).pack(anchor="w")

    def populate_field_map_text(self):
        self.field_map_text.delete("1.0", END)

        lines = []
        lines.append("PDF CODE(S)       EXCEL CELL     FIELD")
        lines.append("-" * 62)

        for item in FIELD_DEFINITIONS:
            lines.append(
                f"{item['expression']:<17} {item['excel_cell']:<13} {item['label']}"
            )

        lines.append("")
        lines.append(f"Total target yellow input cells: {len(FIELD_DEFINITIONS)}")
        lines.append("")
        lines.append("Backend template path:")
        lines.append(str(BACKEND_TEMPLATE_PATH))

        self.field_map_text.insert("1.0", "\n".join(lines))
        self.field_map_text.configure(state="disabled")

    # ============================================================
    # UI EVENTS
    # ============================================================

    def select_pdf(self):
        path = filedialog.askopenfilename(
            title="Select customer PDF",
            filetypes=[("PDF Files", "*.pdf")],
        )

        if path:
            self.pdf_path.set(path)
            self.log(f"Selected PDF: {path}")

    def install_backend_template(self):
        path = filedialog.askopenfilename(
            title="Select blank Excel credit worksheet template",
            filetypes=[("Excel Files", "*.xlsx")],
        )

        if not path:
            return

        try:
            TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
            shutil.copy2(path, BACKEND_TEMPLATE_PATH)
            self.refresh_template_status()
            self.log(f"[TEMPLATE] Installed backend template: {BACKEND_TEMPLATE_PATH}")
            messagebox.showinfo(
                "Template Installed",
                f"Backend template installed:\n{BACKEND_TEMPLATE_PATH}"
            )
        except Exception as e:
            self.log(f"[ERROR] Failed installing template: {e}")
            messagebox.showerror("Template Install Failed", str(e))

    def run_pdf_preview(self):
        path = self.pdf_path.get()

        if not self.validate_pdf_selected(path):
            return

        field_specs = self.get_field_specs()

        self.clear_results_only()
        self.status_text.set("Running PDF preview...")
        self.progress.start(10)

        worker = threading.Thread(
            target=self._pdf_preview_worker,
            args=(path, field_specs),
            daemon=True,
        )
        worker.start()

    def generate_excel(self):
        path = self.pdf_path.get()

        if not self.validate_pdf_selected(path):
            return

        if not BACKEND_TEMPLATE_PATH.exists():
            messagebox.showerror(
                "Missing Backend Template",
                f"Backend Excel template not found:\n{BACKEND_TEMPLATE_PATH}\n\n"
                "Use the Install/Replace Backend Template button first."
            )
            return

        field_specs = self.get_field_specs()

        self.clear_results_only()
        self.status_text.set("Generating Excel...")
        self.progress.start(10)

        worker = threading.Thread(
            target=self._generate_excel_worker,
            args=(path, field_specs),
            daemon=True,
        )
        worker.start()

    def validate_pdf_selected(self, path: str) -> bool:
        if not path or path == "No PDF selected":
            messagebox.showwarning("Missing PDF", "Please select a PDF first.")
            return False

        if not os.path.exists(path):
            messagebox.showerror("PDF Not Found", f"The selected PDF does not exist:\n{path}")
            return False

        return True

    def open_outputs_folder(self):
        try:
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            os.startfile(str(OUTPUT_DIR))
        except Exception as e:
            messagebox.showerror("Open Folder Failed", str(e))

    def save_report(self):
        if not self.latest_report_lines:
            messagebox.showinfo("No Report", "There is no debug report to save yet.")
            return

        default_name = f"credit_ws_debug_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

        save_path = filedialog.asksaveasfilename(
            title="Save Debug Report",
            defaultextension=".txt",
            initialfile=default_name,
            initialdir=str(LOG_DIR),
            filetypes=[("Text Files", "*.txt")],
        )

        if not save_path:
            return

        try:
            with open(save_path, "w", encoding="utf-8") as f:
                f.write("\n".join(self.latest_report_lines))

            messagebox.showinfo("Saved", f"Debug report saved:\n{save_path}")
            self.log(f"Saved debug report: {save_path}")

        except Exception as e:
            messagebox.showerror("Save Failed", str(e))
            self.log(f"[ERROR] Failed to save report: {e}")

    def clear_all(self):
        self.clear_results_only()
        self.console.delete("1.0", END)
        self.latest_report_lines = []
        self.latest_output_excel_path = ""
        self.status_text.set("Ready")

    def clear_results_only(self):
        for item in self.field_tree.get_children():
            self.field_tree.delete(item)

        for item in self.occ_tree.get_children():
            self.occ_tree.delete(item)

        self.latest_field_results = []
        self.latest_occurrence_rows = []
        self.latest_code_occurrences = {}

    # ============================================================
    # FIELD SPECS
    # ============================================================

    def get_field_specs(self) -> list[FieldSpec]:
        specs = []

        for item in FIELD_DEFINITIONS:
            expression = item["expression"].replace(" ", "")
            codes = expression.split("+")
            specs.append(
                FieldSpec(
                    expression=expression,
                    label=item["label"],
                    excel_cell=item["excel_cell"],
                    codes=codes,
                )
            )

        return specs

    def get_requested_codes(self, field_specs: list[FieldSpec]) -> list[str]:
        return sorted(
            set(code for spec in field_specs for code in spec.codes),
            key=lambda value: int(value),
        )

    # ============================================================
    # WORKERS
    # ============================================================

    def _pdf_preview_worker(self, pdf_path: str, field_specs: list[FieldSpec]):
        try:
            self._qlog("=" * 100)
            self._qlog("Starting PDF preview diagnostics V4")
            self._qlog(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self._qlog(f"PDF: {pdf_path}")
            self._qlog(f"Field definitions: {len(field_specs)}")
            self._qlog("=" * 100)

            field_results, occurrences_by_code, readability_info = self.extract_pdf_field_results(
                pdf_path=pdf_path,
                field_specs=field_specs,
            )

            self.publish_results(
                field_results=field_results,
                occurrences_by_code=occurrences_by_code,
                readability_info=readability_info,
                requested_codes=self.get_requested_codes(field_specs),
            )

            self.queue.put(("status", "PDF preview complete"))
            self.queue.put(("done", None))

        except Exception:
            self._qlog("[FATAL ERROR]")
            self._qlog(traceback.format_exc())
            self.queue.put(("status", "PDF preview failed"))
            self.queue.put(("done", None))

    def _generate_excel_worker(self, pdf_path: str, field_specs: list[FieldSpec]):
        try:
            self._qlog("=" * 100)
            self._qlog("Starting PDF to Excel generation V4")
            self._qlog(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self._qlog(f"PDF: {pdf_path}")
            self._qlog(f"Backend template: {BACKEND_TEMPLATE_PATH}")
            self._qlog(f"Field definitions: {len(field_specs)}")
            self._qlog("=" * 100)

            field_results, occurrences_by_code, readability_info = self.extract_pdf_field_results(
                pdf_path=pdf_path,
                field_specs=field_specs,
            )

            self.publish_results(
                field_results=field_results,
                occurrences_by_code=occurrences_by_code,
                readability_info=readability_info,
                requested_codes=self.get_requested_codes(field_specs),
            )

            missing = [r for r in field_results if r.status == "MISSING"]

            if missing:
                self._qlog("[STOP] Excel was not generated because one or more fields are missing.")
                for result in missing:
                    self._qlog(f"  - Missing: {result.expression} {result.label} -> {result.excel_cell}")

                self.queue.put(("status", "Generation stopped: missing fields"))
                self.queue.put(("done", None))
                return

            output_path = self.create_output_excel_path(pdf_path)
            self._qlog(f"[EXCEL] Copying backend template to output: {output_path}")
            shutil.copy2(BACKEND_TEMPLATE_PATH, output_path)

            self.write_results_to_excel(output_path, field_results)

            if self.recalculate_with_excel.get():
                self.recalculate_workbook_with_excel_com(output_path)
            else:
                self._qlog("[EXCEL] Skipped Excel COM recalculation. Workbook formulas will calculate when opened in Excel.")

            self.latest_output_excel_path = str(output_path)

            self._qlog("")
            self._qlog("[SUCCESS] Completed workbook created:")
            self._qlog(str(output_path))

            self.queue.put(("status", "Excel generated successfully"))
            self.queue.put(("done", None))
            self.queue.put(("generated", str(output_path)))

        except Exception:
            self._qlog("[FATAL ERROR]")
            self._qlog(traceback.format_exc())
            self.queue.put(("status", "Excel generation failed"))
            self.queue.put(("done", None))

    # ============================================================
    # PDF EXTRACTION
    # ============================================================

    def extract_pdf_field_results(self, pdf_path: str, field_specs: list[FieldSpec]):
        try:
            import fitz
        except ImportError:
            raise RuntimeError("PyMuPDF is not installed. Run: python -m pip install pymupdf")

        requested_codes = self.get_requested_codes(field_specs)
        self._qlog(f"[REQUESTED CODES] {', '.join(requested_codes)}")

        doc = fitz.open(pdf_path)

        all_words = []
        page_count = len(doc)
        total_text_chars = 0
        pages_with_text = 0
        total_images = 0

        self._qlog(f"[PDF] Page count: {page_count}")

        for page_index in range(page_count):
            page = doc[page_index]
            text = page.get_text("text") or ""
            words_raw = page.get_text("words") or []
            images = page.get_images(full=True) or []

            text_chars = len(text.strip())
            total_text_chars += text_chars
            total_images += len(images)

            if text.strip():
                pages_with_text += 1

            self._qlog(
                f"[PAGE {page_index + 1}] "
                f"text chars={text_chars}, "
                f"words={len(words_raw)}, "
                f"images={len(images)}, "
                f"size={round(page.rect.width, 2)}x{round(page.rect.height, 2)}"
            )

            if page_index == 0:
                preview = text.strip().replace("\n", " ")
                preview = preview[:900] + ("..." if len(preview) > 900 else "")
                self._qlog(f"[PAGE 1 TEXT PREVIEW] {preview}")

            for word in words_raw:
                try:
                    x0, y0, x1, y1, word_text, block_no, line_no, word_no = word
                    clean_text = str(word_text).strip()

                    if clean_text:
                        all_words.append(
                            WordItem(
                                page_index=page_index,
                                page_number=page_index + 1,
                                x0=x0,
                                y0=y0,
                                x1=x1,
                                y1=y1,
                                text=clean_text,
                                block_no=block_no,
                                line_no=line_no,
                                word_no=word_no,
                            )
                        )
                except Exception as e:
                    self._qlog(f"[WARN] Failed parsing word on page {page_index + 1}: {e}")

        doc.close()

        readability = self.calculate_readability(
            page_count=page_count,
            pages_with_text=pages_with_text,
            total_text_chars=total_text_chars,
            total_words=len(all_words),
            total_images=total_images,
        )

        readability_info = {
            "page_count": page_count,
            "pages_with_text": pages_with_text,
            "total_text_chars": total_text_chars,
            "total_words": len(all_words),
            "total_images": total_images,
            "readability": readability,
        }

        self._qlog("-" * 100)
        self._qlog("[READABILITY SUMMARY]")
        self._qlog(f"Pages with selectable text: {pages_with_text}/{page_count}")
        self._qlog(f"Total extracted text characters: {total_text_chars}")
        self._qlog(f"Total extracted word items: {len(all_words)}")
        self._qlog(f"Total embedded images detected: {total_images}")
        self._qlog(f"Readability result: {readability}")
        self._qlog("-" * 100)

        occurrences_by_code = self.find_requested_code_occurrences(
            all_words=all_words,
            requested_codes=requested_codes,
        )

        self.select_best_occurrences(occurrences_by_code)

        for code in requested_codes:
            occurrences = occurrences_by_code.get(code, [])
            self._qlog("")
            self._qlog(f"[CODE {code}] occurrences found: {len(occurrences)}")

            if not occurrences:
                self._qlog("  - Missing from selectable text.")
                continue

            for occ in occurrences:
                selected_text = "YES" if occ.selected else "NO"
                self._qlog(
                    f"  - selected={selected_text}, "
                    f"page={occ.page_number}, "
                    f"x={round(occ.x0, 2)}-{round(occ.x1, 2)}, "
                    f"y={round(occ.y0, 2)}-{round(occ.y1, 2)}, "
                    f"amount='{occ.nearby_amount_text}', "
                    f"score={occ.confidence_score}, "
                    f"note='{occ.note}', "
                    f"context='{occ.nearby_context}'"
                )

        field_results = []

        for spec in field_specs:
            result = self.evaluate_field(spec, occurrences_by_code)
            field_results.append(result)

            self._qlog("")
            self._qlog(f"[FIELD] {spec.expression} = {spec.label} -> {spec.excel_cell}")
            self._qlog(f"  Preview value: {result.display_value}")
            self._qlog(f"  Status: {result.status}")
            self._qlog(f"  Difficulty: {result.difficulty}")
            self._qlog(f"  Notes: {result.notes}")
            self._qlog(f"  Details: {result.component_details}")

        return field_results, occurrences_by_code, readability_info

    def calculate_readability(
        self,
        page_count: int,
        pages_with_text: int,
        total_text_chars: int,
        total_words: int,
        total_images: int,
    ) -> str:
        if total_words == 0 and total_text_chars == 0:
            if total_images > 0:
                return "HARD - likely scanned/image PDF; OCR would be needed"
            return "HARD - no selectable text detected"

        avg_words_per_page = total_words / max(page_count, 1)

        if pages_with_text == page_count and avg_words_per_page >= 30:
            return "EASY - selectable text detected on all pages"

        if pages_with_text > 0 and total_words >= 50:
            return "MEDIUM - some selectable text detected; layout needs testing"

        return "HARD - very limited selectable text detected"

    def normalize_code_token(self, text: str) -> str:
        cleaned = str(text).strip()
        cleaned = cleaned.replace("[", "").replace("]", "")
        cleaned = cleaned.replace("(", "").replace(")", "")
        cleaned = cleaned.replace("{", "").replace("}", "")
        cleaned = cleaned.replace(".", "")
        cleaned = cleaned.strip()
        return cleaned

    def is_amount_like(self, text: str) -> bool:
        cleaned = str(text).strip()

        if not cleaned:
            return False

        possible_code = (
            cleaned.replace("$", "")
            .replace("-", "")
            .replace("(", "")
            .replace(")", "")
            .strip()
        )

        if "," not in cleaned and possible_code.isdigit() and len(possible_code) <= 5:
            return False

        return bool(AMOUNT_PATTERN.match(cleaned))

    def amount_to_number(self, text: str):
        if not text:
            return None

        cleaned = str(text).strip()
        negative = False

        if cleaned.startswith("(") and cleaned.endswith(")"):
            negative = True

        cleaned = cleaned.replace("$", "")
        cleaned = cleaned.replace(",", "")
        cleaned = cleaned.replace("(", "")
        cleaned = cleaned.replace(")", "")
        cleaned = cleaned.strip()

        try:
            value = float(cleaned)
            if negative:
                value *= -1
            return value
        except Exception:
            return None

    def format_number_for_display(self, value: float | None) -> str:
        if value is None:
            return ""

        if abs(value - int(value)) < 0.000001:
            return f"{int(value):,}"

        return f"{value:,.2f}"

    def find_requested_code_occurrences(
        self,
        all_words: list[WordItem],
        requested_codes: list[str],
    ) -> dict[str, list[CodeOccurrence]]:
        requested_set = set(requested_codes)
        occurrences_by_code = {code: [] for code in requested_codes}

        for word in all_words:
            normalized = self.normalize_code_token(word.text)

            if normalized not in requested_set:
                continue

            nearby_amount, nearby_context, note, score = self.find_left_side_amount_preview(
                target_word=word,
                all_words=all_words,
            )

            occurrence = CodeOccurrence(
                code=normalized,
                page_number=word.page_number,
                x0=word.x0,
                y0=word.y0,
                x1=word.x1,
                y1=word.y1,
                nearby_amount_text=nearby_amount,
                nearby_context=nearby_context,
                note=note,
                confidence_score=score,
            )

            occurrences_by_code[normalized].append(occurrence)

        return occurrences_by_code

    def find_left_side_amount_preview(
        self,
        target_word: WordItem,
        all_words: list[WordItem],
    ):
        y_tolerance = 5.75
        max_left_distance = 260

        same_page_words = [w for w in all_words if w.page_index == target_word.page_index]

        same_row_left_words = [
            w for w in same_page_words
            if w.x1 < target_word.x0
            and abs(w.y_center - target_word.y_center) <= y_tolerance
            and (target_word.x0 - w.x1) <= max_left_distance
        ]

        same_row_left_words.sort(key=lambda w: w.x1, reverse=True)

        nearby_context = " ".join(w.text for w in same_row_left_words[:12])

        for candidate in same_row_left_words:
            candidate_text = candidate.text.strip()

            if self.is_amount_like(candidate_text):
                score = 80

                distance = target_word.x0 - candidate.x1
                if distance <= 80:
                    score += 10
                elif distance <= 150:
                    score += 5

                if "," in candidate_text:
                    score += 10

                return (
                    candidate_text,
                    nearby_context,
                    "Nearby amount-like value found to the left",
                    score,
                )

        if nearby_context:
            return (
                "",
                nearby_context,
                "Code found, but no amount-like value was found on the same row. Treat as valid blank unless later rules say otherwise.",
                55,
            )

        return (
            "",
            "",
            "Code found, but left-side row area appears blank or not selectable. Treat as valid blank unless later rules say otherwise.",
            50,
        )

    def select_best_occurrences(self, occurrences_by_code: dict[str, list[CodeOccurrence]]):
        for code, occurrences in occurrences_by_code.items():
            if not occurrences:
                continue

            with_amount = [occ for occ in occurrences if occ.nearby_amount_text]

            if len(occurrences) == 1:
                occurrences[0].selected = True
                continue

            if len(with_amount) == 1:
                with_amount[0].selected = True
                with_amount[0].note += " | Auto-selected because it is the only occurrence with a nearby amount."
                continue

            if len(with_amount) > 1:
                with_amount.sort(key=lambda occ: occ.confidence_score, reverse=True)
                with_amount[0].selected = True
                with_amount[0].note += " | Auto-selected by highest score, but multiple amount matches exist. Needs review."
                continue

            occurrences.sort(key=lambda occ: occ.confidence_score, reverse=True)
            occurrences[0].selected = True
            occurrences[0].note += " | Auto-selected blank candidate, but multiple blank/code-reference matches exist. Needs review."

    # ============================================================
    # FIELD EVALUATION
    # ============================================================

    def evaluate_field(
        self,
        spec: FieldSpec,
        occurrences_by_code: dict[str, list[CodeOccurrence]],
    ) -> FieldResult:
        component_values = []
        component_notes = []
        missing_codes = []
        review_needed = False
        all_components_blank_or_found = True

        for code in spec.codes:
            occurrences = occurrences_by_code.get(code, [])

            if not occurrences:
                missing_codes.append(code)
                all_components_blank_or_found = False
                component_notes.append(f"{code}: MISSING")
                continue

            selected = next((occ for occ in occurrences if occ.selected), occurrences[0])

            if len(occurrences) > 1:
                review_needed = True

            if selected.nearby_amount_text:
                numeric_value = self.amount_to_number(selected.nearby_amount_text)

                if numeric_value is None:
                    review_needed = True
                    component_notes.append(f"{code}: found '{selected.nearby_amount_text}' but could not convert")
                else:
                    component_values.append(numeric_value)
                    component_notes.append(f"{code}: {selected.nearby_amount_text} ({selected.location_text})")
            else:
                component_notes.append(f"{code}: blank ({selected.location_text})")

        if missing_codes:
            return FieldResult(
                expression=spec.expression,
                label=spec.label,
                excel_cell=spec.excel_cell,
                display_value="",
                numeric_value=None,
                should_write_blank=True,
                status="MISSING",
                difficulty="Hard",
                notes=f"Missing code(s): {', '.join(missing_codes)}",
                component_details=" | ".join(component_notes),
            )

        if len(spec.codes) == 1:
            code = spec.codes[0]
            occurrences = occurrences_by_code.get(code, [])
            selected = next((occ for occ in occurrences if occ.selected), occurrences[0])

            if selected.nearby_amount_text:
                numeric_value = self.amount_to_number(selected.nearby_amount_text)

                status = "READY"
                difficulty = "Easy" if len(occurrences) == 1 else "Medium"
                notes = "Single code found with value."

                if len(occurrences) > 1:
                    notes += " Multiple occurrences exist, so final extraction should use position rules."

                return FieldResult(
                    expression=spec.expression,
                    label=spec.label,
                    excel_cell=spec.excel_cell,
                    display_value=selected.nearby_amount_text,
                    numeric_value=numeric_value,
                    should_write_blank=False,
                    status=status,
                    difficulty=difficulty,
                    notes=notes,
                    component_details=" | ".join(component_notes),
                )

            status = "VALID BLANK"
            difficulty = "Easy" if len(occurrences) == 1 else "Medium"
            notes = "Code found with no amount. Per rule, this is a valid blank."

            if len(occurrences) > 1:
                notes += " Multiple occurrences exist, so final extraction should use position rules."

            return FieldResult(
                expression=spec.expression,
                label=spec.label,
                excel_cell=spec.excel_cell,
                display_value="",
                numeric_value=None,
                should_write_blank=True,
                status=status,
                difficulty=difficulty,
                notes=notes,
                component_details=" | ".join(component_notes),
            )

        # Combined-code field.
        if component_values:
            total = sum(component_values)
            display_value = self.format_number_for_display(total)

            if review_needed:
                status = "NEEDS REVIEW"
                difficulty = "Medium"
                notes = "Combined field calculated, but one or more component codes had multiple occurrences."
            else:
                status = "READY"
                difficulty = "Medium"
                notes = "Combined field calculated from component helper codes."

            return FieldResult(
                expression=spec.expression,
                label=spec.label,
                excel_cell=spec.excel_cell,
                display_value=display_value,
                numeric_value=total,
                should_write_blank=False,
                status=status,
                difficulty=difficulty,
                notes=notes,
                component_details=" | ".join(component_notes),
            )

        if all_components_blank_or_found:
            return FieldResult(
                expression=spec.expression,
                label=spec.label,
                excel_cell=spec.excel_cell,
                display_value="",
                numeric_value=None,
                should_write_blank=True,
                status="VALID BLANK",
                difficulty="Medium",
                notes="All component codes were found, but all appear blank. Per rule, this is valid blank.",
                component_details=" | ".join(component_notes),
            )

        return FieldResult(
            expression=spec.expression,
            label=spec.label,
            excel_cell=spec.excel_cell,
            display_value="",
            numeric_value=None,
            should_write_blank=True,
            status="NEEDS REVIEW",
            difficulty="Medium/Hard",
            notes="Combined field could not be confidently calculated.",
            component_details=" | ".join(component_notes),
        )

    # ============================================================
    # RESULT PUBLISHING / SUMMARY
    # ============================================================

    def publish_results(
        self,
        field_results: list[FieldResult],
        occurrences_by_code: dict[str, list[CodeOccurrence]],
        readability_info: dict,
        requested_codes: list[str],
    ):
        self.latest_code_occurrences = occurrences_by_code

        for code in requested_codes:
            for occ in occurrences_by_code.get(code, []):
                self.queue.put(("occurrence", occ))

        for result in field_results:
            self.queue.put(("field_result", result))

        summary_lines = self.build_overall_summary(
            field_results=field_results,
            occurrences_by_code=occurrences_by_code,
            requested_codes=requested_codes,
            readability_info=readability_info,
        )

        self._qlog("")
        self._qlog("=" * 100)
        self._qlog("[OVERALL PDF AUTOMATION READINESS]")
        for line in summary_lines:
            self._qlog(line)
        self._qlog("=" * 100)

    def build_overall_summary(
        self,
        field_results: list[FieldResult],
        occurrences_by_code: dict[str, list[CodeOccurrence]],
        requested_codes: list[str],
        readability_info: dict,
    ) -> list[str]:
        total_fields = len(field_results)
        ready_count = sum(1 for r in field_results if r.status == "READY")
        blank_count = sum(1 for r in field_results if r.status == "VALID BLANK")
        review_count = sum(1 for r in field_results if r.status == "NEEDS REVIEW")
        missing_count = sum(1 for r in field_results if r.status == "MISSING")

        codes_found = sum(1 for code in requested_codes if occurrences_by_code.get(code))
        multiple_codes = [code for code in requested_codes if len(occurrences_by_code.get(code, [])) > 1]

        lines = [
            f"Required Excel yellow input cells checked: {total_fields}",
            f"Unique helper codes requested: {len(requested_codes)}",
            f"Unique helper codes found: {codes_found}/{len(requested_codes)}",
            f"READY fields: {ready_count}",
            f"VALID BLANK fields: {blank_count}",
            f"NEEDS REVIEW fields: {review_count}",
            f"MISSING fields: {missing_count}",
            f"Codes with multiple occurrences: {', '.join(multiple_codes) if multiple_codes else 'None'}",
            f"PDF readability: {readability_info.get('readability')}",
            f"Extracted word count: {readability_info.get('total_words')}",
            f"Extracted text characters: {readability_info.get('total_text_chars')}",
            "",
        ]

        if missing_count == 0:
            lines.append("Excel generation readiness: GOOD.")
            lines.append("Reason: all requested helper codes were found or valid blanks were detected.")
        else:
            lines.append("Excel generation readiness: BLOCKED.")
            lines.append("Reason: one or more helper codes were missing.")

        lines.append("")
        lines.append("Blank fields will be written as true blank Excel cells.")
        lines.append("The backend Excel template is copied first; the original template is not modified.")

        return lines

    # ============================================================
    # EXCEL WRITING
    # ============================================================

    def create_output_excel_path(self, pdf_path: str) -> Path:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        pdf_stem = Path(pdf_path).stem
        safe_pdf_stem = re.sub(r"[^A-Za-z0-9_\-]+", "_", pdf_stem).strip("_")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        filename = f"Completed_Credit_WS_{safe_pdf_stem}_{timestamp}.xlsx"
        return OUTPUT_DIR / filename

    def write_results_to_excel(self, output_path: Path, field_results: list[FieldResult]):
        self._qlog("[EXCEL] Opening copied workbook for writing.")
        wb = load_workbook(output_path)

        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(f"Expected sheet '{SHEET_NAME}' not found. Found sheets: {wb.sheetnames}")

        ws = wb[SHEET_NAME]

        self._qlog("[EXCEL] Writing extracted values into 24 yellow input cells.")

        for result in field_results:
            cell = result.excel_cell

            if result.status == "MISSING":
                raise RuntimeError(f"Cannot write missing field: {result.expression} {result.label}")

            if result.should_write_blank:
                ws[cell].value = None
                self._qlog(f"  - {cell}: {result.expression} {result.label} -> BLANK")
            else:
                ws[cell].value = result.numeric_value
                self._qlog(f"  - {cell}: {result.expression} {result.label} -> {result.display_value}")

        self.force_excel_formula_recalculation_on_open(wb)

        wb.save(output_path)
        self._qlog(f"[EXCEL] Saved workbook: {output_path}")

    def force_excel_formula_recalculation_on_open(self, wb):
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"
            self._qlog("[EXCEL] Formula recalculation flags set for next Excel open.")
        except Exception as e:
            self._qlog(f"[WARN] Could not set formula recalculation flags: {e}")

    def recalculate_workbook_with_excel_com(self, output_path: Path):
        self._qlog("[EXCEL COM] Attempting to open Excel and force recalculation.")

        try:
            import win32com.client
        except Exception as e:
            self._qlog(f"[WARN] pywin32/win32com is not available. Skipping Excel COM recalculation. Details: {e}")
            return

        excel = None
        wb = None

        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(str(output_path))
            excel.CalculateFullRebuild()
            wb.Save()

            self._qlog("[EXCEL COM] Workbook recalculated and saved successfully.")

        except Exception as e:
            self._qlog(f"[WARN] Excel COM recalculation failed. Workbook will still recalculate when opened. Details: {e}")

        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=True)
            except Exception:
                pass

            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass

    # ============================================================
    # CLIPBOARD / EXPORT HELPERS
    # ============================================================

    def clean_clipboard_value(self, value):
        if value is None:
            return ""

        value = str(value)
        value = value.replace("\t", " ")
        value = value.replace("\r", " ")
        value = value.replace("\n", " ")
        value = re.sub(r"\s+", " ", value).strip()
        return value

    def rows_to_tsv(self, rows):
        return "\n".join(
            "\t".join(self.clean_clipboard_value(cell) for cell in row)
            for row in rows
        )

    def copy_text_to_clipboard(self, text, label):
        if not text or not text.strip():
            messagebox.showinfo("Nothing to Copy", f"No {label} available to copy yet.")
            return

        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.root.update()

            self.status_text.set(f"Copied {label} to clipboard")
            messagebox.showinfo(
                "Copied",
                f"{label} copied to your clipboard.\n\nNow press Ctrl + V in ChatGPT.",
            )
            self.log(f"[CLIPBOARD] Copied {label} to clipboard.")

        except Exception as e:
            messagebox.showerror("Clipboard Error", str(e))
            self.log(f"[ERROR] Failed to copy {label} to clipboard: {e}")

    def build_field_results_clipboard_text(self):
        rows = [[
            "PDF Code(s)",
            "Field",
            "Excel Cell",
            "Preview Value",
            "Status",
            "Difficulty",
            "Notes",
            "Component Details",
        ]]

        for result in self.latest_field_results:
            rows.append([
                result.expression,
                result.label,
                result.excel_cell,
                result.display_value,
                result.status,
                result.difficulty,
                result.notes,
                result.component_details,
            ])

        return self.rows_to_tsv(rows)

    def build_raw_occurrences_clipboard_text(self):
        rows = [[
            "Code",
            "Selected",
            "Page",
            "X0",
            "Y0",
            "X1",
            "Y1",
            "Nearby Amount",
            "Confidence Score",
            "Note",
            "Nearby Context",
        ]]

        for occ in self.latest_occurrence_rows:
            rows.append([
                occ.code,
                "YES" if occ.selected else "NO",
                occ.page_number,
                round(occ.x0, 2),
                round(occ.y0, 2),
                round(occ.x1, 2),
                round(occ.y1, 2),
                occ.nearby_amount_text,
                occ.confidence_score,
                occ.note,
                occ.nearby_context,
            ])

        return self.rows_to_tsv(rows)

    def build_console_clipboard_text(self):
        return "\n".join(self.latest_report_lines)

    def copy_field_results_to_clipboard(self):
        self.copy_text_to_clipboard(
            self.build_field_results_clipboard_text(),
            "Field Results",
        )

    def copy_raw_occurrences_to_clipboard(self):
        self.copy_text_to_clipboard(
            self.build_raw_occurrences_clipboard_text(),
            "Raw Helper Code Occurrences",
        )

    def copy_console_to_clipboard(self):
        self.copy_text_to_clipboard(
            self.build_console_clipboard_text(),
            "Debug Console",
        )

    def copy_all_results_to_clipboard(self):
        combined = (
            "FIELD RESULTS\n"
            + self.build_field_results_clipboard_text()
            + "\n\nRAW HELPER CODE OCCURRENCES\n"
            + self.build_raw_occurrences_clipboard_text()
            + "\n\nDEBUG CONSOLE\n"
            + self.build_console_clipboard_text()
        )

        self.copy_text_to_clipboard(
            combined,
            "Field Results + Raw Occurrences + Debug Console",
        )

    # ============================================================
    # QUEUE / LOGGING
    # ============================================================

    def _qlog(self, message: str):
        self.queue.put(("log", message))

    def _poll_queue(self):
        try:
            while True:
                event_type, payload = self.queue.get_nowait()

                if event_type == "log":
                    self.log(payload)

                elif event_type == "field_result":
                    self.add_field_result(payload)

                elif event_type == "occurrence":
                    self.add_occurrence(payload)

                elif event_type == "status":
                    self.status_text.set(payload)

                elif event_type == "generated":
                    self.on_excel_generated(payload)

                elif event_type == "done":
                    self.progress.stop()
                    if self.status_text.get().startswith("Running") or self.status_text.get().startswith("Generating"):
                        self.status_text.set("Complete")

        except queue.Empty:
            pass

        self.root.after(100, self._poll_queue)

    def log(self, message: str):
        timestamp = datetime.now().strftime("%H:%M:%S")
        line = f"[{timestamp}] {message}"
        self.latest_report_lines.append(line)
        self.console.insert(END, line + "\n")
        self.console.see(END)

    def add_field_result(self, result: FieldResult):
        self.latest_field_results.append(result)

        tag = "ready"

        if result.status == "VALID BLANK":
            tag = "blank"
        elif result.status == "NEEDS REVIEW":
            tag = "review"
        elif result.status == "MISSING":
            tag = "missing"

        self.field_tree.insert(
            "",
            END,
            values=(
                result.expression,
                result.label,
                result.excel_cell,
                result.display_value,
                result.status,
                result.difficulty,
                result.notes,
                result.component_details,
            ),
            tags=(tag,),
        )

    def add_occurrence(self, occ: CodeOccurrence):
        self.latest_occurrence_rows.append(occ)

        tag = "selected" if occ.selected else "unselected"

        self.occ_tree.insert(
            "",
            END,
            values=(
                occ.code,
                "YES" if occ.selected else "NO",
                occ.page_number,
                round(occ.x0, 2),
                round(occ.y0, 2),
                round(occ.x1, 2),
                round(occ.y1, 2),
                occ.nearby_amount_text,
                occ.confidence_score,
                occ.note,
                occ.nearby_context,
            ),
            tags=(tag,),
        )

    def on_excel_generated(self, output_path: str):
        messagebox.showinfo(
            "Excel Generated",
            f"Completed Excel workbook created:\n\n{output_path}"
        )

        try:
            os.startfile(output_path)
        except Exception as e:
            self.log(f"[WARN] Could not open output workbook automatically: {e}")


def main():
    root = Tk()
    CreditWorksheetApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()